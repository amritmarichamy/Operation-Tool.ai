from flask import Flask, render_template, request, jsonify, send_file, send_from_directory, Response, flash, redirect, url_for, session
from werkzeug.security import generate_password_hash, check_password_hash
import random
import os, re, csv, io, time, ssl, socket, smtplib, imaplib, email, zipfile, shutil, json, math, sqlite3, tempfile
from email.message import EmailMessage
from email.utils import formatdate
from email import policy as email_policy
from datetime import datetime, timezone, timedelta
from queue import Queue, Empty
from threading import Thread, Lock
from collections import defaultdict

from flask_cors import CORS
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import func, text
from sqlalchemy.exc import OperationalError
from sqlalchemy.pool import NullPool
import openpyxl
import pandas as pd
import requests as http_requests
from bs4 import BeautifulSoup
from urllib.parse import urlparse
from werkzeug.utils import secure_filename
from typing import List, Optional, Set, Tuple

# Optional MX check (OFF by default)
try:
    import dns.resolver  # pip install dnspython
    _dns_available = True
except Exception:
    dns = None
    _dns_available = False

APP_NAME = "Terra Tern Email CRM"

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
# ---- Data directories (runtime data lives under data/) ----
DATA_DIR = os.path.join(BASE_DIR, "data")
os.makedirs(DATA_DIR, exist_ok=True)
CRM_SQLITE_PATH = os.path.join(DATA_DIR, "crm.sqlite3")
UPLOAD_ROOT = os.path.join(DATA_DIR, "uploads")
REPORTS_DIR = os.path.join(DATA_DIR, "reports")
os.makedirs(UPLOAD_ROOT, exist_ok=True)
os.makedirs(REPORTS_DIR, exist_ok=True)

# CRM export column "Enrollment ID" = full JSA URL when candidate has a plain id (override via env).
CRM_JSA_ENROLLMENT_UPDATE_PREFIX = os.environ.get(
    "CRM_JSA_ENROLLMENT_UPDATE_PREFIX",
    "https://backend.terratern.com/jsa-enrollment/update?id=",
).strip()

# Scanner output folder
SCANNER_UPLOAD_FOLDER = os.path.join(DATA_DIR, "scanner_uploads")
SCANNER_OUTPUT_FOLDER = os.path.join(DATA_DIR, "scanner_outputs")
os.makedirs(SCANNER_UPLOAD_FOLDER, exist_ok=True)
os.makedirs(SCANNER_OUTPUT_FOLDER, exist_ok=True)

# ---- Frontend directories ----
FRONTEND_DIR = os.path.join(BASE_DIR, "frontend")
app = Flask(__name__,
            static_folder=os.path.join(FRONTEND_DIR, "static"),
            template_folder=os.path.join(FRONTEND_DIR, "templates"))
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.secret_key = os.getenv("SECRET_KEY", "terra-tern-crm-secret-key-2026")
CORS(app)

# Email settings (Gmail)
SMTP_HOST = "smtp.gmail.com"
SMTP_PORT = 587
IMAP_HOST = "imap.gmail.com"

# OTP SMTP settings
OTP_EMAIL = "amrit.marichamy@terratern.com"
OTP_PASS = "qrop gcxm xrpz jmpu"

# Retry settings
MAX_RETRIES = 5
RETRY_BASE_SECONDS = 3
NETWORK_WAIT_SECONDS = 8

# Validations
ENABLE_MX_CHECK = False            # set True if you also install dnspython and want MX check
STRICT_BLOCK_INVALID_TARGETS = True # block sending if any invalid targets exist

# IMAP bounce scan after sends: slower but removes bounces from CSV reports and reconciles Sent vs Bounced in Analytics.
ENABLE_BOUNCE_CHECK_DEFAULT = True

# Per-candidate daily ceiling (UTC day): backlog is spread over multiple days (e.g. 300 behind → 100×3).
# Also caps each automation run so clients are not flooded in one batch.
SMART_AUTOMATION_BACKLOG_SAFE_CAP = 100
# When the candidate is **not** behind the day curve in any automation bucket, allow more headroom
# (manual / on-track sends); scheduler still sends 0 batch when there is no backlog.
AUTOMATION_DAILY_CAP_WHEN_NO_BACKLOG = 500
# Seconds between each email in automated / scheduled runs (Gmail-friendly pacing)
AUTOMATION_DEFAULT_DELAY_SECONDS = 10

# DB
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///" + CRM_SQLITE_PATH
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
# NOTE:
# Many browser tabs + background worker threads can create a lot of concurrent DB usage.
# SQLite + the default QueuePool can hit pool timeouts (QueuePool size 5 / overflow 10).
# Using NullPool avoids pool exhaustion by opening/closing a DB connection per request.
app.config["SQLALCHEMY_ENGINE_OPTIONS"] = {
    "poolclass": NullPool,
    "connect_args": {"check_same_thread": False, "timeout": 30},
    "pool_pre_ping": True,
}
db = SQLAlchemy(app)


@app.teardown_appcontext
def _cleanup_session(exc=None):
    # Ensure connections are always returned/closed even if a request errors.
    try:
        db.session.remove()
    except Exception:
        pass

DISPOSABLE_DOMAINS = {
    "mailinator.com", "guerrillamail.com", "10minutemail.com", "tempmail.com",
    "yopmail.com", "trashmail.com", "sharklasers.com"
}

EMAIL_RE = re.compile(
    r"^[A-Za-z0-9.!#$%&'*+/=?^_`{|}~-]+@"
    r"[A-Za-z0-9](?:[A-Za-z0-9-]{0,61}[A-Za-z0-9])?"
    r"(?:\.[A-Za-z0-9](?:[A-Za-z0-9-]{0,61}[A-Za-z0-9])?)+$"
)

# Runtime state
run_states = {}        # run_id -> {"stop": bool, "pause": bool}
progress_queues = {}   # run_id -> list of Queue() (fan-out: one Queue per SSE consumer)
progress_lock = Lock() # Lock for thread-safe access to progress_queues
is_bulk_sync_running = False # Global flag for bulk sync
last_bulk_sync_status = ""   # Summary of last bulk sync

def push_progress(run_id: str, msg: str):
    """Broadcast a progress message to ALL SSE consumers listening to this run."""
    run_id_str = str(run_id)
    with progress_lock:
        consumers = progress_queues.get(run_id_str, [])
        for q in consumers:
            try:
                q.put(str(msg))
            except Exception:
                pass


def _bump_send_run_counter(run_id: int, col: str, delta: int = 1):
    """Atomic UPDATE for live Runs & Reports stats (worker threads + SQLite)."""
    if delta <= 0 or col not in ("sent", "failed", "skipped", "bounced"):
        return
    try:
        with app.app_context():
            with db.engine.begin() as conn:
                conn.execute(
                    text(f"UPDATE send_runs SET {col} = COALESCE({col}, 0) + :d WHERE id = :rid"),
                    {"d": int(delta), "rid": int(run_id)},
                )
    except Exception:
        pass


def _set_run_status(run_id: str, status: str):
    """Safely update a run status using a short-lived app context.

    We keep this helper very defensive so it never crashes the worker.
    """
    try:
        rid = int(str(run_id))
    except Exception:
        return
    try:
        with app.app_context():
            run = db.session.get(SendRun, rid)
            if not run:
                return
            # Don't overwrite terminal states.
            if (run.status or "").lower() in ("done", "failed", "stopped", "deleted"):
                return
            if run.status != status:
                run.status = status
                db.session.add(run)
                db.session.commit()
            db.session.remove()
    except Exception:
        try:
            db.session.rollback()
        except Exception:
            pass


def _wait_for_network(run_id: str):
    """Block until SMTP host is reachable. Updates run status as paused_network."""
    run_id_str = str(run_id)
    # Mark run as paused because network is down.
    _set_run_status(run_id_str, "paused_network")
    push_progress(run_id_str, "⏸️ Network down. Run paused and will auto-resume when internet returns.")

    # Wait until SMTP port is reachable again.
    last_ping = 0.0
    while not is_network_up():
        st = run_states.get(run_id_str, {"stop": False, "pause": False})
        if st.get("stop"):
            return False
        # Respect user pause too; still just sleep.
        now = time.time()
        if now - last_ping > 30:
            push_progress(run_id_str, "…still waiting for network")
            last_ping = now
        time.sleep(2.0)

    # Network is back.
    _set_run_status(run_id_str, "running")
    push_progress(run_id_str, "▶️ Network is back. Resuming from where it stopped.")
    return True

def is_network_up(host=SMTP_HOST, port=SMTP_PORT, timeout=5):
    try:
        with socket.create_connection((host, port), timeout=timeout):
            return True
    except Exception:
        return False

def looks_like_email(addr: str) -> bool:
    if not addr or "@" not in addr:
        return False
    return bool(EMAIL_RE.match(addr.strip()))

def has_mx_records(domain: str) -> bool:
    if not ENABLE_MX_CHECK or not _dns_available:
        return True
    try:
        answers = dns.resolver.resolve(domain, "MX")
        return any(answers)
    except Exception:
        return False

def domain_is_allowed(domain: str) -> bool:
    return domain.lower() not in DISPOSABLE_DOMAINS

def is_valid_address_for_send(addr: str):
    addr = (addr or "").strip().lower()
    if not looks_like_email(addr):
        return (False, "invalid_syntax")
    try:
        _, domain = addr.split("@", 1)
    except ValueError:
        return (False, "invalid_syntax")
    if not domain_is_allowed(domain):
        return (False, "disposable_domain")
    if not has_mx_records(domain):
        return (False, "no_mx")
    return (True, "")

def safe_candidate_folder(candidate_id: int):
    folder = os.path.join(UPLOAD_ROOT, f"candidate_{candidate_id}")
    os.makedirs(folder, exist_ok=True)
    return folder


def upload_path_for_api(abs_path: Optional[str]) -> str:
    """Path relative to project root for API/UI (forward slashes). Empty if no file."""
    if not abs_path or not str(abs_path).strip():
        return ""
    raw = str(abs_path).strip()
    # Resolve project-relative paths against BASE_DIR (abspath() alone uses process cwd).
    if not os.path.isabs(raw):
        raw = os.path.normpath(os.path.join(BASE_DIR, raw.replace("/", os.sep)))
    try:
        ap = os.path.normpath(os.path.abspath(raw))
        bd = os.path.normpath(BASE_DIR)
        ap_n = os.path.normcase(ap)
        bd_n = os.path.normcase(bd)
        if ap_n == bd_n or ap_n.startswith(bd_n + os.sep):
            rel = os.path.relpath(ap, bd)
            out = rel.replace("\\", "/")
            if out and out != ".":
                return out
    except (ValueError, OSError, Exception):
        pass
    s = raw.replace("\\", "/")
    low = s.lower()
    if "uploads/" in low:
        i = low.index("uploads/")
        return s[i:].replace("\\", "/")
    try:
        return os.path.basename(s) or s
    except Exception:
        return s


def save_uploaded_file(file_storage, folder, name_hint):
    if not file_storage:
        return None
    filename = file_storage.filename or name_hint
    filename = re.sub(r"[^A-Za-z0-9._-]+", "_", filename)
    path = os.path.join(folder, filename)
    file_storage.save(path)
    return path

def parse_roles(text: str):
    raw = (text or "").replace(";", "\n").replace("|", "\n")
    out = []
    for line in raw.splitlines():
        line = line.strip()
        if not line:
            continue
        parts = [p.strip() for p in line.split(",")] if "," in line else [line]
        for p in parts:
            if p and p not in out:
                out.append(p)
    return out

def send_with_retries(sender_email, sender_password, msg, run_id):
    attempt = 0
    network_paused = False
    while attempt <= MAX_RETRIES:
        state = run_states.get(str(run_id), {"stop": False, "pause": False})
        if state.get("stop"):
            return ("failed", "Stopped by user")
        while state.get("pause"):
            time.sleep(0.6)
            state = run_states.get(str(run_id), state)

#         # --- Network cut handling (auto-pause + auto-resume) ---
        # If the network is down, we DO NOT count this as a failure.
        # We keep the run alive, mark it as paused_network, and resume
        # automatically once the network is back.
        while not is_network_up():
            if not network_paused:
                network_paused = True
                _set_run_status(str(run_id), "paused_network")
                push_progress(run_id, "⏸️ Network down. Run paused automatically. Waiting for internet to come back…")

            state = run_states.get(str(run_id), {"stop": False, "pause": False})
            if state.get("stop"):
                return ("failed", "Stopped by user")
            # If user manually pauses, honor it too.
            while state.get("pause"):
                time.sleep(0.6)
                state = run_states.get(str(run_id), state)

            time.sleep(NETWORK_WAIT_SECONDS)

        if network_paused:
            network_paused = False
            _set_run_status(str(run_id), "running")
            push_progress(run_id, "▶️ Network restored. Resuming email sending…")

        try:
            with smtplib.SMTP(SMTP_HOST, SMTP_PORT, timeout=60) as server:
                server.starttls(context=ssl.create_default_context())
                server.login(sender_email, sender_password)
                mail_opts = ["SMTPUTF8", "DSN"]
                rcpt_opts = ["NOTIFY=FAILURE,DELAY"]
                refused = server.send_message(
                    msg, mail_options=mail_opts, rcpt_options=rcpt_opts
                )
            if refused:
                return ("bounced", f"SMTP rejected recipients: {refused}")
            # SMTP accepted the message; mirror into Gmail Sent so the Sent folder matches reality
            if not append_gmail_sent_folder(sender_email, sender_password, msg):
                push_progress(
                    run_id,
                    "ℹ️ Email accepted by Gmail SMTP. If it does not appear under Sent, enable IMAP "
                    "for this account (Google settings) and ensure the app password works for IMAP.",
                )
            return ("sent", None)

        except smtplib.SMTPRecipientsRefused as e:
            return ("bounced", f"SMTPRecipientsRefused: {e}")
        except smtplib.SMTPAuthenticationError as e:
            return ("failed", f"SMTPAuthenticationError: {e}")
        except smtplib.SMTPException as e:
            attempt += 1
            if attempt > MAX_RETRIES:
                return ("failed", f"SMTPException: {e}")
            backoff = RETRY_BASE_SECONDS * (2 ** (attempt - 1))
            push_progress(run_id, f"ℹ️ SMTP error; retrying in {backoff}s…")
            time.sleep(backoff)
        except (socket.timeout, socket.gaierror, ConnectionError, ssl.SSLError) as e:
            # Treat as transient network issue; if network is actually down,
            # the loop at the top will pause and resume automatically.
            attempt += 1
            if attempt > MAX_RETRIES:
                # One last check: if network looks down now, wait instead of failing.
                if not is_network_up():
                    attempt = 0
                    continue
                return ("failed", f"NetworkError: {e}")
            backoff = RETRY_BASE_SECONDS * (2 ** (attempt - 1))
            push_progress(run_id, f"ℹ️ Network error; retrying in {backoff}s…")
            time.sleep(backoff)
        except Exception as e:
            return ("failed", f"UnexpectedError: {e}")


def append_gmail_sent_folder(sender_email: str, sender_password: str, msg: EmailMessage) -> bool:
    """
    Gmail often does not mirror SMTP-submitted mail into the Sent folder in the web UI.
    Append the same message via IMAP so Sent matches delivery (same app password as SMTP).
    """
    try:
        raw = msg.as_bytes(policy=email_policy.SMTP)
    except Exception:
        try:
            raw = msg.as_bytes()
        except Exception as ex:
            print(f"[Email] Sent-folder copy: cannot serialize message: {ex}")
            return False
    try:
        imap = imaplib.IMAP4_SSL(IMAP_HOST, timeout=60)
        imap.login(sender_email, sender_password)
        for folder in (
            "[Gmail]/Sent Mail",
            "[Gmail]/Sent",
            "Sent",
        ):
            try:
                typ, _ = imap.append(folder, "\\Seen", None, raw)
                if typ == "OK":
                    imap.logout()
                    return True
            except imaplib.IMAP4.error:
                continue
        try:
            imap.logout()
        except Exception:
            pass
    except Exception as ex:
        print(f"[Email] Sent-folder IMAP append failed: {ex}")
    return False


def fetch_bounces_gmail(sender_email, sender_password, since_dt, hr_sent_set):
    bounced_to = set()
    try:
        imap = imaplib.IMAP4_SSL(IMAP_HOST)
        imap.login(sender_email, sender_password)
        imap.select("INBOX")
        since = (since_dt - timedelta(days=1)).strftime("%d-%b-%Y")
        criteria = f'(SINCE "{since}" OR FROM "MAILER-DAEMON" SUBJECT "Undelivered Mail Returned to Sender" SUBJECT "Delivery Status Notification" SUBJECT "failure")'
        typ, data = imap.search(None, criteria)
        if typ != "OK":
            imap.logout()
            return bounced_to

        for num in data[0].split():
            typ, msgdata = imap.fetch(num, "(RFC822)")
            if typ != "OK":
                continue
            msg = email.message_from_bytes(msgdata[0][1])

            failed_candidates = set()
            for h in msg.get_all("X-Failed-Recipients", []):
                for addr in h.split(","):
                    a = addr.strip().lower()
                    if a:
                        failed_candidates.add(a)

            for part in msg.walk():
                ctype = (part.get_content_type() or "").lower()
                if ctype == "message/delivery-status":
                    try:
                        payload = part.get_payload()
                        if isinstance(payload, list):
                            for block in payload:
                                final_rcpt = block.get("Final-Recipient")
                                if final_rcpt and ";" in final_rcpt:
                                    a = final_rcpt.split(";", 1)[1].strip().lower()
                                    if a:
                                        failed_candidates.add(a)
                    except Exception:
                        pass

                if ctype == "text/plain":
                    try:
                        textb = part.get_payload(decode=True) or b""
                    except Exception:
                        textb = b""
                    textt = (textb or b"").decode("utf-8", "ignore")
                    for token in textt.replace("<"," ").replace(">"," ").split():
                        if "@" in token and "." in token and len(token) <= 254:
                            cand = token.strip().strip(",.;:()[]{}<>").lower()
                            if "@" in cand and cand.count("@") == 1:
                                failed_candidates.add(cand)

            for a in failed_candidates:
                if a in hr_sent_set:
                    bounced_to.add(a)

        imap.logout()
    except Exception:
        pass
    return bounced_to

def _dt_naive_for_delta(dt: Optional[datetime]) -> Optional[datetime]:
    """Normalize to naive datetime for day-delta math (matches existing app storage)."""
    if dt is None:
        return None
    if getattr(dt, "tzinfo", None):
        return dt.replace(tzinfo=None)
    return dt


def candidate_service_anchor_dt(
    c: "Candidate",
    workflow_service_start: Optional[datetime] = None,
    workspace_service_start: Optional[datetime] = None,
) -> Optional[datetime]:
    """Calendar anchor for 'days in system': profile start, then 6-month plan, then workspace service start."""
    if c.smart_service_start_date:
        return _dt_naive_for_delta(c.smart_service_start_date)
    if workflow_service_start:
        return _dt_naive_for_delta(workflow_service_start)
    if workspace_service_start:
        return _dt_naive_for_delta(workspace_service_start)
    return None


def candidate_days_for_bucklist(
    c: "Candidate",
    workflow_service_start: Optional[datetime] = None,
    workspace_service_start: Optional[datetime] = None,
) -> Optional[int]:
    """Live days from service anchor when present; otherwise pinned bucklist days or age since created."""
    anchor = candidate_service_anchor_dt(c, workflow_service_start, workspace_service_start)
    if anchor is not None:
        return max(0, (datetime.utcnow() - anchor).days)
    if c.bucklist_days_in_system is not None:
        return max(0, int(c.bucklist_days_in_system))
    if c.created_at:
        start = _dt_naive_for_delta(c.created_at)
        return max(0, (datetime.utcnow() - start).days)
    return None


def candidate_days_in_system_source(
    c: "Candidate",
    workflow_service_start: Optional[datetime] = None,
    workspace_service_start: Optional[datetime] = None,
) -> Optional[str]:
    anchor = candidate_service_anchor_dt(c, workflow_service_start, workspace_service_start)
    if anchor is not None:
        if c.smart_service_start_date:
            return "profile_service_start"
        if workflow_service_start:
            return "workflow_plan"
        if workspace_service_start:
            return "workspace_service"
        return "elapsed"
    if c.bucklist_days_in_system is not None:
        return "bucklist_pin"
    if c.created_at:
        return "created_age"
    return None


def bucket_key_for_days(days: int) -> str:
    d = max(0, int(days))
    if d <= 30:
        return "0-30"
    if d <= 60:
        return "31-60"
    if d <= 90:
        return "61-90"
    if d <= 120:
        return "91-120"
    if d <= 150:
        return "121-150"
    if d <= 180:
        return "151-180"
    return ">180"


BUCKLIST_BUCKET_KEYS = ("0-30", "31-60", "61-90", "91-120", "121-150", "151-180", ">180")

ENROLLMENT_STATUSES = frozenset({"Ongoing", "On Hold", "Completed"})


def normalize_enrollment_status(raw: Optional[str]) -> str:
    s = (raw or "").strip()
    return s if s in ENROLLMENT_STATUSES else "Ongoing"

# Inclusive bounds for validating days_in_system against a selected Bucklist category
BUCKLIST_CATEGORY_BOUNDS = {
    "0-30": (0, 30),
    "31-60": (31, 60),
    "61-90": (61, 90),
    "91-120": (91, 120),
    "121-150": (121, 150),
    "151-180": (151, 180),
    ">180": (181, 999999),
}

# Background scheduler only — bands 91+ keep automated sends; 0–90 days defer (no send, schedule bumped).
SCHEDULER_AUTOMATION_DISABLED_BUCKLIST_BUCKETS = frozenset({"0-30", "31-60", "61-90"})


def candidate_bucklist_bucket_key_for_scheduler(cand: "Candidate", now: Optional[datetime] = None) -> Optional[str]:
    """Service-day band for automation gating.

    When a 6-month workflow plan exists, use ``workflow_plan_elapsed_service_days`` — same as the
    dashboard **Day** / phase curve. Otherwise fall back to bucklist bucket rules like ``/api/bucklist``.
    """
    now = now or datetime.utcnow()
    wp = (
        WorkflowPlan.query.filter_by(candidate_id=cand.id)
        .filter(WorkflowPlan.status.in_(["active", "paused"]))
        .order_by(WorkflowPlan.id.desc())
        .first()
    )
    if wp and getattr(wp, "service_start_date", None):
        elapsed = workflow_plan_elapsed_service_days(wp, now)
        return bucket_key_for_days(int(elapsed))
    plan_starts, ws_starts = candidate_service_start_maps_for_ids([cand.id])
    wf_s, ws_s = plan_starts.get(cand.id), ws_starts.get(cand.id)
    try:
        days = candidate_days_for_bucklist(cand, workflow_service_start=wf_s, workspace_service_start=ws_s)
    except Exception:
        days = None
    if days is None:
        return None
    return bucket_key_for_days(int(days))


def scheduler_should_pause_automation_for_candidate_bucklist(
    cand: "Candidate", now: Optional[datetime] = None
) -> bool:
    key = candidate_bucklist_bucket_key_for_scheduler(cand, now)
    if key is None:
        return False
    return key in SCHEDULER_AUTOMATION_DISABLED_BUCKLIST_BUCKETS


def format_candidate_country_type(smart_country: Optional[str], workspace_country: Optional[str] = None) -> str:
    """Display label: country as stored (e.g. India, UAE). Legacy rows may have 'Work X' — strip the prefix."""
    raw = (smart_country or "").strip() or (workspace_country or "").strip()
    if not raw:
        return ""
    low = raw.lower()
    if low.startswith("work "):
        return raw[5:].strip()
    return raw


# Models
class Candidate(db.Model):
    __tablename__ = "candidates"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(200), nullable=False)
    email = db.Column(db.String(200), nullable=False)
# NOTE: legacy column kept for backward compatibility.
    # UI now uses `pa_member` instead of `app_code`.
    app_code = db.Column(db.String(100), nullable=True)
    pa_member = db.Column(db.String(200), nullable=True)
    rm_member = db.Column(db.String(200), nullable=True)
    placement_officer_member = db.Column(db.String(200), nullable=True)
    app_password = db.Column(db.String(200), nullable=True)
    subject_template = db.Column(db.Text, nullable=True)
    message_template = db.Column(db.Text, nullable=True)
    roles_text = db.Column(db.Text, nullable=True)
    resume_path = db.Column(db.String(500), nullable=True)
    cover_letter_path = db.Column(db.String(500), nullable=True)
    enrollment_id = db.Column(db.String(100), nullable=True)
    enrollment_status = db.Column(db.String(40), nullable=True)
    industry_types = db.Column(db.Text, nullable=True)
    scheduled_time = db.Column(db.DateTime, nullable=True)
    # One-time Smart Automation service profile (days + HR scope derived from these)
    smart_service_start_date = db.Column(db.DateTime, nullable=True)
    # Snapshot of "total applications completed till date" (manual + CRM); updated when user saves profile / enables automation
    smart_baseline_applied = db.Column(db.Integer, default=0)
    smart_country = db.Column(db.String(100), nullable=True)
    smart_industry = db.Column(db.String(150), nullable=True)
    # Bucklist: pinned day count when there is no service-start anchor (profile / 6-month plan / workspace start)
    bucklist_days_in_system = db.Column(db.Integer, nullable=True)
    # Last time a **scheduler-driven** automation run was started (workspace / workflow / scheduled_time). UTC.
    scheduler_automation_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=func.now())
    updated_at = db.Column(db.DateTime, nullable=False, default=func.now(), onupdate=func.now())

    def to_dict_summary(
        self,
        workspace_country: Optional[str] = None,
        workflow_service_start: Optional[datetime] = None,
        workspace_service_start: Optional[datetime] = None,
    ):
        # prefer new columns; fall back to legacy app_code
        pa = (self.pa_member or self.app_code or "")
        # Candidate "Country Type" is profile-only; do not mix in workspace/campaign country (avoids misleading "Work UAE").
        country_type = format_candidate_country_type(self.smart_country, None)
        days_comp = candidate_days_for_bucklist(
            self,
            workflow_service_start=workflow_service_start,
            workspace_service_start=workspace_service_start,
        )
        days_src = candidate_days_in_system_source(
            self,
            workflow_service_start=workflow_service_start,
            workspace_service_start=workspace_service_start,
        )
        return {
            "id": self.id,
            "name": self.name,
            "email": self.email,
            "pa_member": pa,
            "rm_member": (self.rm_member or "").strip(),
            "placement_officer_member": self.placement_officer_member or "",
            "enrollment_id": self.enrollment_id or "",
            "enrollment_status": normalize_enrollment_status(getattr(self, "enrollment_status", None)),
            "industry_types": (self.industry_types or "").strip(),
            "smart_service_start_date": self.smart_service_start_date.isoformat() if self.smart_service_start_date else "",
            "smart_baseline_applied": int(self.smart_baseline_applied or 0),
            "smart_country": (self.smart_country or "").strip(),
            "smart_industry": (self.smart_industry or "").strip(),
            "bucklist_days_in_system": int(self.bucklist_days_in_system) if self.bucklist_days_in_system is not None else None,
            "days_in_system_computed": days_comp,
            "days_in_system_source": days_src,
            "country_type": country_type,
            "workspace_country": (workspace_country or "").strip(),
            "has_app_password": bool(self.app_password),
            "has_resume": bool(self.resume_path),
            "has_cover_letter": bool(self.cover_letter_path),
            "resume_path": upload_path_for_api(self.resume_path),
            "cover_letter_path": upload_path_for_api(self.cover_letter_path),
            "scheduled_time": self.scheduled_time.isoformat() if self.scheduled_time else "",
            "updated_at": self.updated_at.isoformat() if self.updated_at else "",
            "created_at": self.created_at.isoformat() if self.created_at else "",
        }

    def to_dict_detail(self):
        pa = (self.pa_member or self.app_code or "")
        wf_s, ws_s = candidate_aux_service_starts(self.id)
        days_comp = candidate_days_for_bucklist(
            self, workflow_service_start=wf_s, workspace_service_start=ws_s
        )
        days_src = candidate_days_in_system_source(
            self, workflow_service_start=wf_s, workspace_service_start=ws_s
        )
        return {
            "id": self.id,
            "name": self.name,
            "email": self.email,
            "pa_member": pa,
            "rm_member": (self.rm_member or "").strip(),
            "placement_officer_member": self.placement_officer_member or "",
            "enrollment_id": self.enrollment_id or "",
            "enrollment_status": normalize_enrollment_status(getattr(self, "enrollment_status", None)),
            "app_password": self.app_password or "",
            "subject_template": self.subject_template or "",
            "message_template": self.message_template or "",
            "roles_text": self.roles_text or "",
            "industry_types": (self.industry_types or "").strip(),
            "smart_service_start_date": self.smart_service_start_date.isoformat() if self.smart_service_start_date else "",
            "smart_baseline_applied": int(self.smart_baseline_applied or 0),
            "smart_country": (self.smart_country or "").strip(),
            "smart_industry": (self.smart_industry or "").strip(),
            "scheduled_time": self.scheduled_time.isoformat() if self.scheduled_time else "",
            "bucklist_days_in_system": int(self.bucklist_days_in_system) if self.bucklist_days_in_system is not None else None,
            "days_in_system_computed": days_comp,
            "days_in_system_source": days_src,
            "resume_on_file": bool(self.resume_path),
            "cover_on_file": bool(self.cover_letter_path),
            "resume_path": upload_path_for_api(self.resume_path),
            "cover_letter_path": upload_path_for_api(self.cover_letter_path),
            "updated_at": self.updated_at.isoformat() if self.updated_at else "",
            "created_at": self.created_at.isoformat() if self.created_at else "",
        }

class Target(db.Model):
    __tablename__ = "targets"
    id = db.Column(db.Integer, primary_key=True)
    company_name = db.Column(db.String(250), nullable=False)
    hr_email = db.Column(db.String(250), nullable=False)
    country = db.Column(db.String(100), nullable=True)
    hr_name = db.Column(db.String(200), nullable=True)
    target_role = db.Column(db.String(250), nullable=True)
    industry = db.Column(db.String(150), nullable=False, default="Default")
    is_valid = db.Column(db.Boolean, nullable=False, default=True)
    invalid_reason = db.Column(db.String(80), nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=func.now())

    def to_dict(self):
        return {
            "id": self.id,
            "company_name": self.company_name,
            "hr_email": self.hr_email,
            "country": self.country or "",
            "hr_name": self.hr_name or "",
            "target_role": self.target_role or "",
            "industry": self.industry or "Default",
            "is_valid": bool(self.is_valid),
            "invalid_reason": self.invalid_reason or ""
        }

class Industry(db.Model):
    __tablename__ = "industries"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), nullable=False)
    country = db.Column(db.String(100), nullable=False, default="Global")
    created_at = db.Column(db.DateTime, nullable=False, default=func.now())
    __table_args__ = (db.UniqueConstraint('name', 'country', name='_industry_country_uc'),)

class SentHistory(db.Model):
    __tablename__ = "sent_history"
    id = db.Column(db.Integer, primary_key=True)
    candidate_id = db.Column(db.Integer, db.ForeignKey('candidates.id'), nullable=False)
    target_email = db.Column(db.String(250), nullable=False)
    sent_at = db.Column(db.DateTime, nullable=False, default=func.now())
    
    # Composite unique constraint to prevent duplicates per candidate
    __table_args__ = (db.UniqueConstraint('candidate_id', 'target_email', name='_candidate_target_uc'),)


def utc_calendar_start_naive(now: Optional[datetime] = None) -> datetime:
    """UTC midnight (naive), same convention as SentHistory.sent_at / server clock."""
    now = now or datetime.utcnow()
    return datetime(now.year, now.month, now.day)


def emails_sent_today_for_candidate(candidate_id: int, now: Optional[datetime] = None) -> int:
    """Emails already recorded for this candidate since UTC midnight (escalation / daily pacing)."""
    if not candidate_id:
        return 0
    start = utc_calendar_start_naive(now)
    return int(SentHistory.query.filter(SentHistory.candidate_id == candidate_id, SentHistory.sent_at >= start).count())


def candidate_has_automation_backlog_any_bucket(candidate_id: int, now: Optional[datetime] = None) -> bool:
    """True if the candidate is behind the official day curve in any active workflow or enabled workspace."""
    now = now or datetime.utcnow()
    if not candidate_id:
        return False
    wp = (
        WorkflowPlan.query.filter_by(candidate_id=candidate_id)
        .filter(WorkflowPlan.status.in_(["paused", "active"]))
        .order_by(WorkflowPlan.id.desc())
        .first()
    )
    if wp:
        elapsed = workflow_plan_elapsed_service_days(wp, now)
        expected = linear_expected_applications_for_days(elapsed)
        applied = int(wp.total_applied or 0)
        if max(0, expected - applied) > 0:
            return True
    cand = db.session.get(Candidate, candidate_id)
    if not cand:
        return False
    for ws in Workspace.query.filter_by(candidate_id=candidate_id, automation_enabled=True).all():
        wpp = wp
        svc = getattr(cand, "smart_service_start_date", None) or ws.service_start_date
        if not svc:
            svc = now
        if wpp:
            days_elapsed = workflow_plan_elapsed_service_days(wpp, now)
            applied_w = int(wpp.total_applied or 0)
        else:
            days_elapsed = max(0, (now - svc).days)
            applied_w = int(ws.automation_total_sent or 0)
        expected = linear_expected_applications_for_days(days_elapsed)
        if max(0, expected - applied_w) > 0:
            return True
    return False


def scheduler_automation_already_used_today(candidate_id: int, now: Optional[datetime] = None) -> bool:
    """Whether a scheduler-driven automation run already **started** for this candidate since UTC midnight."""
    now = now or datetime.utcnow()
    if not candidate_id:
        return False
    c = db.session.get(Candidate, candidate_id)
    ts = getattr(c, "scheduler_automation_at", None) if c else None
    if not ts:
        return False
    start = utc_calendar_start_naive(now)
    return ts >= start


def remaining_daily_send_quota(candidate_id: Optional[int], now: Optional[datetime] = None) -> int:
    """Remaining sends today (UTC). 100/day while any bucket has curve backlog; higher cap when fully on track."""
    if not candidate_id:
        return int(SMART_AUTOMATION_BACKLOG_SAFE_CAP)
    used = emails_sent_today_for_candidate(candidate_id, now)
    if candidate_has_automation_backlog_any_bucket(candidate_id, now):
        daily_cap = int(SMART_AUTOMATION_BACKLOG_SAFE_CAP)
    else:
        daily_cap = max(int(SMART_AUTOMATION_BACKLOG_SAFE_CAP), int(AUTOMATION_DAILY_CAP_WHEN_NO_BACKLOG))
    return max(0, daily_cap - used)


class User(db.Model):
    __tablename__ = "users"
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    otp_code = db.Column(db.String(10), nullable=True)
    otp_expiry = db.Column(db.DateTime, nullable=True)
    is_verified = db.Column(db.Boolean, default=False)
    is_approved = db.Column(db.Boolean, default=False)
    role = db.Column(db.String(20), default="user")
    # JSON text: list of sidebar tab IDs this user can see. Null = all (admin/manager default).
    allowed_features = db.Column(db.Text, nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=func.now())

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

class Workspace(db.Model):
    __tablename__ = "workspaces"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(250), nullable=False)
    candidate_id = db.Column(db.Integer, db.ForeignKey('candidates.id'), nullable=False)
    industry = db.Column(db.String(150), nullable=False, default="Default")
    country = db.Column(db.String(100), nullable=False, default="Global")
    
    # Automation settings (moved from Candidate)
    automation_enabled = db.Column(db.Boolean, default=False)
    automation_batch_size = db.Column(db.Integer, default=10)
    automation_interval_days = db.Column(db.Integer, default=2)
    automation_next_run = db.Column(db.DateTime, nullable=True)
    automation_target_index = db.Column(db.Integer, default=0)
    automation_max_emails = db.Column(db.Integer, default=1000)
    automation_total_sent = db.Column(db.Integer, default=0)
    
    # New Service Plan fields
    automation_type = db.Column(db.String(20), default="interval") # 'interval' or 'schedule'
    scheduled_days = db.Column(db.String(200), default="Tuesday,Wednesday,Thursday,Friday")
    service_start_date = db.Column(db.DateTime, nullable=True)
    service_end_date = db.Column(db.DateTime, nullable=True)
    monthly_target = db.Column(db.Integer, default=200)
    monthly_sent_count = db.Column(db.Integer, default=0)
    last_month_reset = db.Column(db.DateTime, nullable=True)
    automation_per_run_cap = db.Column(db.Integer, default=100)
    
    created_at = db.Column(db.DateTime, nullable=False, default=func.now())
    updated_at = db.Column(db.DateTime, nullable=False, default=func.now(), onupdate=func.now())

    def to_dict(self):
        cand = db.session.get(Candidate, self.candidate_id)
        cand_name = cand.name if cand else "Unknown Candidate"
            
        return {
            "id": self.id,
            "name": self.name,
            "candidate_id": self.candidate_id,
            "candidate_name": cand_name,
            "industry": self.industry or "Default",
            "country": self.country or "Global",
            "automation_enabled": bool(self.automation_enabled),
            "automation_batch_size": self.automation_batch_size or 10,
            "automation_interval_days": self.automation_interval_days or 2,
            "automation_next_run": self.automation_next_run.isoformat() if self.automation_next_run else "",
            "automation_target_index": self.automation_target_index or 0,
            "automation_max_emails": self.automation_max_emails or 1000,
            "automation_total_sent": self.automation_total_sent or 0,
            "automation_type": self.automation_type or "interval",
            "scheduled_days": self.scheduled_days or "Tuesday,Wednesday,Thursday,Friday",
            "service_start_date": self.service_start_date.isoformat() if self.service_start_date else "",
            "service_end_date": self.service_end_date.isoformat() if self.service_end_date else "",
            "monthly_target": self.monthly_target or 200,
            "monthly_sent_count": self.monthly_sent_count or 0,
            "automation_per_run_cap": self.automation_per_run_cap or 100,
            "created_at": self.created_at.isoformat() if self.created_at else "",
        }

class SendRun(db.Model):
    __tablename__ = "send_runs"
    id = db.Column(db.Integer, primary_key=True)
    mode = db.Column(db.String(20), nullable=False)
    status = db.Column(db.String(20), nullable=False, default="queued")
    created_at = db.Column(db.DateTime, nullable=False, default=func.now())
    started_at = db.Column(db.DateTime, nullable=True)
    ended_at = db.Column(db.DateTime, nullable=True)
    candidate_names = db.Column(db.Text, nullable=True)
    pa_members = db.Column(db.Text, nullable=True)
    placement_officer_members = db.Column(db.Text, nullable=True)
    total_targets = db.Column(db.Integer, nullable=False, default=0)
    sent = db.Column(db.Integer, nullable=False, default=0)
    failed = db.Column(db.Integer, nullable=False, default=0)
    bounced = db.Column(db.Integer, nullable=False, default=0)
    skipped = db.Column(db.Integer, nullable=False, default=0)
    is_deleted = db.Column(db.Boolean, nullable=False, default=False)
    industry = db.Column(db.String(150), nullable=True, default="Default")
    country = db.Column(db.String(100), nullable=True, default="Global")
    user_id = db.Column(db.Integer, nullable=True)  # Track which user started this run

    def to_dict(self):
        return {
            "id": self.id,
            "mode": self.mode,
            "status": self.status,
            "created_at": self.created_at.isoformat() if self.created_at else "",
            "candidate_names": self.candidate_names or "",
            "pa_members": self.pa_members or "",
            "placement_officer_members": self.placement_officer_members or "",
            "total_targets": self.total_targets,
            "sent": self.sent,
            "failed": self.failed,
            "industry": self.industry or "Default",
            "country": self.country or "Global",
            "bounced": self.bounced,
            "skipped": self.skipped,
            "is_deleted": bool(self.is_deleted),
            "user_id": self.user_id,
        }

class RunCandidateReport(db.Model):
    __tablename__ = "run_candidate_reports"
    id = db.Column(db.Integer, primary_key=True)
    run_id = db.Column(db.Integer, nullable=False, index=True)
    candidate_id = db.Column(db.Integer, nullable=False, index=True)
    candidate_name = db.Column(db.String(200), nullable=False)
    report_csv_path = db.Column(db.String(500), nullable=False)
    created_at = db.Column(db.DateTime, nullable=False, default=func.now())

    def to_dict(self):
        return {
            "run_id": self.run_id,
            "candidate_id": self.candidate_id,
            "candidate_name": self.candidate_name,
            "download_url": f"/api/runs/{self.run_id}/candidates/{self.candidate_id}/report"
        }

class JobApplication(db.Model):
    __tablename__ = "job_applications"
    id = db.Column(db.Integer, primary_key=True)
    candidate_id = db.Column(db.Integer, db.ForeignKey('candidates.id'), nullable=False)
    company_name = db.Column(db.String(250), nullable=False)
    job_role = db.Column(db.String(250), nullable=False)
    country = db.Column(db.String(100), nullable=True)
    applied_date = db.Column(db.String(50), nullable=True)
    status = db.Column(db.String(50), nullable=False, default="Applied")
    screenshot_path = db.Column(db.String(500), nullable=True)
    last_sync_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=func.now())

    def to_dict(self):
        return {
            "id": self.id,
            "candidate_id": self.candidate_id,
            "company_name": self.company_name,
            "job_role": self.job_role,
            "country": self.country or "",
            "applied_date": self.applied_date or "",
            "status": self.status or "Applied",
            "screenshot_path": self.screenshot_path or "",
            "last_sync_at": self.last_sync_at.isoformat() if self.last_sync_at else "",
            "created_at": self.created_at.isoformat() if self.created_at else ""
        }

class EmailEvent(db.Model):
    __tablename__ = "email_events"
    id = db.Column(db.Integer, primary_key=True)
    candidate_id = db.Column(db.Integer, db.ForeignKey('candidates.id'), nullable=False)
    job_app_id = db.Column(db.Integer, db.ForeignKey('job_applications.id'), nullable=True)
    event_type = db.Column(db.String(50), nullable=False) # Assessment, Interview, Offer
    company_name = db.Column(db.String(250), nullable=True)
    job_role = db.Column(db.String(250), nullable=True)
    subject = db.Column(db.String(500), nullable=True)
    received_at = db.Column(db.DateTime, nullable=True)
    created_at = db.Column(db.DateTime, nullable=False, default=func.now())

    def to_dict(self):
        return {
            "id": self.id,
            "candidate_id": self.candidate_id,
            "job_app_id": self.job_app_id,
            "event_type": self.event_type,
            "company_name": self.company_name or "",
            "job_role": self.job_role or "",
            "subject": self.subject or "",
            "received_at": self.received_at.isoformat() if self.received_at else "",
            "created_at": self.created_at.isoformat() if self.created_at else ""
        }

# ---- Pending Changes (Approval Workflow) ----
class PendingChange(db.Model):
    __tablename__ = "pending_changes"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=False)
    candidate_id = db.Column(db.Integer, db.ForeignKey('candidates.id'), nullable=False)
    change_type = db.Column(db.String(50), nullable=False, default="candidate_edit")
    change_data = db.Column(db.Text, nullable=False)  # JSON serialized changes
    status = db.Column(db.String(20), nullable=False, default="pending")  # pending, approved, rejected
    created_at = db.Column(db.DateTime, nullable=False, default=func.now())
    reviewed_by = db.Column(db.Integer, db.ForeignKey('users.id'), nullable=True)
    reviewed_at = db.Column(db.DateTime, nullable=True)
    review_note = db.Column(db.Text, nullable=True)

    def to_dict(self):
        submitter = db.session.get(User, self.user_id)
        reviewer = db.session.get(User, self.reviewed_by) if self.reviewed_by else None
        candidate = db.session.get(Candidate, self.candidate_id)
        return {
            "id": self.id,
            "user_id": self.user_id,
            "submitted_by": submitter.username if submitter else "Unknown",
            "candidate_id": self.candidate_id,
            "candidate_name": candidate.name if candidate else "Deleted",
            "change_type": self.change_type,
            "change_data": json.loads(self.change_data) if self.change_data else {},
            "status": self.status,
            "created_at": self.created_at.isoformat() if self.created_at else "",
            "reviewed_by": reviewer.username if reviewer else "",
            "reviewed_at": self.reviewed_at.isoformat() if self.reviewed_at else "",
            "review_note": self.review_note or "",
        }

# ---- 6-Month Workflow Automation ----
WORKFLOW_PHASES = {
    1: {"days": (0, 30),   "target": 100,  "cumulative": 100,  "per_batch": 10,  "alternate": True,  "label": "Phase 1 (0-30 Days)"},
    2: {"days": (31, 60),  "target": 200,  "cumulative": 300,  "per_batch": 20,  "alternate": True,  "label": "Phase 2 (31-60 Days)"},
    3: {"days": (61, 90),  "target": 200,  "cumulative": 500,  "per_batch": 40,  "alternate": True,  "label": "Phase 3 (61-90 Days)"},
    4: {"days": (91, 120), "target": 200,  "cumulative": 700,  "per_batch": 50,  "alternate": True,  "label": "Phase 4 (91-120 Days)"},
    5: {"days": (121,150), "target": 200,  "cumulative": 900,  "per_batch": 100, "alternate": True,  "label": "Phase 5 (121-150 Days)"},
    6: {"days": (151,180), "target": 200,  "cumulative": 1100, "per_batch": 150, "alternate": True,  "label": "Phase 6 (151-180 Days)"},
    7: {"days": (181,9999), "target": 100, "cumulative": 1200, "per_batch": 100, "alternate": False, "label": "Overflow (>180 Days)"},
}


def _workflow_phase_num_for_day(day_num: int) -> int:
    for phase_num, info in WORKFLOW_PHASES.items():
        d_start, d_end = info["days"]
        if d_start <= day_num <= d_end:
            return phase_num
    return 7


def workflow_phases_prediction_payload(
    days_elapsed: int,
    total_applied: int,
    total_target: int = 1200,
    service_start=None,
):
    """Per-phase status + calendar dates for the 6-month / 1200-app Overall Plan UI."""
    days_elapsed = max(0, int(days_elapsed))
    total_applied = max(0, int(total_applied))
    total_target = max(1, int(total_target or 1200))
    current_p = _workflow_phase_num_for_day(days_elapsed)
    overall_pct = min(100.0, round((total_applied / float(total_target)) * 100, 1))
    phases = []
    for pnum in sorted(WORKFLOW_PHASES.keys()):
        info = WORKFLOW_PHASES[pnum]
        d0, d1 = info["days"]
        if pnum < current_p:
            st = "completed"
        elif pnum == current_p:
            st = "current"
        else:
            st = "upcoming"
        dr = f"{d0}–{d1}" if d1 < 9000 else f"{d0}+"
        period_start = ""
        period_end = ""
        if service_start:
            try:
                ps_dt = service_start + timedelta(days=d0)
                period_start = ps_dt.strftime("%d-%m-%Y")
                if d1 >= 9000:
                    period_end = "—"
                else:
                    pe_dt = service_start + timedelta(days=d1)
                    period_end = pe_dt.strftime("%d-%m-%Y")
            except Exception:
                period_start = ""
                period_end = ""
        phases.append({
            "phase": pnum,
            "label": info["label"],
            "days_range": dr,
            "cumulative_target": int(info["cumulative"]),
            "per_batch": int(info["per_batch"]),
            "phase_app_target": int(info["target"]),
            "period_start": period_start,
            "period_end": period_end,
            "status": st,
        })
    return {
        "overall_progress_pct": overall_pct,
        "current_phase": current_p,
        "phases": phases,
    }


def workflow_plan_elapsed_service_days(plan: "WorkflowPlan", now: Optional[datetime] = None) -> int:
    """Elapsed service day for curve + batch math (aligned with profile + bucklist pin + plan start).

    Takes the **maximum** of calendar days since plan start and any authoritative day the user
    set on the candidate (`smart_service_start_date`, `bucklist_days_in_system`). Using only plan
    start would ignore the bucklist pin when both exist.
    """
    now = now or datetime.utcnow()
    if not plan or not plan.service_start_date:
        return 0
    elapsed_plan = max(0, (now - plan.service_start_date).days)
    cand = db.session.get(Candidate, plan.candidate_id) if getattr(plan, "candidate_id", None) else None
    if not cand:
        return elapsed_plan
    extra = []
    if cand.smart_service_start_date:
        anchor = _dt_naive_for_delta(cand.smart_service_start_date)
        if anchor:
            extra.append(max(0, (now - anchor).days))
    if cand.bucklist_days_in_system is not None:
        try:
            extra.append(max(0, int(cand.bucklist_days_in_system)))
        except (TypeError, ValueError):
            pass
    cand_max = max(extra) if extra else 0
    rep_el = 0
    if getattr(plan, "reported_service_day", None) is not None and getattr(plan, "reported_service_anchor_at", None):
        try:
            rep0 = max(0, int(plan.reported_service_day))
            rep_el = rep0 + max(0, (now - plan.reported_service_anchor_at).days)
        except (TypeError, ValueError):
            rep_el = 0
    return max(elapsed_plan, cand_max, rep_el)


def workflow_plan_phase_tracker_rows(plan: "WorkflowPlan"):
    """Per-phase dates + Yes/No vs the official day curve (linear_expected_applications_for_days).

    - **Current** phase: compare total_applied to the curve at **today's** elapsed service day.
    - **Past** phase: compare total_applied to the curve at the **last day** of that phase.
    - **Upcoming**: no comparison yet.
    Row-level ``expected_for_today`` / ``on_track`` use the same curve at ``elapsed_days``.
    """
    svc = plan.service_start_date
    now = datetime.utcnow()
    elapsed = workflow_plan_elapsed_service_days(plan, now)
    applied = int(plan.total_applied or 0)
    expected_for_today = linear_expected_applications_for_days(elapsed)
    on_track_now = applied >= expected_for_today
    rows = []
    for pnum in sorted(WORKFLOW_PHASES.keys()):
        info = WORKFLOW_PHASES[pnum]
        d0, d1 = info["days"]
        cumulative = int(info["cumulative"])
        period_start = ""
        period_end = ""
        if svc:
            try:
                ps_dt = svc + timedelta(days=d0)
                period_start = ps_dt.strftime("%d-%m-%Y")
                if d1 >= 9000:
                    period_end = "—"
                else:
                    pe_dt = svc + timedelta(days=d1)
                    period_end = pe_dt.strftime("%d-%m-%Y")
            except Exception:
                period_start = ""
                period_end = ""

        bench_day = None
        exp_apps = None
        if elapsed < d0:
            roll = "upcoming"
            rtarget_met = None
        elif d1 >= 9000:
            if elapsed < d0:
                roll = "upcoming"
                rtarget_met = None
            else:
                roll = "current"
                bench_day = elapsed
                exp_apps = linear_expected_applications_for_days(elapsed)
                rtarget_met = applied >= exp_apps
        elif elapsed > d1:
            roll = "past"
            bench_day = int(d1)
            exp_apps = linear_expected_applications_for_days(int(d1))
            rtarget_met = applied >= exp_apps
        else:
            roll = "current"
            bench_day = elapsed
            exp_apps = linear_expected_applications_for_days(elapsed)
            rtarget_met = applied >= exp_apps

        if rtarget_met is None:
            achieved_label = "—"
        elif rtarget_met:
            achieved_label = "Yes"
        else:
            achieved_label = "No"

        dr = f"{d0}–{d1}" if d1 < 9000 else f"{d0}+"
        rows.append({
            "phase": pnum,
            "label": info["label"],
            "days_range": dr,
            "period_start": period_start,
            "period_end": period_end,
            "cumulative_target": cumulative,
            "benchmark_day": bench_day,
            "expected_applications": exp_apps,
            "roll": roll,
            "target_met": rtarget_met,
            "achieved_label": achieved_label,
        })
    return {
        "has_plan": True,
        "plan_id": plan.id,
        "candidate_id": plan.candidate_id,
        "candidate_name": (plan.candidate_name or "").strip(),
        "plan_status": (plan.status or "active").strip().lower(),
        "elapsed_days": elapsed,
        "total_applied": applied,
        "total_target": int(plan.total_target or 1200),
        "expected_for_today": expected_for_today,
        "on_track": bool(on_track_now),
        "applications_ahead_or_behind": int(applied - expected_for_today),
        "service_start_date": plan.service_start_date.isoformat() if plan.service_start_date else "",
        "phases": rows,
    }


def workflow_tracker_row_no_plan(cand: Candidate):
    """Dashboard row when the candidate has no workflow plan."""
    return {
        "has_plan": False,
        "plan_id": None,
        "candidate_id": cand.id,
        "candidate_name": (cand.name or "").strip(),
        "plan_status": "none",
        "elapsed_days": None,
        "total_applied": None,
        "total_target": None,
        "service_start_date": "",
        "phases": [],
    }


def linear_expected_applications_for_days(days: int) -> int:
    days = max(0, int(days))
    if days <= 0:
        return 0
    if days > 180:
        return 1200
    knots = [(0, 0), (30, 100), (60, 300), (90, 500), (120, 700), (150, 900), (180, 1100)]
    for i in range(len(knots) - 1):
        d0, a0 = knots[i]
        d1, a1 = knots[i + 1]
        if days <= d1:
            if d1 == d0:
                return a1
            return round(a0 + (a1 - a0) * (days - d0) / (d1 - d0))
    return 1100


def days_bucket_short(days: int) -> str:
    """Human-readable day bucket for Smart Automation dashboards."""
    d = max(0, int(days))
    if d <= 30:
        return "0-30"
    if d <= 60:
        return "31-60"
    if d <= 90:
        return "61-90"
    if d <= 120:
        return "91-120"
    if d <= 150:
        return "121-150"
    if d <= 180:
        return "151-180"
    return "180+"


def compute_smart_automation_plan(
    days: int,
    applied: int,
    max_per_run: int = 100,
    service_cap: int = 1200,
    interval_days: int = 1,
    backlog_safe_cap: int = None,
    candidate_id: Optional[int] = None,
):
    """On-track (no catch-up): batch 0. Behind curve (backlog): min(pending, backlog_safe_cap, remaining); not limited by max_per_run.

    When ``candidate_id`` is set, suggested_batch is also clipped by **remaining daily quota** (100 − sends today).
    """
    if backlog_safe_cap is None:
        backlog_safe_cap = SMART_AUTOMATION_BACKLOG_SAFE_CAP
    backlog_safe_cap = max(1, min(500, int(backlog_safe_cap)))

    days = max(0, int(days))
    applied = max(0, int(applied))
    max_per_run = max(1, min(500, int(max_per_run)))
    interval_days = max(1, min(7, int(interval_days)))
    remaining = max(0, service_cap - applied)
    phase_num = _workflow_phase_num_for_day(days)
    pinfo = WORKFLOW_PHASES.get(phase_num, WORKFLOW_PHASES[7])
    phase_cap = int(pinfo["per_batch"])
    expected = linear_expected_applications_for_days(days)
    on_track_delta = applied - expected
    pending_on_track = max(0, expected - applied)

    if pending_on_track > 0:
        # Backlog: catch up to the curve, capped by backlog_safe_cap (e.g. 100/day), never by max_per_run
        batch = min(pending_on_track, backlog_safe_cap, remaining)
        backlog_mode = True
    else:
        # On track or ahead: curve says no catch-up — use calculated zero (no forced sends)
        batch = 0
        backlog_mode = False

    if days <= 180:
        days_left = max(1, 180 - days)
    else:
        days_left = max(1, int(math.ceil(remaining / max(1, backlog_safe_cap))))

    schedule_slots = max(1, int(math.ceil(days_left / float(interval_days))))

    sent_today = emails_sent_today_for_candidate(candidate_id) if candidate_id else 0
    daily_left = remaining_daily_send_quota(candidate_id)
    batch = min(int(batch), daily_left)

    return {
        "phase": phase_num,
        "phase_label": pinfo["label"],
        "phase_batch_cap": phase_cap,
        "expected_applications_by_now": expected,
        "on_track_delta": on_track_delta,
        "pending_on_track": pending_on_track,
        "remaining_to_cap": remaining,
        "suggested_batch_per_run": int(batch),
        "backlog_mode": backlog_mode,
        "backlog_safe_cap": backlog_safe_cap,
        "alternate_interval_days": interval_days,
        "estimated_alternate_slots_remaining": schedule_slots,
        "days_left_in_180_window": max(0, 180 - days) if days <= 180 else 0,
        "service_cap": service_cap,
        "days_bucket": days_bucket_short(days),
        "sent_today_for_daily_cap": int(sent_today),
        "remaining_daily_quota": int(daily_left),
    }


def workflow_plan_effective_send_batch(
    plan: "WorkflowPlan",
    now: Optional[datetime] = None,
    *,
    scheduled: bool = True,
) -> int:
    """Batch size for 6-month scheduled/manual sends.

    Behind the official day curve: catch up toward today's target, capped at 100 and remaining target.
    On or ahead of curve and ``scheduled`` is True: **no** sends — wait until elapsed days / curve
    raises the expected target (same rule as Smart Automation catch-up).
    On or ahead when ``scheduled`` is False (operator Send-now / manual run): phase playbook per_batch.
    """
    now = now or datetime.utcnow()
    if not plan or not plan.service_start_date:
        return 0
    elapsed = workflow_plan_elapsed_service_days(plan, now)
    applied = int(plan.total_applied or 0)
    tgt = int(plan.total_target or 1200)
    remaining = max(0, tgt - applied)
    if remaining <= 0:
        return 0
    per_run_cap = max(1, min(500, int(SMART_AUTOMATION_BACKLOG_SAFE_CAP)))
    expected = linear_expected_applications_for_days(elapsed)
    pending = max(0, expected - applied)
    phase_num = plan._get_phase_for_day(elapsed)
    phase_info = WORKFLOW_PHASES.get(phase_num, WORKFLOW_PHASES[7])
    phase_batch = int(phase_info["per_batch"])
    if pending > 0:
        raw = min(pending, per_run_cap, remaining)
    elif scheduled:
        raw = 0
    else:
        raw = min(phase_batch, per_run_cap, remaining)
    day_left = remaining_daily_send_quota(getattr(plan, "candidate_id", None), now)
    return min(raw, day_left)


def workflow_next_run_effective_display(
    next_run: Optional[datetime],
    now: Optional[datetime] = None,
    plan_status: str = "active",
) -> Optional[datetime]:
    """If next_run is in the past for an active plan, return the next calendar slot at the same clock time (Sundays skipped). Does not change the database — scheduler still sees the original overdue time."""
    now = now or datetime.utcnow()
    if next_run is None:
        return None
    if (plan_status or "").lower() != "active":
        return next_run
    if next_run > now:
        return next_run
    t = next_run
    guard = 0
    while t <= now and guard < 400:
        t = t + timedelta(days=1)
        while t.weekday() == 6:
            t = t + timedelta(days=1)
        guard += 1
    return t


class WorkflowPlan(db.Model):
    __tablename__ = "workflow_plans"
    id = db.Column(db.Integer, primary_key=True)
    candidate_id = db.Column(db.Integer, db.ForeignKey('candidates.id'), nullable=False)
    candidate_name = db.Column(db.String(200), nullable=True)
    enrollment_id = db.Column(db.String(100), nullable=True)

    # User Preferences
    country = db.Column(db.String(100), nullable=True)
    industry = db.Column(db.String(100), nullable=True)
    scheduled_start_time = db.Column(db.DateTime, nullable=True)

    # Service dates
    service_start_date = db.Column(db.DateTime, nullable=False)
    service_end_date = db.Column(db.DateTime, nullable=True)

    # Status: active, paused, completed, expired
    status = db.Column(db.String(20), default="active")
    current_phase = db.Column(db.Integer, default=1)

    # Progress
    total_target = db.Column(db.Integer, default=1200)
    total_applied = db.Column(db.Integer, default=0)

    # Scheduling
    is_alternate_day = db.Column(db.Boolean, default=True)
    last_run_date = db.Column(db.DateTime, nullable=True)
    next_run_date = db.Column(db.DateTime, nullable=True)

    created_at = db.Column(db.DateTime, nullable=False, default=func.now())
    updated_at = db.Column(db.DateTime, nullable=False, default=func.now(), onupdate=func.now())

    # User-reported "service day" at reported_service_anchor_at; grows as (now - anchor).days are added.
    # Keeps batch sizing aligned with 6-Month calculator when service_start_date / profile drift.
    reported_service_day = db.Column(db.Integer, nullable=True)
    reported_service_anchor_at = db.Column(db.DateTime, nullable=True)

    def _get_phase_for_day(self, day_num):
        """Return phase number for the given elapsed day."""
        for phase_num, info in WORKFLOW_PHASES.items():
            d_start, d_end = info["days"]
            if d_start <= day_num <= d_end:
                return phase_num
        return 7  # overflow

    def get_dynamic_batch_size(self, elapsed):
        remaining_days = max(0, 180 - elapsed)
        if remaining_days <= 0:
            return 0
        now = datetime.utcnow()
        active_days = 0
        for i in range(1, remaining_days + 1):
            day = now + timedelta(days=i)
            if day.weekday() != 6:  # 6 is Sunday
                active_days += 1
        remaining_target = max(0, self.total_target - self.total_applied)
        return math.ceil(remaining_target / max(1, active_days))

    def compute_status(self):
        """Compute current phase + progress details."""
        now = datetime.utcnow()
        elapsed = workflow_plan_elapsed_service_days(self, now)
        phase_num = self._get_phase_for_day(elapsed)
        phase_info = WORKFLOW_PHASES.get(phase_num, WORKFLOW_PHASES[7])

        # Progress percentage
        progress_pct = round((self.total_applied / self.total_target) * 100, 1) if self.total_target > 0 else 0
        progress_pct = min(progress_pct, 100.0)

        # Days remaining in service (180 days)
        service_days = 180
        days_remaining = max(0, service_days - elapsed)

        # Phase progress
        phase_target = phase_info["cumulative"]
        phase_progress_pct = round((self.total_applied / phase_target) * 100, 1) if phase_target > 0 else 0
        phase_progress_pct = min(phase_progress_pct, 100.0)

        remaining_target = max(0, self.total_target - self.total_applied)
        phase_batch = workflow_plan_effective_send_batch(self, now)

        return {
            "elapsed_days": elapsed,
            "current_phase": phase_num,
            "phase_label": phase_info["label"],
            "phase_target": phase_info["cumulative"],
            "phase_batch_size": phase_batch,
            "phase_is_alternate": phase_info["alternate"],
            "overall_progress_pct": progress_pct,
            "phase_progress_pct": phase_progress_pct,
            "days_remaining": days_remaining,
            "is_overflow": phase_num == 7,
        }

    def to_dict(self):
        status_info = self.compute_status()
        now = datetime.utcnow()
        nr_raw = self.next_run_date
        nr_show = workflow_next_run_effective_display(nr_raw, now, self.status or "active")
        return {
            "id": self.id,
            "candidate_id": self.candidate_id,
            "candidate_name": self.candidate_name or "",
            "enrollment_id": self.enrollment_id or "",
            "country": self.country or "",
            "industry": self.industry or "",
            "scheduled_start_time": self.scheduled_start_time.isoformat() if self.scheduled_start_time else "",
            "service_start_date": self.service_start_date.isoformat() if self.service_start_date else "",
            "service_end_date": self.service_end_date.isoformat() if self.service_end_date else "",
            "status": self.status or "active",
            "current_phase": status_info["current_phase"],
            "total_target": self.total_target or 1200,
            "total_applied": self.total_applied or 0,
            "last_run_date": self.last_run_date.isoformat() if self.last_run_date else "",
            "next_run_date": nr_show.isoformat() if nr_show else "",
            "next_run_date_stored": nr_raw.isoformat() if nr_raw else "",
            "created_at": self.created_at.isoformat() if self.created_at else "",
            "updated_at": self.updated_at.isoformat() if self.updated_at else "",
            # Computed fields
            "elapsed_days": status_info["elapsed_days"],
            "phase_label": status_info["phase_label"],
            "phase_target": status_info["phase_target"],
            "phase_batch_size": status_info["phase_batch_size"],
            "phase_is_alternate": status_info["phase_is_alternate"],
            "overall_progress_pct": status_info["overall_progress_pct"],
            "phase_progress_pct": status_info["phase_progress_pct"],
            "days_remaining": status_info["days_remaining"],
            "is_overflow": status_info["is_overflow"],
        }


def sync_workflow_plan_hr_from_candidate(cand: Candidate) -> None:
    """Apply this candidate's HR profile to their 6-month plan and workspaces (that candidate only)."""
    if not cand or not getattr(cand, "id", None):
        return
    cty = (cand.smart_country or "").strip()
    ind = (cand.smart_industry or "").strip()
    cty_f = cty if cty else "Global"
    ind_f = ind if ind else "Default"
    wp = (
        WorkflowPlan.query.filter_by(candidate_id=cand.id)
        .filter(WorkflowPlan.status.in_(["active", "paused"]))
        .order_by(WorkflowPlan.id.desc())
        .first()
    )
    if wp:
        wp.country = cty_f
        wp.industry = ind_f
    for ws in Workspace.query.filter_by(candidate_id=cand.id).all():
        ws.country = cty_f
        ws.industry = ind_f


def candidate_aux_service_starts(candidate_id: int) -> Tuple[Optional[datetime], Optional[datetime]]:
    """Latest active/paused workflow plan start, and latest workspace service_start_date, for one candidate."""
    wf_start = None
    plan = (
        WorkflowPlan.query.filter_by(candidate_id=candidate_id)
        .filter(WorkflowPlan.status.in_(["active", "paused"]))
        .order_by(WorkflowPlan.id.desc())
        .first()
    )
    if plan and plan.service_start_date:
        wf_start = plan.service_start_date
    ws_start = None
    ws = Workspace.query.filter_by(candidate_id=candidate_id).order_by(Workspace.id.desc()).first()
    if ws and getattr(ws, "service_start_date", None):
        ws_start = ws.service_start_date
    return wf_start, ws_start


def candidate_service_start_maps_for_ids(candidate_ids) -> Tuple[dict, dict]:
    """Per candidate id: workflow service_start_date (newest active/paused plan) and workspace service_start_date."""
    if not candidate_ids:
        return {}, {}
    plan_map: dict = {}
    for p in (
        WorkflowPlan.query.filter(WorkflowPlan.candidate_id.in_(candidate_ids))
        .filter(WorkflowPlan.status.in_(["active", "paused"]))
        .order_by(WorkflowPlan.id.desc())
        .all()
    ):
        if p.candidate_id not in plan_map and p.service_start_date:
            plan_map[p.candidate_id] = p.service_start_date
    ws_map: dict = {}
    for ws in Workspace.query.filter(Workspace.candidate_id.in_(candidate_ids)).order_by(Workspace.id.desc()).all():
        if ws.candidate_id not in ws_map and getattr(ws, "service_start_date", None):
            ws_map[ws.candidate_id] = ws.service_start_date
    return plan_map, ws_map


def _column_exists(table, col):
    rows = db.session.execute(text(f"PRAGMA table_info({table})")).fetchall()
    return any(r[1] == col for r in rows)

def migrate_sqlite():
    try:
        if not _column_exists("send_runs", "candidate_names"):
            db.session.execute(text("ALTER TABLE send_runs ADD COLUMN candidate_names TEXT"))
            db.session.commit()
        if not _column_exists("send_runs", "pa_members"):
            db.session.execute(text("ALTER TABLE send_runs ADD COLUMN pa_members TEXT"))
            db.session.commit()
        if not _column_exists("send_runs", "placement_officer_members"):
            db.session.execute(text("ALTER TABLE send_runs ADD COLUMN placement_officer_members TEXT"))
            db.session.commit()
        if not _column_exists("send_runs", "country"):
            db.session.execute(text("ALTER TABLE send_runs ADD COLUMN country VARCHAR(100) DEFAULT 'Global'"))
            db.session.commit()
        if not _column_exists("send_runs", "is_deleted"):
            db.session.execute(text("ALTER TABLE send_runs ADD COLUMN is_deleted INTEGER DEFAULT 0"))
            db.session.commit()
        if not _column_exists("send_runs", "industry"):
            db.session.execute(text("ALTER TABLE send_runs ADD COLUMN industry TEXT DEFAULT 'Default'"))
            db.session.commit()
        if not _column_exists("industries", "country"):
            db.session.execute(text("ALTER TABLE industries ADD COLUMN country VARCHAR(100) DEFAULT 'Global'"))
            db.session.commit()
        
        if not _column_exists("email_events", "job_role"):
            db.session.execute(text("ALTER TABLE email_events ADD COLUMN job_role VARCHAR(250)"))
            db.session.commit()

        # Check if industries table has the correct composite UNIQUE constraint
        # PRAGMA index_list returns list of indexes. Auto-indexes for UNIQUE constraints are included.
        indices = db.session.execute(text("PRAGMA index_list(industries)")).fetchall()
        # We want to find if there's a unique index that is NOT just on 'name'
        is_composite = False
        for idx in indices:
            idx_name = idx[1]
            columns = db.session.execute(text(f"PRAGMA index_info('{idx_name}')")).fetchall()
            col_names = [c[2] for c in columns]
            if "name" in col_names and "country" in col_names:
                is_composite = True
                break
        
        # If not composite, we need to recreate the table to change the UNIQUE constraint
        if not is_composite:
            # 1. Rename old table
            db.session.execute(text("ALTER TABLE industries RENAME TO industries_old"))
            # 2. Create new table with correct constraint
            db.session.execute(text("""
                CREATE TABLE industries (
                    id INTEGER NOT NULL, 
                    name VARCHAR(150) NOT NULL, 
                    country VARCHAR(100) DEFAULT 'Global', 
                    created_at DATETIME NOT NULL, 
                    PRIMARY KEY (id), 
                    UNIQUE (name, country)
                )
            """))
            # 3. Copy data
            db.session.execute(text("INSERT INTO industries (id, name, country, created_at) SELECT id, name, COALESCE(country, 'Global'), created_at FROM industries_old"))
            # 4. Drop old table
            db.session.execute(text("DROP TABLE industries_old"))
            db.session.commit()
            
        if not _column_exists("workspaces", "country"):
            db.session.execute(text("ALTER TABLE workspaces ADD COLUMN country VARCHAR(100) DEFAULT 'Global'"))
            db.session.commit()

        # Users table
        db.create_all()

        if not _column_exists("candidates", "pa_member"):
            db.session.execute(text("ALTER TABLE candidates ADD COLUMN pa_member TEXT"))
            db.session.commit()
            # backfill from legacy app_code
            db.session.execute(text("UPDATE candidates SET pa_member = COALESCE(pa_member, app_code) WHERE app_code IS NOT NULL AND (pa_member IS NULL OR pa_member = '')"))
            db.session.commit()
        if not _column_exists("candidates", "placement_officer_member"):
            db.session.execute(text("ALTER TABLE candidates ADD COLUMN placement_officer_member TEXT"))
            db.session.commit()
        if not _column_exists("candidates", "rm_member"):
            db.session.execute(text("ALTER TABLE candidates ADD COLUMN rm_member TEXT"))
            db.session.commit()
            try:
                db.session.execute(
                    text(
                        "UPDATE candidates SET rm_member = TRIM(placement_officer_member) "
                        "WHERE (rm_member IS NULL OR TRIM(COALESCE(rm_member, '')) = '') "
                        "AND placement_officer_member IS NOT NULL AND TRIM(placement_officer_member) != ''"
                    )
                )
                db.session.commit()
            except Exception:
                db.session.rollback()
        if not _column_exists("candidates", "scheduled_time"):
            db.session.execute(text("ALTER TABLE candidates ADD COLUMN scheduled_time DATETIME"))
            db.session.commit()
        if not _column_exists("candidates", "enrollment_id"):
            db.session.execute(text("ALTER TABLE candidates ADD COLUMN enrollment_id VARCHAR(100)"))
            db.session.commit()
        if not _column_exists("candidates", "industry_types"):
            db.session.execute(text("ALTER TABLE candidates ADD COLUMN industry_types TEXT"))
            db.session.commit()
        if not _column_exists("candidates", "smart_service_start_date"):
            db.session.execute(text("ALTER TABLE candidates ADD COLUMN smart_service_start_date DATETIME"))
            db.session.commit()
        if not _column_exists("candidates", "smart_baseline_applied"):
            db.session.execute(text("ALTER TABLE candidates ADD COLUMN smart_baseline_applied INTEGER DEFAULT 0"))
            db.session.commit()
        if not _column_exists("candidates", "smart_country"):
            db.session.execute(text("ALTER TABLE candidates ADD COLUMN smart_country VARCHAR(100)"))
            db.session.commit()
        if not _column_exists("candidates", "smart_industry"):
            db.session.execute(text("ALTER TABLE candidates ADD COLUMN smart_industry VARCHAR(150)"))
            db.session.commit()
        if not _column_exists("candidates", "bucklist_days_in_system"):
            db.session.execute(text("ALTER TABLE candidates ADD COLUMN bucklist_days_in_system INTEGER"))
            db.session.commit()
        if not _column_exists("candidates", "enrollment_status"):
            db.session.execute(text("ALTER TABLE candidates ADD COLUMN enrollment_status VARCHAR(40)"))
            db.session.commit()
        # Before any ORM load of Candidate (e.g. workflow reported_service backfill below).
        if not _column_exists("candidates", "scheduler_automation_at"):
            db.session.execute(text("ALTER TABLE candidates ADD COLUMN scheduler_automation_at DATETIME"))
            db.session.commit()

        if not _column_exists("targets", "industry"):
            db.session.execute(text("ALTER TABLE targets ADD COLUMN industry TEXT DEFAULT 'Default'"))
            db.session.commit()

        # Workspace migrations
        if not _column_exists("workspaces", "automation_type"):
            db.session.execute(text("ALTER TABLE workspaces ADD COLUMN automation_type TEXT DEFAULT 'interval'"))
            db.session.commit()
        if not _column_exists("workspaces", "scheduled_days"):
            db.session.execute(text("ALTER TABLE workspaces ADD COLUMN scheduled_days TEXT DEFAULT 'Tuesday,Wednesday,Thursday,Friday'"))
            db.session.commit()
        if not _column_exists("workspaces", "service_start_date"):
            db.session.execute(text("ALTER TABLE workspaces ADD COLUMN service_start_date DATETIME"))
            db.session.commit()
        if not _column_exists("workspaces", "service_end_date"):
            db.session.execute(text("ALTER TABLE workspaces ADD COLUMN service_end_date DATETIME"))
            db.session.commit()
        if not _column_exists("workspaces", "monthly_target"):
            db.session.execute(text("ALTER TABLE workspaces ADD COLUMN monthly_target INTEGER DEFAULT 200"))
            db.session.commit()
        if not _column_exists("workspaces", "monthly_sent_count"):
            db.session.execute(text("ALTER TABLE workspaces ADD COLUMN monthly_sent_count INTEGER DEFAULT 0"))
            db.session.commit()
        if not _column_exists("workspaces", "last_month_reset"):
            db.session.execute(text("ALTER TABLE workspaces ADD COLUMN last_month_reset DATETIME"))
            db.session.commit()
        if not _column_exists("workspaces", "automation_per_run_cap"):
            db.session.execute(text("ALTER TABLE workspaces ADD COLUMN automation_per_run_cap INTEGER DEFAULT 100"))
            db.session.commit()

        if not _column_exists("job_applications", "status"):
            db.session.execute(text("ALTER TABLE job_applications ADD COLUMN status VARCHAR(50) DEFAULT 'Applied'"))
            db.session.commit()
        if not _column_exists("job_applications", "last_sync_at"):
            db.session.execute(text("ALTER TABLE job_applications ADD COLUMN last_sync_at DATETIME"))
            db.session.commit()
            
        # Users table migrations
        if not _column_exists("users", "is_verified"):
            db.session.execute(text("ALTER TABLE users ADD COLUMN is_verified INTEGER DEFAULT 0"))
            db.session.commit()
        if not _column_exists("users", "is_approved"):
            db.session.execute(text("ALTER TABLE users ADD COLUMN is_approved INTEGER DEFAULT 0"))
            db.session.commit()
        if not _column_exists("users", "role"):
            db.session.execute(text("ALTER TABLE users ADD COLUMN role VARCHAR(20) DEFAULT 'user'"))
            db.session.commit()
        if not _column_exists("users", "allowed_features"):
            db.session.execute(text("ALTER TABLE users ADD COLUMN allowed_features TEXT"))
            db.session.commit()

        # Promote first user to admin if no admins exist
        try:
            admin_count = User.query.filter_by(role='admin').count()
            if admin_count == 0:
                first_user = User.query.order_by(User.id).first()
                if first_user:
                    first_user.role = 'admin'
                    first_user.is_approved = True
                    db.session.commit()
                    print(f"[Migration] Promoted {first_user.username} to admin.")
        except Exception as e:
            print(f"[Migration] Error promoting admin: {e}")

        # SendRun migrations
        if not _column_exists("send_runs", "user_id"):
            db.session.execute(text("ALTER TABLE send_runs ADD COLUMN user_id INTEGER"))
            db.session.commit()

        if not _column_exists("workflow_plans", "reported_service_day"):
            db.session.execute(text("ALTER TABLE workflow_plans ADD COLUMN reported_service_day INTEGER"))
            db.session.commit()
        if not _column_exists("workflow_plans", "reported_service_anchor_at"):
            db.session.execute(text("ALTER TABLE workflow_plans ADD COLUMN reported_service_anchor_at DATETIME"))
            db.session.commit()

        try:
            anchor_fill = datetime.utcnow()
            for p in WorkflowPlan.query.filter(WorkflowPlan.reported_service_day.is_(None)).all():
                d0 = 0
                if p.service_start_date:
                    d0 = max(0, (anchor_fill - p.service_start_date).days)
                cand = db.session.get(Candidate, p.candidate_id)
                if cand and cand.bucklist_days_in_system is not None:
                    try:
                        d0 = max(d0, int(cand.bucklist_days_in_system))
                    except (TypeError, ValueError):
                        pass
                p.reported_service_day = min(2000, int(d0))
                p.reported_service_anchor_at = anchor_fill
            db.session.commit()
        except Exception as ex:
            db.session.rollback()
            print(f"[Migration] workflow_plans reported_service backfill skipped: {ex}")

        # Database is initialized in the entry point
        pass

        # WorkflowPlan table (auto-created by db.create_all, but ensure columns exist)
        # No extra migrations needed here since db.create_all() handles new tables.

    except Exception:
        db.session.rollback()

with app.app_context():
    db.create_all()
    migrate_sqlite()
    try:
        if not _column_exists("users", "allowed_features"):
            db.session.execute(text("ALTER TABLE users ADD COLUMN allowed_features TEXT"))
            db.session.commit()
            print("[DB] Added users.allowed_features column.")
    except Exception as _afe:
        db.session.rollback()
        print(f"[DB] Could not ensure users.allowed_features: {_afe}")
    # Ensure industry_types exists even if migrate_sqlite() exited early on another error
    try:
        if not _column_exists("candidates", "industry_types"):
            db.session.execute(text("ALTER TABLE candidates ADD COLUMN industry_types TEXT"))
            db.session.commit()
            print("[DB] Added candidates.industry_types column.")
    except Exception as _col_err:
        db.session.rollback()
        print(f"[DB] Could not ensure candidates.industry_types: {_col_err}")
    try:
        if not _column_exists("candidates", "bucklist_days_in_system"):
            db.session.execute(text("ALTER TABLE candidates ADD COLUMN bucklist_days_in_system INTEGER"))
            db.session.commit()
            print("[DB] Added candidates.bucklist_days_in_system column.")
    except Exception as _be:
        db.session.rollback()
        print(f"[DB] Could not ensure candidates.bucklist_days_in_system: {_be}")
    try:
        if not _column_exists("candidates", "enrollment_status"):
            db.session.execute(text("ALTER TABLE candidates ADD COLUMN enrollment_status VARCHAR(40)"))
            db.session.commit()
            print("[DB] Added candidates.enrollment_status column.")
    except Exception as _es:
        db.session.rollback()
        print(f"[DB] Could not ensure candidates.enrollment_status: {_es}")
    try:
        if not _column_exists("candidates", "scheduler_automation_at"):
            db.session.execute(text("ALTER TABLE candidates ADD COLUMN scheduler_automation_at DATETIME"))
            db.session.commit()
            print("[DB] Added candidates.scheduler_automation_at column.")
    except Exception as _sch:
        db.session.rollback()
        print(f"[DB] Could not ensure candidates.scheduler_automation_at: {_sch}")
    
    # Cleanup stale runs on startup
    try:
        stale_runs = SendRun.query.filter(SendRun.status.in_(["running", "paused_network"])).all()
        for r in stale_runs:
            r.status = "stopped"
            r.ended_at = datetime.utcnow()
            db.session.add(r)
        if stale_runs:
            db.session.commit()
            print(f"[Startup] Cleaned up {len(stale_runs)} stale run(s).")
    except Exception as e:
        print(f"[Startup] Error cleaning up stale runs: {e}")
        db.session.rollback()

@app.after_request
def add_no_cache(resp):
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


@app.errorhandler(404)
def handle_http_404(e):
    """API clients expect JSON — Flask’s default 404 is HTML and breaks dashboard.js apiJson()."""
    if request.path.startswith("/api/"):
        return jsonify({
            "error": "Not found",
            "path": request.path,
            "hint": "Restart the Flask server (python server.py) if you just updated the app.",
        }), 404
    return (
        "<!doctype html><title>404</title><h1>Page not found</h1>",
        404,
        {"Content-Type": "text/html; charset=utf-8"},
    )


def _json_response(obj, status=200):
    """Emit JSON with explicit json.dumps so all keys (e.g. bucklist_days_in_system) are never dropped."""
    return Response(
        json.dumps(obj, ensure_ascii=False, default=str),
        status=status,
        mimetype="application/json; charset=utf-8",
    )


# Pages
@app.route("/login")
def login_page():
    if "user_id" in session:
        return redirect(url_for("dashboard"))
    return render_template("login.html")

@app.get("/")
def dashboard():
    if "user_id" not in session:
        return redirect(url_for("login_page"))
    return render_template("dashboard.html")


# AUTH API
@app.post("/api/auth/register")
def api_register():
    data = request.json or {}
    username = data.get("username", "").strip()
    email = data.get("email", "").strip()
    password = data.get("password", "")

    if not username or not email or not password:
        return jsonify({"error": "All fields are required"}), 400
    
    if User.query.filter_by(username=username).first():
        return jsonify({"error": "Username already exists"}), 400
    if User.query.filter_by(email=email).first():
        return jsonify({"error": "Email already exists"}), 400

    user = User(username=username, email=email)
    user.set_password(password)
    user.is_verified = False
    user.is_approved = False
    user.role = "user"
    
    # If first user, make admin
    if User.query.count() == 0:
        user.role = "admin"
        user.is_approved = True

    db.session.add(user)
    db.session.commit()

    # Trigger OTP
    otp = str(random.randint(100000, 999999))
    user.otp_code = otp
    user.otp_expiry = datetime.utcnow() + timedelta(minutes=10)
    db.session.commit()

    try:
        msg = EmailMessage()
        msg["Subject"] = "Verify Your Terra Tern Account"
        msg["From"] = f"JSA Pipeline <{OTP_EMAIL}>"
        msg["To"] = user.email
        msg.set_content(f"Hello {username},\n\nYour verification code is: {otp}\n\nThis code expires in 10 minutes.")
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            server.login(OTP_EMAIL, OTP_PASS)
            server.send_message(msg)
    except Exception as e:
        print(f"[Auth] SMTP Error sending registration OTP: {e}")

    return jsonify({
        "message": "Registration successful. Please verify your email with the OTP sent.",
        "user_id": user.id,
        "requires_verification": True
    }), 201

@app.post("/api/auth/login")
def api_login():
    try:
        data = request.json or {}
        username = data.get("username", "").strip()
        password = data.get("password", "")

        user = User.query.filter_by(username=username).first()
        if not user or not user.check_password(password):
            return jsonify({"error": "Invalid credentials"}), 401
                
        if not user.is_verified:
            # Re-send OTP if not verified
            otp = str(random.randint(100000, 999999))
            user.otp_code = otp
            user.otp_expiry = datetime.now(timezone.utc) + timedelta(minutes=10)
            db.session.commit()
            
            try:
                msg = EmailMessage()
                msg["Subject"] = "Verify Your Terra Tern Account"
                msg["From"] = f"JSA Pipeline <{OTP_EMAIL}>"
                msg["To"] = user.email
                msg.set_content(f"Your verification code is: {otp}\nExpires in 10 minutes.")
                
                with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
                    server.starttls()
                    server.login(OTP_EMAIL, OTP_PASS)
                    server.send_message(msg)
                return jsonify({"status": "unverified", "message": "Verification code sent to your email."}), 200
            except Exception as email_err:
                print(f"[Login] SMTP Error: {email_err}")
                return jsonify({"status": "unverified", "message": "Account exists but failed to send verification email."}), 200

        if not user.is_approved:
            return jsonify({"error": "Your account is pending approval by an administrator."}), 403

        session["user_id"] = user.id
        session["username"] = user.username
        session["role"] = user.role
        return jsonify({"message": "Login successful", "role": user.role}), 200
    except Exception as e:
        print(f"[Login] Critical Error: {e}")
        return jsonify({"error": f"Internal Server Error: {str(e)}"}), 500
                server.send_message(msg)
        except Exception as e:
            print(f"[Auth] SMTP Error re-sending verification OTP: {e}")
            return jsonify({"error": "Failed to re-send verification OTP. Check server logs."}), 500

        return jsonify({
            "error": "Email not verified. A new OTP has been sent.",
            "requires_verification": True,
            "user_id": user.id
        }), 403

    if not user.is_approved:
        return jsonify({"error": "Account pending admin approval. Please wait for an administrator to approve your access."}), 403

    # Generate OTP for login
    otp = str(random.randint(100000, 999999))
    user.otp_code = otp
    user.otp_expiry = datetime.utcnow() + timedelta(minutes=10)
    db.session.commit()

    # Send OTP Email
    try:
        msg = EmailMessage()
        msg["Subject"] = f"Your Login OTP - JSA Pipeline"
        msg["From"] = f"JSA Pipeline <{OTP_EMAIL}>"
        msg["To"] = user.email
        msg.set_content(f"Your OTP for login is: {otp}\nExpires in 10 minutes.")
        
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            server.login(OTP_EMAIL, OTP_PASS)
            server.send_message(msg)
            
        return jsonify({"message": "OTP sent to your email", "user_id": user.id}), 200
    except Exception as e:
        print(f"[Auth] SMTP Error: {e}")
        return jsonify({"error": "Failed to send OTP. Check server logs."}), 500

@app.post("/api/auth/verify-otp")
def api_verify_otp():
    data = request.json or {}
    user_id = data.get("user_id")
    otp = data.get("otp", "").strip()

    user = db.session.get(User, user_id)
    if not user or user.otp_code != otp:
        return jsonify({"error": "Invalid OTP"}), 401

    if user.otp_expiry < datetime.utcnow():
        return jsonify({"error": "OTP expired"}), 401

    # Success
    # Mark as verified if it wasn't already
    user.is_verified = True
    
    user.otp_code = None
    user.otp_expiry = None
    db.session.commit()
    
    # Check if approved before logging in
    if not user.is_approved:
        return jsonify({"message": "Email verified successfully! Your account is now pending admin approval."}), 200

    session['user_id'] = user.id
    session['username'] = user.username
    session['role'] = user.role
    return jsonify({"message": "Login successful", "redirect": url_for("dashboard")}), 200

@app.route("/api/auth/logout")
def api_logout():
    session.clear()
    return redirect(url_for("login_page"))

# --- ADMIN API ---

@app.route("/api/admin/pending-users")
def api_admin_pending_users():
    if session.get('role') != 'admin':
        return jsonify({"error": "Unauthorized"}), 403
    
    users = User.query.filter_by(is_verified=True, is_approved=False).all()
    return jsonify([{
        "id": u.id,
        "username": u.username,
        "email": u.email,
        "created_at": u.created_at.strftime("%Y-%m-%d %H:%M:%S")
    } for u in users])


@app.route("/api/admin/users")
def api_admin_all_users():
    """All CRM accounts (for admin overview)."""
    if session.get('role') != 'admin':
        return jsonify({"error": "Unauthorized"}), 403
    users = User.query.order_by(User.created_at.desc()).all()
    out = []
    for u in users:
        out.append({
            "id": u.id,
            "username": u.username,
            "email": u.email,
            "role": (u.role or "user"),
            "is_verified": bool(u.is_verified),
            "is_approved": bool(u.is_approved),
            "created_at": u.created_at.strftime("%Y-%m-%d %H:%M:%S") if u.created_at else "",
        })
    return jsonify(out)

@app.route("/api/admin/approve-user/<int:user_id>", methods=["POST"])
def api_admin_approve_user(user_id):
    if session.get('role') != 'admin':
        return jsonify({"error": "Unauthorized"}), 403
    
    user = User.query.get(user_id)
    if not user:
        return jsonify({"error": "User not found"}), 444
    
    user.is_approved = True
    db.session.commit()
    
    # Optionally send approval email
    try:
        msg = EmailMessage()
        msg["Subject"] = "Account Approved - Terra Tern"
        msg["From"] = f"JSA Admin <{OTP_EMAIL}>"
        msg["To"] = user.email
        msg.set_content(f"Hello {user.username},\n\nYour account has been approved by an administrator. You can now log in to the CRM.")
        with smtplib.SMTP(SMTP_HOST, SMTP_PORT) as server:
            server.starttls()
            server.login(OTP_EMAIL, OTP_PASS)
            server.send_message(msg)
    except Exception as e:
        print(f"[Admin] SMTP Error sending approval email: {e}")

    return jsonify({"message": f"User {user.username} approved successfully."})

@app.route("/api/admin/reject-user/<int:user_id>", methods=["POST"])
def api_admin_reject_user(user_id):
    if session.get('role') != 'admin':
        return jsonify({"error": "Unauthorized"}), 403
    
    user = User.query.get(user_id)
    if not user:
        return jsonify({"error": "User not found"}), 444
    
    username = user.username
    db.session.delete(user)
    db.session.commit()
    
    return jsonify({"message": f"User {username} rejected and removed."})

@app.get("/api/auth/me")
def api_auth_me():
    if "user_id" not in session:
        return jsonify({"authenticated": False}), 401
    
    user = db.session.get(User, session["user_id"])
    if not user:
        return jsonify({"authenticated": False}), 401
        
    allowed = None
    if user.allowed_features:
        try:
            allowed = json.loads(user.allowed_features)
        except Exception:
            pass
            
    # Keep session role updated
    session["role"] = user.role
            
    return jsonify(
        {
            "authenticated": True,
            "username": user.username,
            "user_id": user.id,
            "role": user.role or "user",
            "allowed_features": allowed
        }
    )


def _admin_role_user_count() -> int:
    """Users with admin role (case-insensitive)."""
    return User.query.filter(db.func.lower(User.role) == "admin").count()


@app.post("/api/admin/remove-user/<int:user_id>")
def api_admin_remove_user(user_id: int):
    """Permanently delete a CRM account (admin only)."""
    if session.get("role") != "admin":
        return jsonify({"error": "Unauthorized"}), 403
    sid = session.get("user_id")
    if sid is not None and int(sid) == int(user_id):
        return jsonify({"error": "You cannot remove your own account."}), 400
    user = User.query.get(user_id)
    if not user:
        return jsonify({"error": "User not found"}), 404
    if (user.role or "").strip().lower() == "admin" and _admin_role_user_count() <= 1:
        return jsonify({"error": "Cannot remove the last admin account."}), 400
    username = user.username
    try:
        db.session.delete(user)
        db.session.commit()
    except Exception as ex:
        db.session.rollback()
        return jsonify({"error": str(ex) or "Delete failed"}), 500
    return jsonify({"message": f"User {username} removed."})


# ---- RBAC: Role & Feature Management ----

@app.post("/api/admin/set-role/<int:user_id>")
def api_admin_set_role(user_id: int):
    """Change a user's role (admin only). Roles: admin, manager, user."""
    if session.get("role") != "admin":
        return jsonify({"error": "Unauthorized — admin only"}), 403
    data = request.json or {}
    new_role = (data.get("role") or "").strip().lower()
    if new_role not in ("admin", "manager", "user"):
        return jsonify({"error": "Invalid role. Must be admin, manager, or user."}), 400
    user = db.session.get(User, user_id)
    if not user:
        return jsonify({"error": "User not found"}), 404
    # Prevent demoting the last admin
    if (user.role or "").lower() == "admin" and new_role != "admin" and _admin_role_user_count() <= 1:
        return jsonify({"error": "Cannot demote the last admin."}), 400
    user.role = new_role
    db.session.commit()
    return jsonify({"message": f"User {user.username} role set to {new_role}."})


@app.post("/api/admin/set-features/<int:user_id>")
def api_admin_set_features(user_id: int):
    """Set which sidebar features a user can see (admin only). Payload: {features: [...tab IDs]}."""
    if session.get("role") != "admin":
        return jsonify({"error": "Unauthorized — admin only"}), 403
    data = request.json or {}
    features = data.get("features")
    user = db.session.get(User, user_id)
    if not user:
        return jsonify({"error": "User not found"}), 404
    if features is None or features == []:
        user.allowed_features = None  # null = all features
    else:
        user.allowed_features = json.dumps(features)
    db.session.commit()
    return jsonify({"message": f"Features updated for {user.username}."})


@app.get("/api/auth/my-permissions")
def api_auth_my_permissions():
    """Returns current user's role and allowed features."""
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 401
    user = db.session.get(User, session["user_id"])
    if not user:
        return jsonify({"error": "User not found"}), 404
    allowed = None
    if user.allowed_features:
        try:
            allowed = json.loads(user.allowed_features)
        except Exception:
            allowed = None
    return jsonify({
        "role": user.role or "user",
        "username": user.username,
        "allowed_features": allowed,  # null means all
    })


# ---- Approval Workflow ----

@app.get("/api/admin/pending-changes")
def api_admin_pending_changes():
    """List pending data changes (admin/manager only)."""
    role = session.get("role", "user")
    if role not in ("admin", "manager"):
        return jsonify({"error": "Unauthorized"}), 403
    changes = PendingChange.query.filter_by(status="pending").order_by(PendingChange.created_at.desc()).all()
    return jsonify([c.to_dict() for c in changes])


@app.get("/api/admin/all-changes")
def api_admin_all_changes():
    """All changes history (admin/manager only)."""
    role = session.get("role", "user")
    if role not in ("admin", "manager"):
        return jsonify({"error": "Unauthorized"}), 403
    changes = PendingChange.query.order_by(PendingChange.created_at.desc()).limit(200).all()
    return jsonify([c.to_dict() for c in changes])


@app.post("/api/admin/approve-change/<int:change_id>")
def api_admin_approve_change(change_id: int):
    """Approve a pending change — applies the data to the candidate."""
    role = session.get("role", "user")
    if role not in ("admin", "manager"):
        return jsonify({"error": "Unauthorized"}), 403
    pc = db.session.get(PendingChange, change_id)
    if not pc:
        return jsonify({"error": "Change not found"}), 404
    if pc.status != "pending":
        return jsonify({"error": f"Change already {pc.status}."}), 400
    # Apply the change to the candidate
    candidate = db.session.get(Candidate, pc.candidate_id)
    if not candidate:
        pc.status = "rejected"
        pc.review_note = "Candidate no longer exists"
        pc.reviewed_by = session.get("user_id")
        pc.reviewed_at = datetime.utcnow()
        db.session.commit()
        return jsonify({"error": "Candidate no longer exists. Change rejected."}), 400
    try:
        changes = json.loads(pc.change_data)
        _ALLOWED_FIELDS = {
            "name", "email", "pa_member", "rm_member", "placement_officer_member",
            "roles_text", "subject_template", "message_template", "enrollment_id",
            "enrollment_status", "industry_types", "smart_country", "smart_industry",
            "bucklist_days_in_system",
        }
        for key, val in changes.items():
            if key in _ALLOWED_FIELDS:
                setattr(candidate, key, val)
        pc.status = "approved"
        pc.reviewed_by = session.get("user_id")
        pc.reviewed_at = datetime.utcnow()
        db.session.commit()
        return jsonify({"message": f"Change approved and applied to {candidate.name}."})
    except Exception as ex:
        db.session.rollback()
        return jsonify({"error": f"Failed to apply change: {ex}"}), 500


@app.post("/api/admin/reject-change/<int:change_id>")
def api_admin_reject_change(change_id: int):
    """Reject a pending change."""
    role = session.get("role", "user")
    if role not in ("admin", "manager"):
        return jsonify({"error": "Unauthorized"}), 403
    data = request.json or {}
    pc = db.session.get(PendingChange, change_id)
    if not pc:
        return jsonify({"error": "Change not found"}), 404
    if pc.status != "pending":
        return jsonify({"error": f"Change already {pc.status}."}), 400
    pc.status = "rejected"
    pc.reviewed_by = session.get("user_id")
    pc.reviewed_at = datetime.utcnow()
    pc.review_note = (data.get("note") or "").strip() or "Rejected by admin"
    db.session.commit()
    return jsonify({"message": "Change rejected."})



@app.get("/__health__")
def health():
    return jsonify({
        "app": APP_NAME,
        "status": "ok",
        "port": 8080,
        "build": "smart-automation-v2",
        "features": {
            "smart_automation_dashboard": True,
            "api_paths": ["/api/smart-automation/dashboard", "/api/sa/dashboard"],
        },
    })

# API Candidates
@app.get("/api/candidates")
def api_list_candidates():
    q = Candidate.query.order_by(Candidate.name.asc()).all()
    if not q:
        return _json_response([])
    cids = [c.id for c in q]
    ws_map = {}
    for ws in Workspace.query.filter(Workspace.candidate_id.in_(cids)).order_by(Workspace.id.desc()).all():
        if ws.candidate_id not in ws_map:
            ws_map[ws.candidate_id] = (ws.country or "").strip()
    plan_starts, ws_starts = candidate_service_start_maps_for_ids(cids)
    return _json_response(
        [
            c.to_dict_summary(
                workspace_country=ws_map.get(c.id),
                workflow_service_start=plan_starts.get(c.id),
                workspace_service_start=ws_starts.get(c.id),
            )
            for c in q
        ]
    )


def _bucklist_header_map(header_row):
    """Map name, email, days column indices from first row."""
    m = {}
    if not header_row:
        return m
    for i, cell in enumerate(header_row):
        if cell is None:
            continue
        s = str(cell).strip().lower()
        if "email" in s or "e-mail" in s:
            m["email"] = i
        elif "day" in s:
            m["days"] = i
    if "name" not in m:
        for i, cell in enumerate(header_row):
            if cell is None:
                continue
            s = str(cell).strip().lower()
            if "name" in s and "email" not in s:
                m["name"] = i
                break
    if "name" not in m and len(header_row) >= 1:
        m["name"] = 0
    if "email" not in m and len(header_row) >= 2:
        m["email"] = 1
    if "days" not in m and len(header_row) >= 3:
        m["days"] = 2
    return m


@app.get("/api/bucklist")
def api_bucklist():
    """Candidates grouped by days-in-system buckets (Bucklist grid)."""
    q = Candidate.query.order_by(Candidate.name.asc()).all()
    cids = [c.id for c in q]
    plan_starts, ws_starts = candidate_service_start_maps_for_ids(cids)
    buckets = {k: [] for k in BUCKLIST_BUCKET_KEYS}
    unassigned = []
    for c in q:
        wf_s, ws_s = plan_starts.get(c.id), ws_starts.get(c.id)
        try:
            bdis = int(c.bucklist_days_in_system) if c.bucklist_days_in_system is not None else None
        except (TypeError, ValueError):
            bdis = None
        try:
            days = candidate_days_for_bucklist(
                c, workflow_service_start=wf_s, workspace_service_start=ws_s
            )
        except Exception:
            days = None
        raw_src = None
        try:
            raw_src = candidate_days_in_system_source(
                c, workflow_service_start=wf_s, workspace_service_start=ws_s
            )
        except Exception:
            raw_src = None
        ds_src = {
            "profile_service_start": "service_start",
            "workflow_plan": "workflow_plan",
            "workspace_service": "workspace_service",
            "elapsed": "service_start",
            "bucklist_pin": "bucklist",
            "created_age": "estimated",
        }.get(raw_src or "", raw_src or "estimated")
        try:
            ct = format_candidate_country_type(c.smart_country, None)
        except Exception:
            ct = ""
        try:
            sb = int(c.smart_baseline_applied or 0)
        except (TypeError, ValueError):
            sb = 0
        try:
            entry = {
                "id": c.id,
                "name": c.name or "",
                "email": c.email or "",
                "days": days,
                "bucklist_days_in_system": bdis,
                "country_type": ct,
                "industry_types": (c.industry_types or "").strip(),
                "updated_at": c.updated_at.isoformat() if c.updated_at else "",
                "days_source": ds_src,
                "smart_baseline_applied": sb,
                "smart_country": (c.smart_country or "").strip(),
                "smart_industry": (c.smart_industry or "").strip(),
            }
        except Exception:
            entry = {
                "id": c.id,
                "name": getattr(c, "name", None) or "",
                "email": getattr(c, "email", None) or "",
                "days": days,
                "bucklist_days_in_system": bdis,
                "country_type": "",
                "industry_types": "",
                "updated_at": "",
                "days_source": ds_src or "estimated",
                "smart_baseline_applied": sb,
                "smart_country": "",
                "smart_industry": "",
            }
        if days is None:
            unassigned.append(entry)
            continue
        key = bucket_key_for_days(days)
        if key in buckets:
            buckets[key].append(entry)
        else:
            unassigned.append(entry)
    return jsonify({"buckets": buckets, "unassigned": unassigned})


@app.post("/api/bucklist/assign", strict_slashes=False)
@app.post("/api/bucklist-assign", strict_slashes=False)
def api_bucklist_assign():
    """Set bucklist_days_in_system for an existing candidate (Not assigned → pick range + days)."""
    data = request.json or {}
    try:
        cid = int(data.get("candidate_id"))
    except (TypeError, ValueError):
        return jsonify({"error": "candidate_id is required"}), 400
    category = (data.get("category") or "").strip()
    if category not in BUCKLIST_CATEGORY_BOUNDS:
        return jsonify({"error": "Invalid or missing category"}), 400
    lo, hi = BUCKLIST_CATEGORY_BOUNDS[category]
    days_raw = data.get("days_in_system")
    if days_raw is None or (isinstance(days_raw, str) and not str(days_raw).strip()):
        if hi > 100000:
            days_val = 200
        else:
            days_val = max(lo, min((lo + hi) // 2, hi))
    else:
        try:
            days_val = int(days_raw)
        except (TypeError, ValueError):
            return jsonify({"error": "days_in_system must be a whole number"}), 400
    if not (lo <= days_val <= hi):
        return jsonify(
            {"error": f"Days must be between {lo} and {hi} for category {category}."}
        ), 400
    cand = Candidate.query.get(cid)
    if not cand:
        return jsonify({"error": "Candidate not found"}), 404
    cand.bucklist_days_in_system = days_val
    db.session.commit()
    return jsonify({"message": "Assigned", "candidate_id": cand.id}), 200


@app.post("/api/bucklist/candidate")
def api_bucklist_save_candidate_profile():
    """Same fields as POST /api/candidates (multipart) plus category + days_in_system for Bucklist bucket."""
    category = (request.form.get("category") or "").strip()
    if category not in BUCKLIST_CATEGORY_BOUNDS:
        return jsonify({"error": "Invalid or missing category"}), 400
    try:
        days_val = int(request.form.get("days_in_system") or 0)
    except (TypeError, ValueError):
        return jsonify({"error": "days_in_system must be a whole number"}), 400
    lo, hi = BUCKLIST_CATEGORY_BOUNDS[category]
    if not (lo <= days_val <= hi):
        return jsonify(
            {"error": f"Days in system must be between {lo} and {hi} for category {category}."}
        ), 400

    name = (request.form.get("name") or "").strip()
    email_addr = (request.form.get("email") or "").strip()
    pa_member = (request.form.get("pa_member") or request.form.get("app_code") or "").strip()
    rm_member = (request.form.get("rm_member") or "").strip()
    placement_officer_member = (request.form.get("placement_officer_member") or "").strip()
    app_password = (request.form.get("app_password") or "").strip()
    subject_template = request.form.get("subject_template") or ""
    message_template = request.form.get("message_template") or ""
    roles_text = request.form.get("roles_text") or ""
    enrollment_id = (request.form.get("enrollment_id") or "").strip()
    industry_types = (
        (request.form.get("industry_types") or request.form.get("industryTypes") or "").strip()
    )
    smart_country = (
        request.form.get("smart_country")
        or request.form.get("bucklist_smart_country")
        or request.form.get("country_type")
        or ""
    ).strip()

    if not name or not email_addr:
        return jsonify({"error": "name and email are required"}), 400
    email_addr = email_addr.lower()

    resume_file = request.files.get("resume")
    cover_file = request.files.get("coverLetter")

    existing = Candidate.query.filter(func.lower(Candidate.email) == email_addr).first()
    if existing:
        existing.name = name
        existing.bucklist_days_in_system = days_val
        existing.pa_member = pa_member
        existing.app_code = pa_member
        existing.rm_member = rm_member
        existing.placement_officer_member = placement_officer_member
        existing.app_password = app_password
        existing.subject_template = subject_template
        existing.message_template = message_template
        existing.roles_text = roles_text
        existing.enrollment_id = enrollment_id
        existing.industry_types = industry_types if industry_types else None
        existing.smart_country = smart_country if smart_country else None
        folder = safe_candidate_folder(existing.id)
        if resume_file and getattr(resume_file, "filename", None):
            existing.resume_path = save_uploaded_file(resume_file, folder, "resume.pdf")
        if cover_file and getattr(cover_file, "filename", None):
            existing.cover_letter_path = save_uploaded_file(cover_file, folder, "cover_letter.pdf")
        db.session.commit()
        return jsonify({"message": "Candidate updated", "candidate_id": existing.id}), 200

    cand = Candidate(
        name=name,
        email=email_addr,
        pa_member=pa_member,
        rm_member=rm_member,
        placement_officer_member=placement_officer_member,
        app_code=pa_member,
        app_password=app_password,
        subject_template=subject_template,
        message_template=message_template,
        roles_text=roles_text,
        enrollment_id=enrollment_id or None,
        industry_types=industry_types if industry_types else None,
        smart_country=smart_country if smart_country else None,
        bucklist_days_in_system=days_val,
    )
    db.session.add(cand)
    db.session.flush()
    folder = safe_candidate_folder(cand.id)
    if resume_file and getattr(resume_file, "filename", None):
        cand.resume_path = save_uploaded_file(resume_file, folder, "resume.pdf")
    if cover_file and getattr(cover_file, "filename", None):
        cand.cover_letter_path = save_uploaded_file(cover_file, folder, "cover_letter.pdf")
    db.session.commit()
    return jsonify({"message": "Candidate created", "candidate_id": cand.id}), 201


@app.get("/api/bucklist/template")
def api_bucklist_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bucklist"
    ws.append(["Name", "Email", "Days in system"])
    ws.append(["Example Candidate", "candidate@example.com", 45])
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return send_file(
        bio,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="bucklist_upload_template.xlsx",
    )


@app.post("/api/bucklist/upload")
def api_bucklist_upload():
    """Excel: Name, Email, Days in system — upserts candidates and sets bucklist_days_in_system."""
    f = request.files.get("file") or request.files.get("excel")
    if not f or not getattr(f, "filename", None):
        return jsonify({"error": "file is required"}), 400
    fn = (f.filename or "").lower()
    if not fn.endswith(".xlsx"):
        return jsonify({"error": "Upload .xlsx only (use the template)"}), 400
    wb = openpyxl.load_workbook(f.stream)
    sheet = wb.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return jsonify({"error": "Empty file"}), 400
    m = _bucklist_header_map(rows[0])
    i_name = m.get("name", 0)
    i_email = m.get("email", 1)
    i_days = m.get("days", 2)

    created = 0
    updated = 0
    skipped = 0
    for row in rows[1:]:
        if not row:
            skipped += 1
            continue
        if all(x is None or (isinstance(x, str) and not str(x).strip()) for x in row):
            continue
        name = (str(row[i_name]).strip() if len(row) > i_name and row[i_name] is not None else "") if i_name is not None else ""
        email_addr = ""
        if len(row) > i_email and row[i_email] is not None:
            email_addr = str(row[i_email]).strip().lower()
        days_raw = row[i_days] if len(row) > i_days else None
        try:
            if days_raw is None or str(days_raw).strip() == "":
                days_val = None
            else:
                days_val = int(float(days_raw))
        except (ValueError, TypeError):
            days_val = None
        if not name or not email_addr:
            skipped += 1
            continue
        if days_val is None:
            skipped += 1
            continue
        days_val = max(0, days_val)

        existing = Candidate.query.filter(func.lower(Candidate.email) == email_addr).first()
        if existing:
            existing.name = name
            existing.bucklist_days_in_system = days_val
            updated += 1
        else:
            cand = Candidate(
                name=name,
                email=email_addr,
                pa_member="",
                app_code="",
                bucklist_days_in_system=days_val,
            )
            db.session.add(cand)
            created += 1
    db.session.commit()
    return jsonify(
        {
            "message": "Bucklist import complete",
            "created": created,
            "updated": updated,
            "skipped": skipped,
        }
    ), 200


@app.get("/api/candidates/<int:candidate_id>")
def api_get_candidate(candidate_id):
    c = Candidate.query.get_or_404(candidate_id)
    return _json_response(c.to_dict_detail())


@app.get("/api/candidates/<int:candidate_id>/download")
def api_candidate_download_files(candidate_id):
    """ZIP of resume + cover letter (used by 6-Month workflow Report button). Always returns a valid ZIP."""
    c = Candidate.query.get_or_404(candidate_id)
    base = _safe_zip_name(c.name or f"candidate_{c.id}")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    mem = io.BytesIO()
    n = 0
    with zipfile.ZipFile(mem, mode="w", compression=zipfile.ZIP_DEFLATED) as z:
        if c.resume_path and os.path.exists(c.resume_path):
            z.write(c.resume_path, f"resume_{base}.pdf")
            n += 1
        if c.cover_letter_path and os.path.exists(c.cover_letter_path):
            z.write(c.cover_letter_path, f"cover_letter_{base}.pdf")
            n += 1
        if n == 0:
            z.writestr(
                "README.txt",
                "No resume or cover letter PDF was found for this candidate.\n"
                "Upload both files on the candidate form, then download again.\n",
            )
    mem.seek(0)
    return send_file(
        mem,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"candidate_{c.id}_{base}_{ts}.zip",
    )


@app.post("/api/candidates")
def api_save_candidate():
    cid_raw = request.form.get("id")
    name = (request.form.get("name") or "").strip()
    email_addr = (request.form.get("email") or "").strip()
    # UI uses pa_member; accept legacy app_code too.
    pa_member = (request.form.get("pa_member") or request.form.get("app_code") or "").strip()
    rm_member = (request.form.get("rm_member") or "").strip()
    placement_officer_member = (request.form.get("placement_officer_member") or "").strip()
    app_password = (request.form.get("app_password") or "").strip()
    subject_template = request.form.get("subject_template") or ""
    message_template = request.form.get("message_template") or ""
    roles_text = request.form.get("roles_text") or ""
    enrollment_id = (request.form.get("enrollment_id") or "").strip()
    enrollment_status = normalize_enrollment_status(request.form.get("enrollment_status"))
    industry_types = (
        (request.form.get("industry_types") or request.form.get("industryTypes") or "").strip()
    )

    if not name or not email_addr:
        return jsonify({"error": "name and email are required"}), 400

    if cid_raw:
        cand = Candidate.query.get(int(cid_raw))
        if not cand:
            return jsonify({"error": "Candidate not found"}), 404
    else:
        # Prevent staff from creating new candidates directly
        if session.get("role") == "user":
            return jsonify({"error": "Staff cannot create new candidates. Please contact an admin."}), 403
        cand = Candidate()
        db.session.add(cand)

    if session.get("role") == "user":
        # Staff edits require approval
        changes = {
            "name": name,
            "email": email_addr,
            "pa_member": pa_member,
            "rm_member": rm_member,
            "placement_officer_member": placement_officer_member,
            "app_password": app_password,
            "subject_template": subject_template,
            "message_template": message_template,
            "roles_text": roles_text,
            "enrollment_id": enrollment_id,
            "enrollment_status": enrollment_status,
            "industry_types": industry_types,
        }
        pc = PendingChange(
            user_id=session.get("user_id"),
            candidate_id=cand.id,
            change_data=json.dumps(changes)
        )
        db.session.add(pc)
        db.session.commit()
        return jsonify({"message": "Change submitted for admin approval.", "pending": True})

    cand.name = name
    cand.email = email_addr
    cand.pa_member = pa_member
    cand.rm_member = rm_member
    cand.placement_officer_member = placement_officer_member
    cand.app_code = pa_member
    cand.app_password = app_password
    cand.subject_template = subject_template
    cand.message_template = message_template
    cand.roles_text = roles_text
    cand.enrollment_id = enrollment_id
    cand.enrollment_status = enrollment_status
    cand.industry_types = industry_types if industry_types else None

    if not cid_raw:
        db.session.flush()

    # Update scheduled_time if provided
    st_raw = request.form.get("scheduled_time")
    if st_raw:
        try:
            # datetime-local format: YYYY-MM-DDTHH:MM
            cand.scheduled_time = datetime.strptime(st_raw, "%Y-%m-%dT%H:%M")
        except Exception:
            cand.scheduled_time = None
    else:
        cand.scheduled_time = None

    folder = safe_candidate_folder(cand.id)
    resume_file = request.files.get("resume")
    cover_file = request.files.get("coverLetter")

    if resume_file:
        cand.resume_path = save_uploaded_file(resume_file, folder, "resume.pdf")
    if cover_file:
        cand.cover_letter_path = save_uploaded_file(cover_file, folder, "cover_letter.pdf")

    # Country + days: dashboard prefers POST .../profile-fields (JSON). Also accept multipart so saves
    # work if the server was not restarted after that route was added, or if a proxy drops JSON.
    if "smart_country" in request.form or "country_type" in request.form or "bucklist_smart_country" in request.form:
        sc = (
            (request.form.get("smart_country") or request.form.get("country_type") or request.form.get("bucklist_smart_country") or "").strip()
        )
        cand.smart_country = sc if sc else None
    if "smart_industry" in request.form or "bucklist_smart_industry" in request.form:
        si = (
            (request.form.get("smart_industry") or request.form.get("bucklist_smart_industry") or "").strip()
        )
        cand.smart_industry = si if si else None
    if "bucklist_days_in_system" in request.form:
        bd_raw = request.form.get("bucklist_days_in_system")
        if bd_raw is None or (isinstance(bd_raw, str) and not str(bd_raw).strip()):
            cand.bucklist_days_in_system = None
        else:
            try:
                cand.bucklist_days_in_system = int(float(bd_raw))
            except (TypeError, ValueError):
                return jsonify({"error": "bucklist_days_in_system must be a number or empty"}), 400

    sync_workflow_plan_hr_from_candidate(cand)

    db.session.commit()
    return _json_response({"message": "Saved", "candidate": cand.to_dict_detail()}), 200


@app.post("/api/candidates/<int:candidate_id>/profile-fields", strict_slashes=False)
def api_candidate_profile_fields(candidate_id):
    """Authoritative save for HR targets (smart_country, smart_industry) and days in system (bucklist_days_in_system)."""
    cand = db.session.get(Candidate, candidate_id)
    if not cand:
        return jsonify({"error": "Candidate not found"}), 404
    data = request.get_json(force=True, silent=True) or {}
    
    if session.get("role") == "user":
        # Intercept and store as PendingChange
        changes = {}
        if "smart_country" in data:
            sc = (data.get("smart_country") or "").strip()
            changes["smart_country"] = sc if sc else None
        if "smart_industry" in data:
            si = (data.get("smart_industry") or "").strip()
            changes["smart_industry"] = si if si else None
        if "bucklist_days_in_system" in data:
            bd = data.get("bucklist_days_in_system")
            if bd is None:
                changes["bucklist_days_in_system"] = None
            else:
                try:
                    changes["bucklist_days_in_system"] = int(float(bd))
                except (TypeError, ValueError):
                    return jsonify({"error": "bucklist_days_in_system must be a number or null"}), 400
        if not changes:
            return jsonify({"message": "No changes submitted"}), 200
            
        pc = PendingChange(
            user_id=session.get("user_id"),
            candidate_id=cand.id,
            change_data=json.dumps(changes)
        )
        db.session.add(pc)
        db.session.commit()
        return jsonify({"message": "Change submitted for admin approval.", "pending": True})
        
    if "smart_country" in data:
        sc = (data.get("smart_country") or "").strip()
        cand.smart_country = sc if sc else None
    if "smart_industry" in data:
        si = (data.get("smart_industry") or "").strip()
        cand.smart_industry = si if si else None
    if "bucklist_days_in_system" in data:
        bd = data.get("bucklist_days_in_system")
        if bd is None:
            cand.bucklist_days_in_system = None
        else:
            try:
                cand.bucklist_days_in_system = int(float(bd))
            except (TypeError, ValueError):
                return jsonify({"error": "bucklist_days_in_system must be a number or null"}), 400
    if "smart_country" in data or "smart_industry" in data:
        sync_workflow_plan_hr_from_candidate(cand)
    db.session.commit()
    return _json_response({"message": "ok", "candidate": cand.to_dict_detail()}), 200


@app.delete("/api/candidates/<int:candidate_id>")
def api_delete_candidate(candidate_id):
    c = Candidate.query.get_or_404(candidate_id)
    db.session.delete(c)
    db.session.commit()
    return jsonify({"message": "Deleted"}), 200

@app.post("/api/candidates/import")
def api_import_candidates_xlsx():
    f = request.files.get("excel")
    if not f:
        return jsonify({"error": "excel file is required"}), 400

    wb = openpyxl.load_workbook(f.stream)
    sheet = wb.active

    created = 0
    updated = 0
    # Supported Excel layouts:
    # NEW: A Name, B Email, C PA Member, D Placement Officer Member, E App Password, F Subject, G Message, H Roles
    # OLD: A Name, B Email, C App Code, D App Password, E Subject, F Message, G Roles
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = (str(row[0]).strip() if row and row[0] else "")
        email_addr = (str(row[1]).strip() if len(row) > 1 and row[1] else "")

        pa_member = ""
        placement_officer_member = ""
        app_password = ""
        subject_template = ""
        message_template = ""
        roles_text = ""
        industry = "Default"

        if len(row) >= 9:
            pa_member = (str(row[2]).strip() if row[2] else "")
            placement_officer_member = (str(row[3]).strip() if row[3] else "")
            app_password = (str(row[4]).strip() if row[4] else "")
            subject_template = (str(row[5]) if row[5] else "")
            message_template = (str(row[6]) if row[6] else "")
            roles_text = (str(row[7]) if row[7] else "")
            industry = (str(row[8]).strip() if row[8] else "Default")
        else:
            # backward-compatible
            pa_member = (str(row[2]).strip() if len(row) > 2 and row[2] else "")
            app_password = (str(row[3]).strip() if len(row) > 3 and row[3] else "")
            subject_template = (str(row[4]) if len(row) > 4 and row[4] else "")
            message_template = (str(row[5]) if len(row) > 5 and row[5] else "")
            roles_text = (str(row[6]) if len(row) > 6 and row[6] else "")

        rm_member = ""
        if len(row) > 9 and row[9] is not None:
            rm_member = str(row[9]).strip()

        scheduled_time = None
        if len(row) >= 9:
            st_raw = str(row[8]).strip() if row[8] else ""
            if st_raw:
                try:
                    # Expecting ISO or something parseable
                    scheduled_time = datetime.fromisoformat(st_raw.replace("Z", "+00:00"))
                except Exception:
                    try:
                        scheduled_time = datetime.strptime(st_raw, "%Y-%m-%d %H:%M:%S")
                    except Exception:
                        scheduled_time = None

        if not name or not email_addr:
            continue

        existing = Candidate.query.filter(func.lower(Candidate.email) == email_addr.lower()).first()
        if existing:
            existing.name = name
            existing.pa_member = pa_member
            existing.placement_officer_member = placement_officer_member
            if rm_member or len(row) > 9:
                existing.rm_member = rm_member
            existing.app_code = pa_member

            if app_password:
                existing.app_password = app_password
            if subject_template:
                existing.subject_template = subject_template
            if message_template:
                existing.message_template = message_template
            if roles_text:
                existing.roles_text = roles_text
            existing.industry_types = industry
            db.session.add(existing)
            updated += 1
        else:
            cand = Candidate(
                name=name,
                email=email_addr,
                pa_member=pa_member,
                rm_member=rm_member,
                placement_officer_member=placement_officer_member,
                app_code=pa_member,
                app_password=app_password,
                subject_template=subject_template, message_template=message_template, roles_text=roles_text,
                industry_types=industry or None,
            )
            db.session.add(cand)
            db.session.commit()
            
            # Create a target if it's actually target data (this logic seems mixed, let's just update the objects)
            # Actually, Candidate and Target are separate. If this is a Target import:
            t = Target(
                company_name=name, hr_email=email_addr, industry=industry, is_valid=True
            )
            db.session.add(t)
            created += 1

    db.session.commit()
    return jsonify({"message": "Imported", "created": created, "updated": updated}), 200


def _safe_zip_name(s: str, max_len: int = 60) -> str:
    s = re.sub(r"[^A-Za-z0-9._-]+", "_", (s or "").strip())
    s = s.strip("_") or "file"
    return s[:max_len]


def api_backup_database_sqlite():
    """Full CRM database file (.sqlite3). Same naming family as /api/candidates/backup_zip."""
    db_path = CRM_SQLITE_PATH
    if not os.path.isfile(db_path):
        return jsonify({"error": "Database file not found"}), 404

    fd, tmp_path = tempfile.mkstemp(suffix=".sqlite3")
    os.close(fd)
    blob = b""
    try:
        src = sqlite3.connect(db_path, timeout=120)
        dst = sqlite3.connect(tmp_path)
        try:
            src.backup(dst)
        finally:
            dst.close()
            src.close()
        with open(tmp_path, "rb") as bf:
            blob = bf.read()
    except Exception as e:
        return jsonify({"error": f"Backup failed: {e}"}), 500
    finally:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass

    if not blob:
        return jsonify({"error": "Backup produced an empty file"}), 500

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    mem = io.BytesIO(blob)
    mem.seek(0)
    return send_file(
        mem,
        mimetype="application/vnd.sqlite3",
        as_attachment=True,
        download_name=f"crm_backup_{ts}.sqlite3",
    )


_SQLITE_BACKUP_PATHS = (
    ("api_crm_backup_sqlite", "/api/candidates/backup_sqlite"),
    ("api_crm_sqlite_backup", "/api/candidates/sqlite_backup"),
    ("api_crm_backup_database", "/api/backup/database"),
)
for _ep, _path in _SQLITE_BACKUP_PATHS:
    app.add_url_rule(
        _path,
        endpoint=_ep,
        view_func=api_backup_database_sqlite,
        methods=["GET"],
        strict_slashes=False,
    )


@app.get("/api/candidates/backup_zip")
def api_candidates_backup_zip():
    """Download a full backup of candidates as ZIP:
    - candidates_backup.xlsx (includes PA Member, App Password, templates, roles)
    - resume/ + cover_letter/ folders with PDFs (if present)
    """
    cands = Candidate.query.order_by(Candidate.id.asc()).all()
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    mem = io.BytesIO()
    with zipfile.ZipFile(mem, mode="w", compression=zipfile.ZIP_DEFLATED) as z:
        # Build Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Candidates"

        headers = [
            "Candidate Name",
            "Sender Email",
            "PA Member",
            "RM Member",
            "Gmail App Password",
            "Placement Officer Member",
            "Industry Types",
            "Roles (one per line) - rotates across Targets",
            "Subject Template",
            "Message Template",
            "Scheduled Time (ISO)",
            "Resume (PDF)",
            "Cover Letter (PDF)",
        ]
        ws.append(headers)

        for c in cands:
            pa = (c.pa_member or c.app_code or "")
            resume_ref = ""
            cover_ref = ""

            # Add PDFs into zip with stable names (avoid collisions)
            if c.resume_path and os.path.exists(c.resume_path):
                resume_ref = f"resume/candidate_{c.id}_{_safe_zip_name(c.name)}.pdf"
                try:
                    z.write(c.resume_path, resume_ref)
                except Exception:
                    resume_ref = ""

            if c.cover_letter_path and os.path.exists(c.cover_letter_path):
                cover_ref = f"cover_letter/candidate_{c.id}_{_safe_zip_name(c.name)}.pdf"
                try:
                    z.write(c.cover_letter_path, cover_ref)
                except Exception:
                    cover_ref = ""

            ws.append([
                c.name or "",
                c.email or "",
                pa,
                (c.rm_member or "").strip(),
                c.app_password or "",
                c.placement_officer_member or "",
                (c.industry_types or "").strip(),
                c.roles_text or "",
                c.subject_template or "",
                c.message_template or "",
                c.scheduled_time.isoformat() if c.scheduled_time else "",
                resume_ref,
                cover_ref,
            ])

        xbuf = io.BytesIO()
        wb.save(xbuf)
        z.writestr("candidates_backup.xlsx", xbuf.getvalue())

        z.writestr(
            "README.txt",
            "This backup contains candidate sender emails, PA, RM member, placement officer, templates, roles, and PDFs.\n"
            "Important: It also includes Gmail App Passwords (sensitive). Store this ZIP securely.\n"
        )

    mem.seek(0)
    return send_file(
        mem,
        mimetype="application/zip",
        as_attachment=True,
        download_name=f"candidates_backup_{ts}.zip"
    )

@app.post("/api/candidates/restore_backup_zip")
def api_candidates_restore_backup_zip():
    """Restore candidates from a ZIP created by /api/candidates/backup_zip.

    - Mode: merge (default) updates existing candidates by sender email, creates missing.
    - replace deletes current candidates first, then restores.
#     - Restores PDF files into uploads/candidate_{id}/resume.pdf & cover_letter.pdf
    """
    f = request.files.get("backup_zip")
    mode = (request.form.get("mode") or "merge").strip().lower()
    if not f:
        return jsonify({"error": "backup_zip file is required"}), 400

    try:
        zf = zipfile.ZipFile(f.stream)
    except Exception:
        return jsonify({"error": "Invalid ZIP file"}), 400

    # Find the Excel inside the ZIP
    excel_name = None
    names = zf.namelist()
    if "candidates_backup.xlsx" in names:
        excel_name = "candidates_backup.xlsx"
    else:
        for n in names:
            if n.lower().endswith(".xlsx"):
                excel_name = n
                break
    if not excel_name:
        return jsonify({"error": "candidates_backup.xlsx not found in ZIP"}), 400

    try:
        wb = openpyxl.load_workbook(io.BytesIO(zf.read(excel_name)))
        ws = wb.active
    except Exception:
        return jsonify({"error": "Could not read Excel from ZIP"}), 400

    def norm(h: str) -> str:
        h = (h or "").strip().lower()
        h = re.sub(r"\s+", " ", h)
        h = re.sub(r"[^a-z0-9 ]+", "", h)
        return h

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        return jsonify({"error": "Excel has no header row"}), 400

    headers = [str(x).strip() if x is not None else "" for x in header_row]
    idx = {norm(h): i for i, h in enumerate(headers) if h}
    rm_col = idx.get(norm("rm member"))

    def get(row, key, default=""):
        i = idx.get(norm(key))
        if i is None or i >= len(row):
            return default
        val = row[i]
        return "" if val is None else str(val)

    # Optional destructive replace
    if mode == "replace":
        try:
            # Delete candidates (keep targets/runs)
            db.session.query(Candidate).delete()
            db.session.commit()
        except Exception:
            db.session.rollback()
            return jsonify({"error": "Failed to clear existing candidates"}), 500

        # Clear candidate upload folders
        try:
            for name in os.listdir(UPLOAD_ROOT):
                if name.startswith("candidate_"):
                    shutil.rmtree(os.path.join(UPLOAD_ROOT, name), ignore_errors=True)
        except Exception:
            pass

    created = 0
    updated = 0
    files_restored = 0

    # Helper: safe read of a file within ZIP (no zip-slip paths)
    def safe_zip_read(member: str) -> bytes:
        member = (member or "").replace("\\", "/")
        member = member.lstrip("/")
        if not member or ".." in member.split("/"):
            raise ValueError("unsafe path")
        return zf.read(member)

    for row in ws.iter_rows(min_row=2, values_only=True):
        name = get(row, "Candidate Name").strip()
        email_addr = get(row, "Sender Email").strip().lower()
        if not name or not email_addr:
            continue

        pa_member = get(row, "PA Member").strip()
        placement_officer_member = get(row, "Placement Officer Member").strip()
        app_password = get(row, "Gmail App Password").strip()
        industry_types = get(row, "Industry Types").strip()
        roles_text = get(row, "Roles (one per line) - rotates across Targets")
        subject_template = get(row, "Subject Template")
        message_template = get(row, "Message Template")
        cover_ref = get(row, "Cover Letter (PDF)").strip()
        st_raw = get(row, "Scheduled Time (ISO)").strip()
        scheduled_time = None
        if st_raw:
            try:
                scheduled_time = datetime.fromisoformat(st_raw.replace("Z", "+00:00"))
            except Exception:
                scheduled_time = None

        cand = Candidate.query.filter_by(email=email_addr).first()
        if cand:
            cand.name = name
            cand.pa_member = pa_member
            cand.placement_officer_member = placement_officer_member
            if rm_col is not None:
                cell = row[rm_col] if rm_col < len(row) else None
                cand.rm_member = "" if cell is None else str(cell).strip()
            cand.app_password = app_password
            cand.subject_template = subject_template
            cand.message_template = message_template
            cand.roles_text = roles_text
            cand.scheduled_time = scheduled_time
            cand.industry_types = industry_types or None
            # keep legacy in sync
            cand.app_code = pa_member
            updated += 1
        else:
            create_kwargs = dict(
                name=name,
                email=email_addr,
                pa_member=pa_member,
                placement_officer_member=placement_officer_member,
                app_password=app_password,
                subject_template=subject_template,
                message_template=message_template,
                roles_text=roles_text,
                app_code=pa_member,
                scheduled_time=scheduled_time,
                industry_types=industry_types or None,
            )
            if rm_col is not None:
                cell = row[rm_col] if rm_col < len(row) else None
                create_kwargs["rm_member"] = "" if cell is None else str(cell).strip()
            cand = Candidate(**create_kwargs)
            db.session.add(cand)
            db.session.flush()  # allocate cand.id
            created += 1

        # Restore PDFs (if present in zip)
        folder = safe_candidate_folder(cand.id)
        try:
            if resume_ref:
                data = safe_zip_read(resume_ref)
                if data:
                    pth = os.path.join(folder, "resume.pdf")
                    with open(pth, "wb") as wf:
                        wf.write(data)
                    cand.resume_path = pth
                    files_restored += 1
        except Exception:
            pass

        try:
            if cover_ref:
                data = safe_zip_read(cover_ref)
                if data:
                    pth = os.path.join(folder, "cover_letter.pdf")
                    with open(pth, "wb") as wf:
                        wf.write(data)
                    cand.cover_letter_path = pth
                    files_restored += 1
        except Exception:
            pass

    try:
        db.session.commit()
    except Exception:
        db.session.rollback()
        return jsonify({"error": "Restore failed while saving to database"}), 500

    return jsonify({"message": "Restored", "created": created, "updated": updated, "files_restored": files_restored}), 200



# API Targets
@app.get("/api/targets/summary")
def api_targets_summary():
    total = Target.query.count()
    invalid = Target.query.filter(Target.is_valid == False).count()
    return jsonify({"total": total, "invalid": invalid})

@app.post("/api/candidates/clear")
def api_clear_candidates():
    """Remove all candidates and dependent rows (workspaces, apps, events, plans, reports)."""
    try:
        EmailEvent.query.delete()
        JobApplication.query.delete()
        SentHistory.query.delete()
        Workspace.query.delete()
        WorkflowPlan.query.delete()
        RunCandidateReport.query.delete()
        Candidate.query.delete()
        db.session.commit()
    except Exception as ex:
        db.session.rollback()
        return jsonify({"error": str(ex)}), 500
    try:
        if os.path.isdir(UPLOAD_ROOT):
            for name in os.listdir(UPLOAD_ROOT):
                if name.startswith("candidate_"):
                    shutil.rmtree(os.path.join(UPLOAD_ROOT, name), ignore_errors=True)
    except Exception:
        pass
    return jsonify({"message": "All candidates cleared"}), 200


@app.post("/api/candidates/transfer-placement-officer")
def api_transfer_placement_officer():
    """Bulk reassign placement_officer_member from one value to another (optional scope: candidate_ids)."""
    data = request.json or {}
    from_po = (data.get("from_po") or "").strip()
    to_po = (data.get("to_po") or "").strip()
    if not from_po:
        return jsonify({"error": "from_po is required"}), 400
    if not to_po:
        return jsonify({"error": "to_po is required"}), 400
    if from_po.lower() == to_po.lower():
        return jsonify({"error": "from and to must be different"}), 400

    ids = data.get("candidate_ids")
    if ids is not None and isinstance(ids, list) and len(ids) > 0:
        id_set = []
        for x in ids:
            try:
                id_set.append(int(x))
            except (TypeError, ValueError):
                pass
        if not id_set:
            return jsonify({"error": "candidate_ids must be non-empty integers"}), 400
        q = Candidate.query.filter(Candidate.id.in_(id_set))
    else:
        q = Candidate.query

    updated = 0
    from_lower = from_po.lower()
    for c in q.all():
        cur = (c.placement_officer_member or "").strip()
        if cur.lower() == from_lower:
            c.placement_officer_member = to_po
            updated += 1
    db.session.commit()
    return jsonify({"message": f"Transferred {updated} candidate(s)", "updated": updated}), 200


@app.post("/api/candidates/transfer-rm-member")
def api_transfer_rm_member():
    """Bulk reassign rm_member from one value to another (optional scope: candidate_ids)."""
    data = request.json or {}
    from_rm = (data.get("from_rm") or "").strip()
    to_rm = (data.get("to_rm") or "").strip()
    if not from_rm:
        return jsonify({"error": "from_rm is required"}), 400
    if not to_rm:
        return jsonify({"error": "to_rm is required"}), 400
    if from_rm.lower() == to_rm.lower():
        return jsonify({"error": "from and to must be different"}), 400

    ids = data.get("candidate_ids")
    if ids is not None and isinstance(ids, list) and len(ids) > 0:
        id_set = []
        for x in ids:
            try:
                id_set.append(int(x))
            except (TypeError, ValueError):
                pass
        if not id_set:
            return jsonify({"error": "candidate_ids must be non-empty integers"}), 400
        q = Candidate.query.filter(Candidate.id.in_(id_set))
    else:
        q = Candidate.query

    updated = 0
    from_lower = from_rm.lower()
    for c in q.all():
        cur = (c.rm_member or "").strip()
        if cur.lower() == from_lower:
            c.rm_member = to_rm
            updated += 1
    db.session.commit()
    return jsonify({"message": f"Transferred {updated} candidate(s)", "updated": updated}), 200


@app.get("/api/targets")
def api_list_targets():
    q = Target.query.order_by(Target.company_name.asc()).all()
    return jsonify([t.to_dict() for t in q])


KNOWN_CRM_COUNTRY_KEYS = frozenset(
    x.lower()
    for x in (
        "Canada",
        "Germany",
        "Australia",
        "Austria",
        "Luxembourg",
        "Netherlands",
        "Ireland",
        "Sweden",
        "UAE",
        "Switzerland",
        "India",
    )
)


def canonical_company_name_for_crm(name: Optional[str]) -> str:
    """
    Map variants like 'Amazon - A', 'Amazon - B', 'Amazon - C' -> 'Amazon' for CRM deduplication.
    Rule: if the name ends with ' - X' where X is a single letter, drop that suffix.
    """
    s = (name or "").strip()
    if not s:
        return s
    parts = re.split(r"\s*-\s*", s)
    if len(parts) >= 2:
        last = parts[-1].strip()
        if len(last) == 1 and last.isalpha():
            return " - ".join(parts[:-1]).strip()
    return s


def _targets_for_crm_export():
    """Apply optional industry / country filters (matches dashboard drill-down)."""
    industry_filter = (request.args.get("industry") or "").strip()
    country_filter = (request.args.get("country") or "").strip()
    country_other = (request.args.get("country_other") or "").strip().lower() in ("1", "true", "yes")

    q = Target.query
    if industry_filter:
        q = q.filter(func.lower(Target.industry) == industry_filter.lower())
    rows = q.order_by(Target.company_name.asc()).all()

    if country_other:

        def _is_other(c: str) -> bool:
            x = (c or "").strip().lower()
            if not x or x == "global":
                return True
            return x not in KNOWN_CRM_COUNTRY_KEYS

        rows = [t for t in rows if _is_other(t.country or "")]
    elif country_filter:
        cf = country_filter.lower()
        rows = [t for t in rows if (t.country or "").strip().lower() == cf]

    return rows


@app.get("/api/targets/export/crm")
def api_export_targets_crm():
    """CSV for company CRM: full rows or deduped by canonical_company (see canonical_company_name_for_crm)."""
    mode = (request.args.get("mode") or "full").strip().lower()
    if mode not in ("full", "dedup"):
        mode = "full"

    rows = _targets_for_crm_export()
    buf = io.StringIO()
    buf.write("\ufeff")
    w = csv.writer(buf)

    if mode == "full":
        w.writerow(
            [
                "canonical_company",
                "company_name",
                "hr_email",
                "hr_name",
                "country",
                "industry",
                "target_role",
                "is_valid",
            ]
        )
        for t in rows:
            w.writerow(
                [
                    canonical_company_name_for_crm(t.company_name),
                    t.company_name or "",
                    t.hr_email or "",
                    t.hr_name or "",
                    t.country or "",
                    t.industry or "",
                    t.target_role or "",
                    "yes" if t.is_valid else "no",
                ]
            )
    else:
        groups = defaultdict(list)
        for t in rows:
            groups[canonical_company_name_for_crm(t.company_name)].append(t)

        w.writerow(
            [
                "canonical_company",
                "company_name_variants",
                "contact_count",
                "hr_emails",
                "hr_names",
                "countries",
                "industries",
                "target_roles",
            ]
        )
        for can in sorted(groups.keys(), key=lambda x: (x or "").lower()):
            items = groups[can]
            variants = sorted({(x.company_name or "").strip() for x in items if (x.company_name or "").strip()})
            emails = sorted({(x.hr_email or "").strip().lower() for x in items if (x.hr_email or "").strip()})
            names = sorted({(x.hr_name or "").strip() for x in items if (x.hr_name or "").strip()})
            countries = sorted({(x.country or "").strip() for x in items if (x.country or "").strip()})
            industries = sorted({(x.industry or "").strip() for x in items if (x.industry or "").strip()})
            roles = sorted({(x.target_role or "").strip() for x in items if (x.target_role or "").strip()})
            w.writerow(
                [
                    can,
                    "; ".join(variants),
                    len(items),
                    "; ".join(emails),
                    "; ".join(names),
                    "; ".join(countries),
                    "; ".join(industries),
                    "; ".join(roles),
                ]
            )

    fn = f"crm_targets_{mode}_{datetime.utcnow().strftime('%Y%m%d_%H%M')}.csv"
    return Response(
        buf.getvalue().encode("utf-8"),
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f'attachment; filename="{fn}"'},
    )


@app.post("/api/targets")
def api_upsert_target():
    data = request.json or {}
    tid = data.get("id")
    company_name = (data.get("company_name") or "").strip()
    hr_email = (data.get("hr_email") or "").strip().lower()
    country = (data.get("country") or "").strip() or "Germany"
    hr_name = (data.get("hr_name") or "").strip() or "Nil"
    target_role = (data.get("target_role") or "").strip()
    industry = (data.get("industry") or "").strip() or "Default"

    if not company_name or not hr_email:
        return jsonify({"error": "company_name and hr_email are required"}), 400

    ok, reason = is_valid_address_for_send(hr_email)

    if tid:
        t = Target.query.get(int(tid))
        if not t:
            return jsonify({"error": "Target not found"}), 404
        t.company_name = company_name
        t.hr_email = hr_email
        t.country = country
        t.hr_name = hr_name
        t.target_role = target_role
        t.industry = industry
        t.is_valid = bool(ok)
        t.invalid_reason = (reason if not ok else None)
        db.session.add(t)
        db.session.commit()
        return jsonify({"message": "Updated", "target": t.to_dict()}), 200

    t = Target(
        company_name=company_name,
        hr_email=hr_email,
        country=country,
        hr_name=hr_name,
        target_role=target_role,
        industry=industry,
        is_valid=bool(ok),
        invalid_reason=(reason if not ok else None),
    )
    db.session.add(t)
    db.session.commit()
    return jsonify({"message": "Created", "target": t.to_dict()}), 200

@app.get("/api/targets/<int:target_id>")
def api_get_target(target_id):
    t = Target.query.get_or_404(target_id)
    return jsonify(t.to_dict())

@app.delete("/api/targets/<int:target_id>")
def api_delete_target(target_id):
    t = Target.query.get_or_404(target_id)
    db.session.delete(t)
    db.session.commit()
    return jsonify({"message": "Deleted"}), 200

@app.post("/api/targets/clear")
def api_clear_targets():
    Target.query.delete()
    db.session.commit()
    return jsonify({"message": "Cleared"}), 200

@app.post("/api/targets/import")
def api_import_targets_xlsx():
    f = request.files.get("excel")
    if not f:
        return jsonify({"error": "excel file is required"}), 400

    form_industry = request.form.get("industry", "").strip()
    form_country = request.form.get("country", "").strip()

    wb = openpyxl.load_workbook(f.stream)
    sheet = wb.active

    # Target.query.delete()
    # db.session.commit()

    inserted = 0
    duplicates = 0
    invalid = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row: continue
        company = str(row[0]).strip() if row[0] else ""
        if not company: continue

        # Flexible column parsing
        b = str(row[1]).strip() if len(row) > 1 and row[1] else ""
        c = str(row[2]).strip() if len(row) > 2 and row[2] else ""

        if looks_like_email(b):
            target_role = ""
            hr_email = b.lower()
            row_country = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            hr_name = str(row[3]).strip() if len(row) > 3 and row[3] else "Nil"
            row_industry = str(row[4]).strip() if len(row) > 4 and row[4] else ""
        else:
            target_role = b
            hr_email = c.lower()
            row_country = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            hr_name = str(row[4]).strip() if len(row) > 4 and row[4] else "Nil"
            row_industry = str(row[5]).strip() if len(row) > 5 and row[5] else ""

        # Priority: form_country (UI context) > row_country (Excel) > "Germany" (fallback)
        country = form_country or row_country or "Germany"
        industry = form_industry or row_industry or "Default"

        if not hr_email: continue

        # Global deduplication by email as requested
        exists = Target.query.filter_by(hr_email=hr_email).first()
        if exists:
            duplicates += 1
            continue

        ok, reason = is_valid_address_for_send(hr_email)
        t = Target(
            company_name=company, target_role=target_role, hr_email=hr_email,
            country=country, hr_name=hr_name, industry=industry, is_valid=bool(ok),
            invalid_reason=(reason if not ok else None),
        )
        db.session.add(t)
        inserted += 1
        if not ok: invalid += 1

    db.session.commit()
    return jsonify({"message": "Imported", "inserted": inserted, "duplicates": duplicates, "invalid": invalid}), 200


@app.get("/api/targets/industries")
def api_list_industries():
    # Merge derived industries from targets and manual industries
    targets = Target.query.all()
    unique_names = {} # name -> [countries]
    for t in targets:
        n = (t.industry or "Default").strip()
        c = (t.country or "Global").strip()
        if n not in unique_names: unique_names[n] = set()
        unique_names[n].add(c)
    
    manual = Industry.query.all()
    for m in manual:
        if m.name not in unique_names: unique_names[m.name] = set()
        unique_names[m.name].add(m.country)
        
    res = []
    for name, countries in unique_names.items():
        res.append({"name": name, "countries": list(countries)})

    return jsonify(res)

@app.post("/api/targets/industries")
def api_add_industry():
    data = request.json or {}
    name = (data.get("name") or "").strip()
    country = (data.get("country") or "Global").strip()
    if not name:
        return jsonify({"error": "name is required"}), 400
    
    existing = Industry.query.filter_by(name=name, country=country).first()
    if existing:
        return jsonify({"message": "Industry already exists for this country"}), 200
    
    try:
        new_ind = Industry(name=name, country=country)
        db.session.add(new_ind)
        db.session.commit()
        return jsonify({"message": "Industry created", "id": new_ind.id}), 201
    except Exception as e:
        db.session.rollback()
        # If it's a unique constraint violation, provide a cleaner message
        if "UNIQUE constraint failed" in str(e):
            return jsonify({"error": f"Industry '{name}' already exists for {country}"}), 400
        return jsonify({"error": str(e)}), 500

@app.delete("/api/targets/industries/<string:name>")
def api_delete_industry(name):
    country = request.args.get("country")
    
    # 1. Delete associated targets first to clear data
    target_query = Target.query.filter(func.lower(Target.industry) == name.lower())
    if country:
        target_query = target_query.filter(func.lower(Target.country) == country.lower())
    
    deleted_targets = target_query.delete(synchronize_session=False)
    
    # 2. Delete the manual Industry entry if it exists
    industry_query = Industry.query.filter(func.lower(Industry.name) == name.lower())
    if country:
        industry_query = industry_query.filter(func.lower(Industry.country) == country.lower())
    
    ind = industry_query.first()
    if ind:
        db.session.delete(ind)
        
    db.session.commit()
    return jsonify({"message": "Deleted", "targets_removed": deleted_targets}), 200

@app.post("/api/targets/industries/cleanup")
def api_cleanup_industries():
    """Delete all manual Industry entries that have 0 associated target contacts."""
    manual = Industry.query.all()
    removed = []
    for ind in manual:
        count = Target.query.filter(func.lower(Target.industry) == ind.name.lower()).count()
        if count == 0:
            removed.append(ind.name)
            db.session.delete(ind)
    db.session.commit()
    return jsonify({"message": "Cleanup complete", "removed": removed, "count": len(removed)}), 200

# API Workspaces
@app.get("/api/workspaces")
def api_list_workspaces():
    ws = Workspace.query.order_by(Workspace.id.desc()).all()
    return jsonify([w.to_dict() for w in ws])

@app.post("/api/workspaces")
def api_save_workspace():
    data = request.json or {}
    wid = data.get("id")
    name = (data.get("name") or "").strip()
    candidate_id = data.get("candidate_id")
    industry = (data.get("industry") or "Default").strip()
    country = (data.get("country") or "Global").strip()
    
    if not name or not candidate_id:
        return jsonify({"error": "name and candidate_id are required"}), 400
        
    if wid:
        w = db.session.get(Workspace, int(wid))
        if not w: return jsonify({"error": "Workspace not found"}), 404
    else:
        w = Workspace()
        # Defaults for new workspace
        w.automation_batch_size = 10
        w.automation_interval_days = 2
        w.automation_max_emails = 1000
        
    w.name = name
    w.candidate_id = int(candidate_id)
    w.industry = industry
    w.country = country
    
    # Optional updates
    if "automation_enabled" in data: w.automation_enabled = bool(data["automation_enabled"])
    if "automation_batch_size" in data: w.automation_batch_size = int(data["automation_batch_size"])
    if "automation_interval_days" in data: w.automation_interval_days = int(data["automation_interval_days"])
    if "automation_max_emails" in data: w.automation_max_emails = int(data["automation_max_emails"])
    if "automation_per_run_cap" in data: w.automation_per_run_cap = max(1, min(500, int(data["automation_per_run_cap"])))
    
    # New Service Plan fields
    if "automation_type" in data: w.automation_type = data["automation_type"]
    if "scheduled_days" in data: w.scheduled_days = data["scheduled_days"]
    if "monthly_target" in data: w.monthly_target = int(data["monthly_target"])

    if "service_start_date" in data and data["service_start_date"]:
        try:
            w.service_start_date = datetime.fromisoformat(data["service_start_date"].replace("Z", ""))
        except Exception: pass
    if "service_end_date" in data and data["service_end_date"]:
        try:
            w.service_end_date = datetime.fromisoformat(data["service_end_date"].replace("Z", ""))
        except Exception: pass

    if "automation_next_run" in data and data["automation_next_run"]:
        try:
            # datetime-local sends "YYYY-MM-DDTHH:MM"
            w.automation_next_run = datetime.fromisoformat(data["automation_next_run"].replace("Z", ""))
        except Exception as e:
            print(f"Error parsing date: {e}")
            
    if w.automation_enabled and not w.automation_next_run:
        w.automation_next_run = datetime.now()
        
    db.session.add(w)
    db.session.commit()
    return jsonify({"message": "Saved", "workspace": w.to_dict()}), 200

@app.delete("/api/workspaces/<int:workspace_id>")
def api_delete_workspace(workspace_id):
    w = Workspace.query.get_or_404(workspace_id)
    db.session.delete(w)
    db.session.commit()
    return jsonify({"message": "Deleted"}), 200

@app.post("/api/targets/paste")
def api_paste_targets():
    inserted = 0
    duplicates = 0
    invalid = 0

    delim = "\t" if "\t" in text_raw and text_raw.count("\t") > text_raw.count(",") else ","
    reader = csv.reader(io.StringIO(text_raw), delimiter=delim)
    rows = list(reader)
    if not rows:
        return jsonify({"error": "No rows found"}), 400

    first = [c.strip().lower() for c in rows[0]]
    headerish = any(x in ("company", "company_name") for x in first) and any("email" in x for x in first)
    start = 1 if headerish else 0

    target_industry = (request.json or {}).get("industry", "Default")

    for r in rows[start:]:
        if not r or len(r) < 2: continue
        company = (r[0] or "").strip()
        if not company: continue

        if len(r) >= 3 and looks_like_email((r[1] or "").strip()):
            target_role = ""
            hr_email = (r[1] or "").strip().lower()
            country = (r[2] or "Germany").strip()
            hr_name = (r[3] or "Nil").strip() if len(r) > 3 else "Nil"
            industry = target_industry
        else:
            target_role = (r[1] or "").strip()
            hr_email = (r[2] or "").strip().lower() if len(r) > 2 else ""
            country = (r[3] or "Germany").strip() if len(r) > 3 else "Germany"
            hr_name = (r[4] or "Nil").strip() if len(r) > 4 else "Nil"
            industry = (r[5] or target_industry).strip() if len(r) > 5 else target_industry

        if not hr_email: continue

        # Global deduplication
        exists = Target.query.filter_by(hr_email=hr_email).first()
        if exists:
            duplicates += 1
            continue

        ok, reason = is_valid_address_for_send(hr_email)
        t = Target(
            company_name=company, target_role=target_role, hr_email=hr_email,
            country=country, hr_name=hr_name, industry=industry, is_valid=bool(ok),
            invalid_reason=(reason if not ok else None),
        )
        db.session.add(t)
        inserted += 1
        if not ok: invalid += 1

    db.session.commit()
    return jsonify({"message": "Pasted", "inserted": inserted, "duplicates": duplicates, "invalid": invalid}), 200

# Runs & Reports
@app.get("/api/runs")
def api_runs():
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 401
    # Also accept manifest=1 (some clients strip the word "download" from query strings).
    _dl = (request.args.get("download") or "").strip().lower()
    if _dl in ("crm_manifest_csv", "crm_manifest_tsv", "crm_manifest") or (request.args.get("manifest") or "").strip() == "1":
        return api_run_reports_crm_manifest_bulk()
    include_deleted = (request.args.get("include_deleted") or "").strip() in ("1", "true", "yes")
    q = SendRun.query
    if not include_deleted:
        # show only active (non-deleted) runs
        q = q.filter((SendRun.is_deleted == False) | (SendRun.is_deleted.is_(False)))

    # Full run history for Analytics (LAN/team). Optional ?limit= (default 10k, max 50k).
    try:
        lim = int((request.args.get("limit") or "10000").strip())
    except ValueError:
        lim = 10000
    lim = max(1, min(lim, 50000))

    rows = q.order_by(SendRun.id.desc()).limit(lim).all()
    return jsonify([r.to_dict() for r in rows])


def _parse_iso_to_naive_utc(s: str) -> Optional[datetime]:
    if not s or not str(s).strip():
        return None
    t = str(s).strip()
    if t.endswith("Z"):
        t = t[:-1] + "+00:00"
    try:
        dt = datetime.fromisoformat(t)
    except ValueError:
        return None
    if dt.tzinfo is not None:
        dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
    return dt


def _crm_manifest_reps_for_request() -> List[RunCandidateReport]:
    q = SendRun.query.filter((SendRun.is_deleted == False) | (SendRun.is_deleted.is_(False)))
    rolling = request.args.get("rolling_days", type=int)
    start_s = (request.args.get("start") or "").strip()
    end_s = (request.args.get("end") or "").strip()

    if rolling is not None and rolling > 0:
        cutoff = datetime.utcnow() - timedelta(days=float(rolling))
        q = q.filter(SendRun.created_at >= cutoff)
    elif start_s and end_s:
        start_dt = _parse_iso_to_naive_utc(start_s)
        end_dt = _parse_iso_to_naive_utc(end_s)
        if start_dt and end_dt:
            q = q.filter(SendRun.created_at >= start_dt, SendRun.created_at <= end_dt)

    runs = q.order_by(SendRun.id.desc()).limit(50000).all()
    run_ids = [r.id for r in runs]

    for rid in run_ids:
        try:
            _backfill_run_reports_if_missing(rid)
        except Exception as ex:
            print(f"[crm-manifest-bulk] backfill run {rid}: {ex}")

    if not run_ids:
        return []
    return (
        RunCandidateReport.query.filter(RunCandidateReport.run_id.in_(run_ids))
        .order_by(RunCandidateReport.run_id.desc(), RunCandidateReport.candidate_name.asc())
        .all()
    )


@app.get("/api/runs/reports/crm-manifest-bulk")
def api_run_reports_crm_manifest_bulk():
    """All candidate report rows for runs whose created_at falls in the optional date window (matches Analytics date filter)."""
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 401

    reps = _crm_manifest_reps_for_request()
    body = _crm_manifest_csv_bytes_for_reps(reps)
    buf = io.BytesIO(body)
    buf.seek(0)
    stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    return send_file(
        buf,
        mimetype="text/csv; charset=utf-8",
        as_attachment=True,
        download_name=f"crm_upload_manifest_all_runs_{stamp}.csv",
    )


@app.get("/api/crm-manifest-data")
def api_crm_manifest_data():
    """JSON rows for CRM manifest (browser builds CSV) — reliable when CSV GET routes are misrouted."""
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 401
    reps = _crm_manifest_reps_for_request()
    rows = _crm_manifest_row_dicts_for_reps(reps)
    return jsonify({"schema": "crm_manifest_v1", "row_count": len(rows), "rows": rows})


@app.route("/api/reports/crm-manifest", methods=["GET", "POST"])
def api_reports_crm_manifest_entry():
    """CRM upload CSV only — same query params as bulk (rolling_days | start & end). Never returns /api/runs JSON."""
    return api_run_reports_crm_manifest_bulk()


@app.get("/api/crm-manifest.csv")
def api_crm_manifest_csv_file():
    """Bulk CRM manifest CSV only — never the JSON array from GET /api/runs. Query: rolling_days | start & end (ISO)."""
    return api_run_reports_crm_manifest_bulk()


@app.get("/api/crm-manifest-export")
def api_crm_manifest_export_no_ext():
    """Same as /api/crm-manifest.csv (some proxies block paths with a dot before the extension)."""
    return api_run_reports_crm_manifest_bulk()


@app.get("/api/crm-upload-manifest-bulk")
def api_crm_upload_manifest_bulk():
    """Same export as /api/runs/reports/crm-manifest-bulk (explicit second route so registration never depends on add_url_rule)."""
    return api_run_reports_crm_manifest_bulk()


@app.get("/api/runs/<int:run_id>/reports")
def api_run_reports(run_id):
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 401
    try:
        _backfill_run_reports_if_missing(run_id)
    except Exception as ex:
        print(f"[api_run_reports] backfill failed for run {run_id}: {ex}")
    q = RunCandidateReport.query.filter(RunCandidateReport.run_id == run_id).order_by(RunCandidateReport.candidate_name.asc()).all()
    return jsonify([x.to_dict() for x in q])


def _crm_manifest_enrollment_cell(cand: Optional[Candidate]) -> str:
    """JSA enrollment URL for company CRM (column still labeled Enrollment ID)."""
    if not cand:
        return ""
    raw = (cand.enrollment_id or "").strip().replace("\r", " ").replace("\n", " ")
    if not raw:
        return ""
    low = raw.lower()
    if low.startswith("http://") or low.startswith("https://"):
        return raw.strip()
    prefix = CRM_JSA_ENROLLMENT_UPDATE_PREFIX or "https://backend.terratern.com/jsa-enrollment/update?id="
    return f"{prefix}{raw}"


def _crm_manifest_report_abs_path(rep: RunCandidateReport) -> str:
    """Absolute on-disk path to this run’s candidate CSV (under project reports/)."""
    p = (rep.report_csv_path or "").strip()
    if not p:
        return ""
    try:
        return os.path.abspath(os.path.normpath(p))
    except Exception:
        return p


def _crm_manifest_row_dicts_for_reps(reps: List[RunCandidateReport]) -> List[dict]:
    out: List[dict] = []
    for rep in reps:
        cand = Candidate.query.get(rep.candidate_id)
        enr_cell = _crm_manifest_enrollment_cell(cand)
        name = rep.candidate_name or (cand.name if cand else "") or ""
        name = str(name).replace("\r", " ").replace("\n", " ").strip()
        path_disp = _crm_manifest_report_abs_path(rep)
        out.append({"enrollment_id": enr_cell, "candidates": name, "report_path": path_disp})
    return out


def _crm_manifest_csv_bytes_for_reps(reps: List[RunCandidateReport]) -> bytes:
    """UTF-8 CSV with BOM for Excel: Enrollment ID (JSA URL), Candidates, Report Path (absolute CSV path)."""
    rows = _crm_manifest_row_dicts_for_reps(reps)
    sio = io.StringIO()
    w = csv.writer(sio, lineterminator="\r\n")
    w.writerow(["Enrollment ID", "Candidates", "Report Path"])
    for row in rows:
        w.writerow([row["enrollment_id"], row["candidates"], row["report_path"]])
    return ("\ufeff" + sio.getvalue()).encode("utf-8")


@app.get("/api/runs/<int:run_id>/reports/crm-manifest")
def api_run_reports_crm_manifest(run_id: int):
    """One CSV for company CRM: JSA enrollment URL, candidate name, absolute report CSV path."""
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 401
    try:
        _backfill_run_reports_if_missing(run_id)
    except Exception as ex:
        print(f"[crm-manifest] backfill failed for run {run_id}: {ex}")
    reps = (
        RunCandidateReport.query.filter(RunCandidateReport.run_id == run_id)
        .order_by(RunCandidateReport.candidate_name.asc())
        .all()
    )
    body = _crm_manifest_csv_bytes_for_reps(reps)
    buf = io.BytesIO(body)
    buf.seek(0)
    return send_file(
        buf,
        mimetype="text/csv; charset=utf-8",
        as_attachment=True,
        download_name=f"run_{run_id}_crm_upload_manifest.csv",
    )
@app.delete("/api/runs/<int:run_id>")
def api_delete_run(run_id: int):
    run = SendRun.query.get(run_id)
    if not run:
        return jsonify({"error": "Run not found"}), 404

    run_id_str = str(run_id)
    # Request stop (safe even if not running)
    st = run_states.setdefault(run_id_str, {"stop": False, "pause": False})
    st["stop"] = True
    st["pause"] = False

    # Soft-delete so worker threads can still commit safely
    run.is_deleted = True
    run.status = "deleted"

    # Remove stored reports (DB rows + files)
    try:
        reps = RunCandidateReport.query.filter(RunCandidateReport.run_id == run_id).all()
        for rep in reps:
            try:
                if rep.report_csv_path and os.path.exists(rep.report_csv_path):
                    os.remove(rep.report_csv_path)
            except Exception:
                pass
        RunCandidateReport.query.filter(RunCandidateReport.run_id == run_id).delete()
        db.session.commit()
    except Exception:
        db.session.rollback()

    # Remove run folder if exists
    try:
        folder = os.path.join(REPORTS_DIR, f"run_{run_id}")
        shutil.rmtree(folder, ignore_errors=True)
    except Exception:
        pass

    try:
        db.session.add(run)
        db.session.commit()
    except Exception:
        db.session.rollback()

    push_progress(run_id_str, "🗑️ Run deleted by user.")
    return jsonify({"status": "deleted", "run_id": run_id}), 200


@app.delete("/api/runs")
def api_delete_all_runs():
    """Delete ALL runs and their associated reports."""
    try:
        runs = SendRun.query.filter(SendRun.is_deleted == False).all()
        deleted_count = 0

        for run in runs:
            run_id = run.id
            run_id_str = str(run_id)

            # Stop any running process
            st = run_states.setdefault(run_id_str, {"stop": False, "pause": False})
            st["stop"] = True
            st["pause"] = False

            # Soft-delete
            run.is_deleted = True
            run.status = "deleted"

            # Remove report files
            reps = RunCandidateReport.query.filter(RunCandidateReport.run_id == run_id).all()
            for rep in reps:
                try:
                    if rep.report_csv_path and os.path.exists(rep.report_csv_path):
                        os.remove(rep.report_csv_path)
                except Exception:
                    pass
            RunCandidateReport.query.filter(RunCandidateReport.run_id == run_id).delete()

            # Remove run folder
            try:
                folder = os.path.join(REPORTS_DIR, f"run_{run_id}")
                shutil.rmtree(folder, ignore_errors=True)
            except Exception:
                pass

            deleted_count += 1

        db.session.commit()
        return jsonify({"status": "ok", "deleted": deleted_count}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"error": str(e)}), 500


@app.post("/api/runs/control")
def api_run_control():
    data = request.json or {}
    run_id = str(data.get("runId") or "")
    action = data.get("action")
    if not run_id:
        return jsonify({"error": "runId is required"}), 400
    if run_id not in run_states:
        run_states[run_id] = {"stop": False, "pause": False}
    state = run_states[run_id]
    if action == "stop":
        state["stop"] = True
        # Also update DB status immediately so UI unlocks even if worker is stuck
        try:
            with app.app_context():
                run = db.session.get(SendRun, int(run_id))
                if run and run.status not in ("done", "failed", "stopped", "deleted"):
                    run.status = "stopped"
                    run.ended_at = datetime.utcnow()
                    db.session.add(run)
                    db.session.commit()
        except Exception:
            pass
        push_progress(run_id, "ℹ️ STOP requested by user.")
    elif action == "pause":
        state["pause"] = True
        push_progress(run_id, "ℹ️ PAUSE requested by user.")
    elif action == "resume":
        state["pause"] = False
        push_progress(run_id, "ℹ️ RESUME requested by user.")
    else:
        return jsonify({"error": "Invalid action"}), 400
    return jsonify({"status": f"{action} applied for run {run_id}"}), 200

@app.get("/api/runs/<run_id>/progress")
def api_run_progress(run_id):
    run_id_str = str(run_id)
    consumer_q = Queue()
    # Register this consumer
    with progress_lock:
        if run_id_str not in progress_queues:
            progress_queues[run_id_str] = []
        progress_queues[run_id_str].append(consumer_q)

    def stream():
        yield "data: CONNECTED\n\n"
        try:
            while True:
                try:
                    msg = consumer_q.get(timeout=60)
                    yield f"data: {msg}\n\n"
                except Empty:
                    yield ":\n\n"
        finally:
            # Unregister this consumer on disconnect
            with progress_lock:
                try:
                    progress_queues.get(run_id_str, []).remove(consumer_q)
                except ValueError:
                    pass
    return Response(stream(), mimetype="text/event-stream")

def _disambiguate_duplicate_company_names_in_rows(rows):
    """First row keeps the original name; each further identical name (case-insensitive) gets ' a', ' b', …."""
    counts = {}
    out = []
    for r in rows:
        r = dict(r)
        base = (r.get("company_name") or "").strip()
        key = base.lower()
        seen = counts.get(key, 0)
        if seen == 0:
            disp = base
        else:
            suf = chr(ord("a") + seen - 1) if seen <= 26 else str(seen)
            disp = f"{base} {suf}" if base else suf
        counts[key] = seen + 1
        r["company_name"] = disp
        out.append(r)
    return out


def _write_candidate_csv(path, rows):
    headers = ["company_name", "role", "country", "date", "hr_name"]
    rows = _disambiguate_duplicate_company_names_in_rows(rows)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(headers)
        for r in rows:
            w.writerow([
                r.get("company_name",""),
                r.get("role",""),
                r.get("country",""),
                r.get("date",""),
                r.get("hr_name","Nil") or "Nil"
            ])


def _norm_candidate_name_chunk(s: str) -> str:
    s = (s or "").replace("\u00a0", " ").replace("\u2003", " ").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _loose_run_name_matches_candidate(hint_name: str, cand: Candidate) -> bool:
    """Match stored run name to CRM row: full name, tokens, or short initials (e.g. Hemalatha C ~ Hemalatha Chandrasekar)."""
    h = _norm_candidate_name_chunk(hint_name).lower()
    nm = _norm_candidate_name_chunk(cand.name or "").lower()
    if not h or not nm:
        return False
    if h == nm:
        return True
    h_parts = [p.rstrip(".") for p in h.split() if p]
    c_parts = [p.rstrip(".") for p in nm.split() if p]
    if not h_parts or not c_parts:
        return False
    used = [False] * len(c_parts)
    for hp in h_parts:
        ok = False
        for i, cp in enumerate(c_parts):
            if used[i]:
                continue
            if cp == hp:
                used[i] = True
                ok = True
                break
            if len(hp) <= 2 and len(cp) > len(hp) and cp.startswith(hp):
                used[i] = True
                ok = True
                break
        if not ok:
            return False
    return True


def _candidate_ids_from_send_run(run: SendRun) -> list:
    """Resolve comma-separated names on SendRun to (candidate_id, display_name) pairs."""
    out = []
    seen = set()
    for raw in (run.candidate_names or "").split(","):
        name = _norm_candidate_name_chunk(raw)
        if not name:
            continue
        c = Candidate.query.filter(func.lower(Candidate.name) == name.lower()).first()
        if not c:
            c = Candidate.query.filter(Candidate.name.ilike(name)).first()
        if not c:
            like = "%".join(name.split())
            if like:
                # Prefer shortest name so "Chandra" beats "… Rajachandran" when both match %Chandra%
                c = (
                    Candidate.query.filter(Candidate.name.ilike(f"%{like}%"))
                    .order_by(func.length(Candidate.name))
                    .first()
                )
        if not c:
            first = name.split()[0] if name.split() else ""
            first_clean = re.sub(r"[^\w]", "", first).lower()
            if len(first_clean) >= 2:
                cand_list = (
                    Candidate.query.filter(func.lower(Candidate.name).like(f"{first_clean}%"))
                    .order_by(func.length(Candidate.name))
                    .limit(40)
                    .all()
                )
                for cand in cand_list:
                    if _loose_run_name_matches_candidate(name, cand):
                        c = cand
                        break
        if not c:
            continue
        if c.id in seen:
            continue
        seen.add(c.id)
        out.append((c.id, (c.name or name).strip()))
    return out


def _name_token_matches_hint(hint: str, cand: Candidate) -> bool:
    """True if hint string refers to this candidate (single token, full name, or Hemalatha C style)."""
    hint = _norm_candidate_name_chunk(hint)
    if not hint:
        return False
    hlow = hint.lower()
    nm = _norm_candidate_name_chunk(cand.name or "").lower()
    if hlow == nm:
        return True
    if hlow in nm.split():
        return True
    return _loose_run_name_matches_candidate(hint, cand)


def _candidate_ids_for_run_from_sent_history(run: SendRun) -> list:
    """Infer candidates who sent during the run window when candidate_names resolution failed."""
    anchor = run.started_at or run.created_at
    if not anchor:
        return []
    end_a = run.ended_at or anchor
    if end_a < anchor:
        end_a = anchor
    t0 = anchor - timedelta(hours=24)
    t1 = end_a + timedelta(hours=24)

    rows = (
        db.session.query(SentHistory.candidate_id, func.count(SentHistory.id).label("cnt"))
        .filter(SentHistory.sent_at >= t0, SentHistory.sent_at <= t1)
        .group_by(SentHistory.candidate_id)
        .order_by(func.count(SentHistory.id).desc())
        .all()
    )
    if not rows:
        ts0 = (run.created_at or anchor) - timedelta(days=3)
        ts1 = (run.ended_at or run.started_at or run.created_at or datetime.utcnow()) + timedelta(
            days=21
        )
        if ts1 < ts0:
            ts1 = ts0 + timedelta(days=21)
        rows = (
            db.session.query(SentHistory.candidate_id, func.count(SentHistory.id).label("cnt"))
            .filter(SentHistory.sent_at >= ts0, SentHistory.sent_at <= ts1)
            .group_by(SentHistory.candidate_id)
            .order_by(func.count(SentHistory.id).desc())
            .all()
        )
    if not rows:
        return []

    hints = [re.sub(r"\s+", " ", p).strip() for p in (run.candidate_names or "").split(",") if p.strip()]

    if len(rows) == 1:
        cid = rows[0][0]
        c = Candidate.query.get(cid)
        return [(cid, (c.name or "").strip())] if c else []

    if not hints:
        if len(rows) > 8:
            return []
        out = []
        seen = set()
        for r in rows:
            cid = r[0]
            if cid in seen:
                continue
            c = Candidate.query.get(cid)
            if not c:
                continue
            seen.add(cid)
            out.append((c.id, (c.name or "").strip()))
        return out

    out = []
    seen = set()
    candidate_ids = [r[0] for r in rows]
    for cid in candidate_ids:
        c = Candidate.query.get(cid)
        if not c or c.id in seen:
            continue
        if any(_name_token_matches_hint(h, c) for h in hints):
            seen.add(c.id)
            out.append((c.id, (c.name or "").strip()))
    return out


def _finalize_hist_csv_rows(rows: list, run: SendRun) -> list:
    """Keep sends that fall in the run interval (with padding); drop internal _ts. If clip removes everything, keep all."""
    if not rows or not any("_ts" in r for r in rows):
        return rows
    rs = run.started_at or run.created_at
    re = run.ended_at or rs
    if not rs or not re:
        return [{k: v for k, v in r.items() if k != "_ts"} for r in rows]
    lo = rs - timedelta(minutes=45)
    hi = re + timedelta(minutes=45)
    clipped = [r for r in rows if r.get("_ts") is not None and lo <= r["_ts"] <= hi]
    use = clipped if clipped else rows
    return [{k: v for k, v in r.items() if k != "_ts"} for r in use]


def _csv_rows_from_sent_history_for_run(run: SendRun, candidate_id: int) -> list:
    """Rebuild CRM CSV rows from SentHistory + targets (and JobApplication fallback)."""
    anchor = run.started_at or run.created_at
    if not anchor:
        anchor = datetime.utcnow()
    end_a = run.ended_at or anchor
    if end_a < anchor:
        end_a = anchor
    t0 = anchor - timedelta(hours=6)
    t1 = end_a + timedelta(hours=6)
    if (run.status or "").lower() in ("running", "queued", "paused_network"):
        t1 = max(t1, datetime.utcnow())

    cand = Candidate.query.get(candidate_id)
    cand_roles = parse_roles((cand.roles_text if cand else "") or "") or ["Candidate"]

    def hist_to_rows(hist_list):
        rows_out = []
        for i, h in enumerate(hist_list):
            em = (h.target_email or "").strip()
            tgt = Target.query.filter(func.lower(Target.hr_email) == func.lower(em)).first()
            if not tgt:
                tgt = Target.query.filter_by(hr_email=h.target_email).first()
            company = (tgt.company_name if tgt else "") or ""
            role_from_tgt = (tgt.target_role if tgt and tgt.target_role else "") or ""
            role_from_tgt = (role_from_tgt or "").strip()
            role = role_from_tgt if role_from_tgt else cand_roles[i % len(cand_roles)]
            country = (tgt.country if tgt else "") or ""
            hrn = (tgt.hr_name if tgt else "") or "Nil"
            d = h.sent_at.strftime("%d-%m-%Y") if h.sent_at else ""
            rows_out.append({
                "company_name": company,
                "role": role,
                "country": country,
                "date": d,
                "hr_name": hrn or "Nil",
                "hr_email": em,
                "_ts": h.sent_at,
            })
        return rows_out

    hist = (
        SentHistory.query.filter(
            SentHistory.candidate_id == candidate_id,
            SentHistory.sent_at >= t0,
            SentHistory.sent_at <= t1,
        )
        .order_by(SentHistory.sent_at.asc())
        .all()
    )
    rows = _finalize_hist_csv_rows(hist_to_rows(hist), run)
    if rows:
        return rows

    day0 = utc_calendar_start_naive(anchor)
    day1 = day0 + timedelta(days=1)
    hist = (
        SentHistory.query.filter(
            SentHistory.candidate_id == candidate_id,
            SentHistory.sent_at >= day0,
            SentHistory.sent_at < day1,
        )
        .order_by(SentHistory.sent_at.asc())
        .all()
    )
    rows = _finalize_hist_csv_rows(hist_to_rows(hist), run)
    if rows:
        return rows

    # Loose fallback: from run start (clock skew / odd ended_at) — cap rows to avoid dragging in very old sends
    if int(run.sent or 0) > 0:
        ts = run.started_at or run.created_at
        if ts:
            cap = min(max(int(run.sent or 0) + 80, 120), 2500)
            hist = (
                SentHistory.query.filter(
                    SentHistory.candidate_id == candidate_id,
                    SentHistory.sent_at >= ts - timedelta(minutes=15),
                )
                .order_by(SentHistory.sent_at.asc())
                .limit(cap)
                .all()
            )
            rows = _finalize_hist_csv_rows(hist_to_rows(hist), run)
            if rows:
                return rows

    # Stopped / imported runs: wide bracket from run creation through end (fixes empty Reports when timestamps are off)
    if int(run.sent or 0) > 0:
        ts_start = (run.created_at or anchor) - timedelta(days=3)
        ts_end = (run.ended_at or run.started_at or run.created_at or datetime.utcnow()) + timedelta(days=21)
        if ts_end < ts_start:
            ts_end = ts_start + timedelta(days=21)
        hist = (
            SentHistory.query.filter(
                SentHistory.candidate_id == candidate_id,
                SentHistory.sent_at >= ts_start,
                SentHistory.sent_at <= ts_end,
            )
            .order_by(SentHistory.sent_at.asc())
            .limit(5000)
            .all()
        )
        rows = _finalize_hist_csv_rows(hist_to_rows(hist), run)
        if rows:
            return rows

    rows = []
    ja_narrow = (
        JobApplication.query.filter(
            JobApplication.candidate_id == candidate_id,
            JobApplication.created_at >= t0,
            JobApplication.created_at <= t1,
        )
        .order_by(JobApplication.created_at.asc())
        .all()
    )
    for ja in ja_narrow:
        rows.append({
            "company_name": ja.company_name or "",
            "role": ja.job_role or "",
            "country": ja.country or "",
            "date": ja.applied_date or (ja.created_at.strftime("%d-%m-%Y") if ja.created_at else ""),
            "hr_name": "Nil",
            "hr_email": "",
        })
    if rows:
        return rows
    if int(run.sent or 0) > 0:
        ts_start = (run.created_at or anchor) - timedelta(days=3)
        ts_end = (run.ended_at or run.started_at or run.created_at or datetime.utcnow()) + timedelta(days=21)
        if ts_end < ts_start:
            ts_end = ts_start + timedelta(days=21)
        ja_wide = (
            JobApplication.query.filter(
                JobApplication.candidate_id == candidate_id,
                JobApplication.created_at >= ts_start,
                JobApplication.created_at <= ts_end,
            )
            .order_by(JobApplication.created_at.asc())
            .limit(5000)
            .all()
        )
        for ja in ja_wide:
            rows.append({
                "company_name": ja.company_name or "",
                "role": ja.job_role or "",
                "country": ja.country or "",
                "date": ja.applied_date or (ja.created_at.strftime("%d-%m-%Y") if ja.created_at else ""),
                "hr_name": "Nil",
                "hr_email": "",
            })
    return rows


def _db_commit_retry(attempts: int = 8) -> bool:
    """SQLite can briefly lock under concurrent workers / WAL; retry before giving up."""
    for i in range(attempts):
        try:
            db.session.commit()
            return True
        except OperationalError as e:
            db.session.rollback()
            msg = str(e).lower()
            if "locked" not in msg and "busy" not in msg:
                raise
            time.sleep(0.08 * (i + 1))
    return False


def _csv_roles_column_all_empty(p: str) -> bool:
    """True if CSV has data rows but every role cell is blank (older backfill used target_role only)."""
    try:
        with open(p, newline="", encoding="utf-8", errors="ignore") as f:
            r = csv.reader(f)
            rows = list(r)
        if len(rows) < 2:
            return False
        header = [x.strip().lower() for x in (rows[0] or [])]
        try:
            ri = header.index("role")
        except ValueError:
            return False
        any_role = False
        for row in rows[1:]:
            if len(row) > ri and (row[ri] or "").strip():
                any_role = True
                break
        return not any_role
    except Exception:
        return False


def _run_report_row_is_stale(run: SendRun, rep: RunCandidateReport) -> bool:
    """True if CSV is missing, unreadable, header-only while run has sends, or role column empty for all data rows."""
    p = rep.report_csv_path or ""
    if not p or not os.path.exists(p):
        return True
    try:
        with open(p, encoding="utf-8", errors="ignore") as f:
            n_lines = sum(1 for _ in f)
        if n_lines <= 1 and int(run.sent or 0) > 0:
            return True
        # Do not require run.sent: empty-role CSVs must always be rebuilt.
        if n_lines > 1 and _csv_roles_column_all_empty(p):
            return True
    except Exception:
        return True
    return False


def _backfill_run_reports_if_missing(run_id: int) -> None:
    """Ensure run_candidate_reports + CSV files exist. Never delete DB report rows — only add or refresh files in place."""
    run = SendRun.query.get(run_id)
    if not run or run.is_deleted:
        return

    report_folder = os.path.join(REPORTS_DIR, f"run_{run_id}")
    os.makedirs(report_folder, exist_ok=True)

    for rep in list(RunCandidateReport.query.filter_by(run_id=run_id).all()):
        if not _run_report_row_is_stale(run, rep):
            continue
        cand = Candidate.query.get(rep.candidate_id)
        display = (cand.name if cand else None) or rep.candidate_name or "Candidate"
        rows = _csv_rows_from_sent_history_for_run(run, rep.candidate_id)
        safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", display)[:60]
        csv_path = os.path.join(report_folder, f"{safe_name}_{rep.candidate_id}.csv")
        _write_candidate_csv(csv_path, rows)
        rep.candidate_name = display
        rep.report_csv_path = csv_path
    _db_commit_retry()

    existing = {r.candidate_id for r in RunCandidateReport.query.filter_by(run_id=run_id).all()}
    pairs = _candidate_ids_from_send_run(run)
    if not pairs:
        pairs = _candidate_ids_for_run_from_sent_history(run)
    if not pairs:
        return

    added = False
    for cid, cand_name in pairs:
        if cid in existing:
            continue
        rows = _csv_rows_from_sent_history_for_run(run, cid)
        safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", cand_name)[:60]
        csv_path = os.path.join(report_folder, f"{safe_name}_{cid}.csv")
        _write_candidate_csv(csv_path, rows)
        db.session.add(
            RunCandidateReport(
                run_id=int(run_id),
                candidate_id=cid,
                candidate_name=cand_name,
                report_csv_path=csv_path,
            )
        )
        added = True
    if added:
        _db_commit_retry()

    # Roles-only fixes for files that are not flagged fully "stale" but still have empty role column
    run = SendRun.query.get(run_id)
    if run and not run.is_deleted:
        for rep in list(RunCandidateReport.query.filter_by(run_id=run_id).all()):
            p = rep.report_csv_path or ""
            if not p or not os.path.exists(p):
                continue
            if not _csv_roles_column_all_empty(p):
                continue
            rows = _csv_rows_from_sent_history_for_run(run, rep.candidate_id)
            _write_candidate_csv(p, rows)


def _persist_run_candidate_report_row(run_id_str: str, cand: dict, report_folder: str, sent_rows: list) -> None:
    """Write CSV + upsert run_candidate_reports so Analytics always has a row per candidate."""
    cand_name = cand.get("name") or "Candidate"
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", cand_name)[:60]
    csv_path = os.path.join(report_folder, f"{safe_name}_{cand['id']}.csv")
    _write_candidate_csv(csv_path, sent_rows)
    with app.app_context():
        rid = int(run_id_str)
        cid = int(cand["id"])
        rep = RunCandidateReport.query.filter_by(run_id=rid, candidate_id=cid).first()
        if rep:
            rep.candidate_name = cand_name
            rep.report_csv_path = csv_path
        else:
            db.session.add(
                RunCandidateReport(
                    run_id=rid,
                    candidate_id=cid,
                    candidate_name=cand_name,
                    report_csv_path=csv_path,
                )
            )
        _db_commit_retry()


def candidate_worker(run_id_str, cand, targets_data, delay_seconds, enable_bounce_check, report_folder, started_utc, results_list, progress_lock, progress_count):
    """Worker function for a SINGLE candidate in a run."""
    try:
        cand_name = cand["name"]
        sender_email = (cand.get("email") or "").strip()
        sender_pass = (cand.get("app_password") or "").strip()
        
        if not sender_email or not sender_pass:
            push_progress(run_id_str, f"❌ [{cand_name}] Missing email/app password. Skipping.")
            try:
                _persist_run_candidate_report_row(run_id_str, cand, report_folder, [])
            except Exception:
                pass
            results_list.append({"sent": 0, "failed": 0, "skipped": 0, "bounced": 0})
            return

        resume_path = cand.get("resume_path") or ""
        cover_path = cand.get("cover_letter_path") or ""
        if not resume_path or not os.path.exists(resume_path) or not cover_path or not os.path.exists(cover_path):
            push_progress(run_id_str, f"❌ [{cand_name}] Missing resume/cover letter. Skipping.")
            try:
                _persist_run_candidate_report_row(run_id_str, cand, report_folder, [])
            except Exception:
                pass
            results_list.append({"sent": 0, "failed": 0, "skipped": 0, "bounced": 0})
            return

        with open(resume_path, "rb") as f: resume_data = f.read()
        with open(cover_path, "rb") as f: cover_data = f.read()
        resume_filename = os.path.basename(resume_path)
        cover_filename = os.path.basename(cover_path)

        subject_template = cand.get("subject_template") or ""
        message_template = cand.get("message_template") or ""
        cand_roles = parse_roles(cand.get("roles_text") or "") or ["Candidate"]

        sent_rows = []
        hr_sent_set = set()
        
        # Local stats: bounced = SMTP recipient refused + optional Gmail IMAP NDR scan after sends
        c_stats = {"sent": 0, "failed": 0, "skipped": 0, "bounced": 0}

        push_progress(run_id_str, f"ℹ️ [{cand_name}] Started parallel outreach.")

        for i, t in enumerate(targets_data):
            # Check for stop/pause
            state = run_states.get(run_id_str, {"stop": False, "pause": False})
            if state.get("stop"): break
            while state.get("pause"):
                time.sleep(1)
                state = run_states.get(run_id_str, state)

            # Outreach Guard (SentHistory check)
            with app.app_context():
                exists = SentHistory.query.filter_by(candidate_id=cand["id"], target_email=t["hr_email"]).first()
                if exists:
                    with progress_lock:
                        progress_count[0] += 1
                        push_progress(run_id_str, f"PROGRESS:{progress_count[0]}")
                    push_progress(run_id_str, f"⏩ [{cand_name}] Skipping {t['hr_email']} - already sent.")
                    c_stats["skipped"] += 1
                    _bump_send_run_counter(int(run_id_str), "skipped", 1)
                    continue

            today_str = datetime.now().strftime("%d-%m-%Y")
            role_used = cand_roles[i % len(cand_roles)]
            pa_member = cand.get("pa_member") or ""
            rm_member = cand.get("rm_member") or ""
            po_member = cand.get("placement_officer_member") or ""

            # Prepare subject/body
            subject = (subject_template
                       .replace("{role}", role_used)
                       .replace("{company}", t.get("company_name") or "")
                       .replace("{hr_name}", t.get("hr_name") or "")
                       .replace("{pa_member}", pa_member)
                       .replace("{rm_member}", rm_member)
                       .replace("{placement_officer_member}", po_member)
                       .replace("{candidate}", cand_name))
            
            body = (message_template
                    .replace("{role}", role_used)
                    .replace("{company}", t.get("company_name") or "")
                    .replace("{hr_name}", t.get("hr_name") or "Team")
                    .replace("{pa_member}", pa_member)
                    .replace("{rm_member}", rm_member)
                    .replace("{placement_officer_member}", po_member)
                    .replace("{candidate}", cand_name))

            # Prepare Email
            msg = EmailMessage()
            msg["Subject"] = subject
            msg["From"] = f"{cand_name} <{sender_email}>"
            msg["To"] = t["hr_email"]
            msg["Date"] = formatdate(localtime=True)

            is_html = "<" in body and ">" in body
            if is_html:
                plain = re.sub('<[^<]+?>', '', body) # very simple strip
                msg.set_content(plain)
                msg.add_alternative(body, subtype="html")
            else:
                msg.set_content(body)

            msg.add_attachment(resume_data, maintype="application", subtype="pdf", filename=resume_filename)
            msg.add_attachment(cover_data, maintype="application", subtype="pdf", filename=cover_filename)

            try:
                # Actual send with retry
                status, err = send_with_retries(sender_email, sender_pass, msg, run_id_str)
                if status == "sent":
                    c_stats["sent"] += 1
                    rid = int(run_id_str)
                    with app.app_context():
                        db.session.add(SentHistory(candidate_id=cand["id"], target_email=t["hr_email"]))
                        db.session.add(JobApplication(
                            candidate_id=cand["id"],
                            company_name=t["company_name"],
                            job_role=role_used,
                            country=t.get("country",""),
                            applied_date=today_str,
                            status="Applied"
                        ))
                        wp = (
                            WorkflowPlan.query.filter_by(candidate_id=cand["id"])
                            .filter(WorkflowPlan.status.in_(["active", "paused"]))
                            .order_by(WorkflowPlan.id.desc())
                            .first()
                        )
                        if wp:
                            wp.total_applied = (wp.total_applied or 0) + 1
                        ri = (cand.get("run_industry") or "Default").strip() or "Default"
                        rc = (cand.get("run_country") or "Global").strip() or "Global"
                        ws_auto = Workspace.query.filter_by(
                            candidate_id=cand["id"],
                            automation_enabled=True,
                            industry=ri,
                            country=rc,
                        ).first()
                        if ws_auto:
                            ws_auto.automation_total_sent = (ws_auto.automation_total_sent or 0) + 1
                        db.session.commit()
                    _bump_send_run_counter(rid, "sent", 1)
                    
                    sent_rows.append({
                        "company_name": t["company_name"], "role": role_used, "country": t.get("country",""),
                        "date": today_str, "hr_name": t.get("hr_name","Nil"), "hr_email": t["hr_email"]
                    })
                    hr_sent_set.add(t["hr_email"].lower())
                    push_progress(run_id_str, f"✅ [{cand_name}] Sent to {t['hr_email']}")
                    push_progress(run_id_str, f"SENT:{c_stats['sent']}")
                elif status == "bounced":
                    c_stats["bounced"] += 1
                    push_progress(run_id_str, f"📭 [{cand_name}] Bounced / rejected: {t['hr_email']}: {err or 'recipient rejected'}")
                    _bump_send_run_counter(int(run_id_str), "bounced", 1)
                else:
                    c_stats["failed"] += 1
                    push_progress(run_id_str, f"❌ [{cand_name}] Failed to {t['hr_email']}: {err}")
                    _bump_send_run_counter(int(run_id_str), "failed", 1)
                
                with progress_lock:
                    progress_count[0] += 1
                    push_progress(run_id_str, f"PROGRESS:{progress_count[0]}")
            except Exception as e:
                c_stats["failed"] += 1
                push_progress(run_id_str, f"❌ [{cand_name}] Thread error: {e}")
                _bump_send_run_counter(int(run_id_str), "failed", 1)

            if i < len(targets_data) - 1:
                time.sleep(delay_seconds)

        # Optional IMAP bounce scan: drops bounced rows from CSV and moves counts from Sent → Bounced
        if enable_bounce_check and hr_sent_set:
            push_progress(run_id_str, f"ℹ️ [{cand_name}] Checking bounces...")
            bounces = fetch_bounces_gmail(sender_email, sender_pass, started_utc, hr_sent_set)
            imap_n = len(bounces)
            if imap_n:
                sent_rows = [r for r in sent_rows if r.get("hr_email", "").lower() not in bounces]
                c_stats["sent"] = max(0, c_stats["sent"] - imap_n)
                c_stats["bounced"] += imap_n

        # Save Report
        _persist_run_candidate_report_row(run_id_str, cand, report_folder, sent_rows)

        push_progress(
            run_id_str,
            f"🏁 [{cand_name}] Thread finished. Sent: {c_stats['sent']}, Failed: {c_stats['failed']}, Bounced: {c_stats['bounced']}",
        )
        results_list.append(dict(c_stats))

    except Exception as e:
        push_progress(run_id_str, f"💥 [{cand.get('name', 'Unknown')}] Fatal Thread Error: {e}")
        results_list.append({"sent": 0, "failed": 1, "skipped": 0, "bounced": 0})
        try:
            _persist_run_candidate_report_row(run_id_str, cand, report_folder, [])
        except Exception:
            pass

def run_sender_thread(run_id: int, candidate_ids, delay_seconds: int, enable_bounce_check: bool, target_limit: int = None, target_offset: int = 0, industry: str = None, country: str = None):
    run_id_str = str(run_id)
    started_utc = datetime.now(timezone.utc)
    push_progress(run_id_str, f"RUN:{run_id_str}")

    with app.app_context():
        run = SendRun.query.get(run_id)
        if not run: return
        run.status = "running"; run.started_at = datetime.utcnow()
        db.session.commit()

# Targets
        t_query = Target.query.filter_by(is_valid=True)
        if industry and industry != "Default": t_query = t_query.filter_by(industry=industry)
        if country and country != "Global": t_query = t_query.filter_by(country=country)

        # Refinement: If single candidate (workflow mode), exclude already sent
        if len(candidate_ids) == 1:
            cid = candidate_ids[0]
            sent_emails = [h.target_email for h in SentHistory.query.filter_by(candidate_id=cid).all()]
            if sent_emails:
                t_query = t_query.filter(~Target.hr_email.in_(sent_emails))

        eff_limit = target_limit
        if candidate_ids:
            quotas = [remaining_daily_send_quota(cid, datetime.utcnow()) for cid in candidate_ids]
            q_use = min(quotas) if quotas else int(SMART_AUTOMATION_BACKLOG_SAFE_CAP)
            if q_use <= 0:
                eff_limit = 0
            elif eff_limit is not None:
                eff_limit = min(int(eff_limit), q_use)
            else:
                eff_limit = q_use

        if eff_limit == 0:
            targets_data = []
        else:
            if target_offset:
                t_query = t_query.offset(target_offset)
            if eff_limit is not None and eff_limit > 0:
                t_query = t_query.limit(eff_limit)
            targets_data = [t.to_dict() for t in t_query.all()]
        
        # Snapshot candidates
        candidates_data = []
        for cid in candidate_ids:
            c = Candidate.query.get(cid)
            if c:
                d = c.to_dict_detail()
                d["pa_member"] = (c.pa_member or c.app_code or "")
                d["resume_path"] = c.resume_path
                d["cover_letter_path"] = c.cover_letter_path
                d["run_industry"] = (industry or "Default").strip() or "Default"
                d["run_country"] = (country or "Global").strip() or "Global"
                candidates_data.append(d)

        run.total_targets = len(targets_data)
        run.candidate_names = ", ".join([c["name"] for c in candidates_data])
        db.session.commit()

    if not targets_data or not candidates_data:
        push_progress(run_id_str, "❌ No targets or candidates found. Run aborted.")
        push_progress(run_id_str, "DONE")
        return

    report_folder = os.path.join(REPORTS_DIR, f"run_{run_id}")
    os.makedirs(report_folder, exist_ok=True)
    push_progress(run_id_str, f"TOTAL:{len(candidates_data) * len(targets_data)}")

    results_list = []

    # Spawn threads for parallel runs
    progress_lock = Lock()
    progress_count = [0]
    threads = []
    for cand in candidates_data:
        t = Thread(target=candidate_worker, args=(run_id_str, cand, targets_data, delay_seconds, enable_bounce_check, report_folder, started_utc, results_list, progress_lock, progress_count))
        t.start()
        threads.append(t)
    
    for t in threads:
        t.join()

    # Final update
    with app.app_context():
        run = SendRun.query.get(run_id)
        if run:
            total_sent = sum(r["sent"] for r in results_list)
            total_failed = sum(r["failed"] for r in results_list)
            total_skipped = sum(r["skipped"] for r in results_list)
            total_bounced = sum(r["bounced"] for r in results_list)
            
            run.sent = total_sent
            run.failed = total_failed
            run.skipped = total_skipped
            run.bounced = total_bounced
            run.status = "done"; run.ended_at = datetime.utcnow()
            db.session.commit()

        push_progress(run_id_str, f"🏁 All candidate threads completed. Totals -> Sent: {total_sent}, Failed: {total_failed}, Bounced: {total_bounced}")

        push_progress(run_id_str, "DONE")

def run_scheduler_loop():
    """Background loop to check for scheduled candidates and automation via Workspaces."""
    import time
    time.sleep(10) # Wait for server to fully start
    while True:
        try:
            with app.app_context():
                now = datetime.utcnow()
                # 1. Manual Scheduled (counts toward one scheduler automation / candidate / UTC day)
                due_cands = Candidate.query.filter(Candidate.scheduled_time != None, Candidate.scheduled_time <= now).all()
                for cand in due_cands:
                    cid = cand.id
                    if scheduler_should_pause_automation_for_candidate_bucklist(cand, now):
                        cand.scheduled_time = now + timedelta(days=1)
                        db.session.commit()
                        print(
                            f"[Scheduler] Candidate #{cid}: scheduled_time deferred — bucklist 0–90 band (automation off)."
                        )
                        continue
                    if scheduler_automation_already_used_today(cid, now):
                        cand.scheduled_time = now + timedelta(days=1)
                        db.session.commit()
                        continue
                    cand.scheduled_time = None
                    db.session.commit()
                    auto_start_single_run(cid, from_scheduler=True)

                # 2. Workspace automation — at most one send batch per candidate per UTC day across all buckets
                due_ws = [
                    ws
                    for ws in Workspace.query.filter_by(automation_enabled=True).all()
                    if not ws.automation_next_run or ws.automation_next_run <= now
                ]
                by_cand_ws = defaultdict(list)
                for ws in due_ws:
                    by_cand_ws[ws.candidate_id].append(ws)

                for cid, wss in by_cand_ws.items():
                    cand = db.session.get(Candidate, cid)
                    if not cand:
                        for ws in wss:
                            ws.automation_next_run = now + timedelta(
                                days=max(1, int(ws.automation_interval_days or 1))
                            )
                        db.session.commit()
                        continue
                    if scheduler_should_pause_automation_for_candidate_bucklist(cand, now):
                        for ws in wss:
                            ws.automation_next_run = now + timedelta(
                                days=max(1, int(ws.automation_interval_days or 1))
                            )
                        db.session.commit()
                        print(
                            f"[Scheduler] Workspace automation deferred for candidate #{cid} — service day band 0–90 (automation off)."
                        )
                        continue
                    if scheduler_automation_already_used_today(cid, now):
                        for ws in wss:
                            ws.automation_next_run = now + timedelta(
                                days=max(1, int(ws.automation_interval_days or 1))
                            )
                        db.session.commit()
                        continue

                    wp = (
                        WorkflowPlan.query.filter_by(candidate_id=cand.id)
                        .filter(WorkflowPlan.status.in_(["paused", "active"]))
                        .order_by(WorkflowPlan.id.desc())
                        .first()
                    )
                    applied_cap = int(wp.total_applied or 0) if wp else max(
                        int(x.automation_total_sent or 0) for x in wss
                    )
                    if applied_cap >= 1200:
                        for ws in wss:
                            ws.automation_enabled = False
                            ws.automation_next_run = now + timedelta(
                                days=max(1, int(ws.automation_interval_days or 1))
                            )
                        db.session.commit()
                        continue

                    best_ws = None
                    best_batch = -1
                    for ws in wss:
                        svc = getattr(cand, "smart_service_start_date", None) or ws.service_start_date
                        if not svc:
                            svc = now
                        if wp:
                            days_elapsed = workflow_plan_elapsed_service_days(wp, now)
                            applied = int(wp.total_applied or 0)
                        else:
                            days_elapsed = max(0, (now - svc).days)
                            applied = int(ws.automation_total_sent or 0)
                        cap = int(ws.automation_per_run_cap or 100)
                        interval = max(1, int(ws.automation_interval_days or 1))
                        plan = compute_smart_automation_plan(
                            days_elapsed,
                            applied,
                            max_per_run=cap,
                            interval_days=interval,
                            candidate_id=cand.id,
                        )
                        batch = int(plan["suggested_batch_per_run"])
                        ws.automation_batch_size = batch
                        ws.automation_max_emails = max(0, int(plan["remaining_to_cap"]))
                        if batch > best_batch:
                            best_batch = batch
                            best_ws = ws

                    if best_batch <= 0 or not best_ws:
                        for ws in wss:
                            ws.automation_next_run = now + timedelta(
                                days=max(1, int(ws.automation_interval_days or 1))
                            )
                        db.session.commit()
                        continue

                    auto_start_single_run(
                        best_ws.candidate_id,
                        limit=best_batch,
                        offset=best_ws.automation_target_index,
                        industry=best_ws.industry,
                        country=best_ws.country,
                        from_scheduler=True,
                    )
                    for ws in wss:
                        ws.automation_next_run = now + timedelta(
                            days=max(1, int(ws.automation_interval_days or 1))
                        )
                    db.session.commit()

                # 3. Workflow Plans (6-Month Service)
                try:
                    active_plans = WorkflowPlan.query.filter(WorkflowPlan.status == "active").all()
                    if active_plans:
                        print(f"[Workflow Scheduler] Checking {len(active_plans)} active plans...")
                    for plan in active_plans:
                        # Skip if already completed
                        if plan.total_applied >= plan.total_target:
                            plan.status = "completed"
                            db.session.commit()
                            continue

                        # Check if it's time to run
                        should_run = False
                        if plan.next_run_date:
                            if plan.next_run_date <= now:
                                should_run = True
                            else:
                                # Not due yet
                                pass 
                        else:
                            should_run = True  # first run

                        if not should_run:
                            continue

                        remaining_plan = max(0, int(plan.total_target or 1200) - int(plan.total_applied or 0))
                        if remaining_plan <= 0:
                            plan.status = "completed"
                            db.session.commit()
                            continue

                        # Skip execution on Sundays but advance schedule
                        if now.weekday() == 6:
                            print(f"[Workflow] Skipping Sunday for Plan #{plan.id}")
                            new_run = (plan.next_run_date or now) + timedelta(days=1)
                            while new_run <= now: new_run += timedelta(days=1)
                            plan.next_run_date = new_run
                            db.session.commit()
                            continue

                        cand_plan = db.session.get(Candidate, plan.candidate_id)
                        if cand_plan and scheduler_should_pause_automation_for_candidate_bucklist(cand_plan, now):
                            new_run = (plan.next_run_date or now) + timedelta(days=1)
                            while new_run <= now:
                                new_run += timedelta(days=1)
                            while new_run.weekday() == 6:
                                new_run += timedelta(days=1)
                            plan.next_run_date = new_run
                            db.session.commit()
                            print(
                                f"[Workflow] Plan #{plan.id}: deferred — bucklist 0–90 band (scheduled automation off)."
                            )
                            continue

                        elapsed = workflow_plan_elapsed_service_days(plan, now)
                        phase_num = plan._get_phase_for_day(elapsed)
                        phase_info = WORKFLOW_PHASES.get(phase_num, WORKFLOW_PHASES[7])
                        batch_size = workflow_plan_effective_send_batch(plan, now)

                        if batch_size <= 0:
                            new_run = (plan.next_run_date or now) + timedelta(days=1)
                            while new_run <= now:
                                new_run += timedelta(days=1)
                            while new_run.weekday() == 6:
                                new_run += timedelta(days=1)
                            plan.next_run_date = new_run
                            db.session.commit()
                            print(
                                f"[Workflow] Plan #{plan.id}: no sends (on/ahead of day curve, daily cap, or target met). Next: {new_run}"
                            )
                            continue

                        if scheduler_automation_already_used_today(plan.candidate_id, now):
                            new_run = (plan.next_run_date or now) + timedelta(days=1)
                            while new_run <= now:
                                new_run += timedelta(days=1)
                            while new_run.weekday() == 6:
                                new_run += timedelta(days=1)
                            plan.next_run_date = new_run
                            db.session.commit()
                            print(
                                f"[Workflow] Plan #{plan.id}: deferred — scheduler automation already ran today for this candidate."
                            )
                            continue

                        pend = max(0, linear_expected_applications_for_days(elapsed) - int(plan.total_applied or 0))
                        catch = "catch-up " if pend > 0 else ""
                        print(f"[Workflow] Triggering Plan #{plan.id} ({plan.candidate_name}): Phase {phase_num}, {catch}Batch {batch_size}")
                        auto_start_single_run(
                            plan.candidate_id,
                            limit=batch_size,
                            offset=0,
                            industry=plan.industry,
                            country=plan.country,
                            from_scheduler=True,
                        )

                        # Update plan
                        plan.current_phase = phase_num
                        plan.last_run_date = now

                        # Calculate next run date (advance by 1 day)
                        new_run = (plan.next_run_date or now) + timedelta(days=1)
                        # Ensure we don't fall behind if the server was off
                        while new_run <= now:
                            new_run += timedelta(days=1)
                        plan.next_run_date = new_run

                        if plan.total_applied >= plan.total_target:
                            plan.status = "completed"

                        db.session.commit()
                        print(f"[Workflow] Plan #{plan.id} updated. Next run: {plan.next_run_date}")

                except Exception as wf_err:
                    print(f"[Workflow Scheduler] Error: {wf_err}")
                    db.session.rollback()

        except Exception as e:
            print(f"[Scheduler] Critical Error: {e}")
            try:
                with app.app_context(): db.session.rollback()
            except: pass
        time.sleep(60)

def auto_start_single_run(
    candidate_id,
    limit=None,
    offset=0,
    industry=None,
    country=None,
    from_scheduler=False,
):
    """Helper to start a single candidate run (possibly a batch from a workflow)."""
    try:
        with app.app_context():
            cand = db.session.get(Candidate, candidate_id)
            if not cand: return
            if from_scheduler and scheduler_should_pause_automation_for_candidate_bucklist(
                cand, datetime.utcnow()
            ):
                print(
                    f"[Automation] Blocked scheduler send for candidate #{candidate_id} — service day band 0–90."
                )
                return

            names = cand.name
            pa = (cand.pa_member or cand.app_code or "").strip()
            po = (cand.placement_officer_member or "").strip()

            mode_label = f"batch_{limit}" if limit else "single"
            if industry: mode_label += f"_{industry}"

            run = SendRun(
                mode=mode_label,
                status="queued",
                candidate_names=names,
                pa_members=pa,
                placement_officer_members=po,
                industry=industry,
                country=country,
            )
            db.session.add(run)
            if from_scheduler:
                cand.scheduler_automation_at = datetime.utcnow()
            db.session.commit()

            run_states[str(run.id)] = {"stop": False, "pause": False}
            push_progress(str(run.id), "CONNECTED")

            delay = AUTOMATION_DEFAULT_DELAY_SECONDS
            enable_bounce = ENABLE_BOUNCE_CHECK_DEFAULT

            t = Thread(target=run_sender_thread, args=(run.id, [candidate_id], delay, enable_bounce, limit, offset, industry, country), daemon=True)
            t.start()
    except Exception as e:
        print(f"[Automation] auto_start_single_run error: {e}")


def run_workspace_automation_tick(ws: Workspace, now: Optional[datetime] = None):
    """One Smart Automation tick for a workspace (same rules as scheduler). Returns status dict."""
    now = now or datetime.utcnow()
    cand = db.session.get(Candidate, ws.candidate_id)
    interval = max(1, int(ws.automation_interval_days or 1))
    if not cand:
        ws.automation_next_run = now + timedelta(days=interval)
        db.session.commit()
        return {"ok": True, "advanced": True, "message": "No sender candidate — next run bumped."}
    svc = getattr(cand, "smart_service_start_date", None) or ws.service_start_date
    if not svc:
        svc = now
    wp = (
        WorkflowPlan.query.filter_by(candidate_id=cand.id)
        .filter(WorkflowPlan.status.in_(["paused", "active"]))
        .order_by(WorkflowPlan.id.desc())
        .first()
    )
    if wp:
        days_elapsed = workflow_plan_elapsed_service_days(wp, now)
    else:
        days_elapsed = max(0, (now - svc).days)
    applied = int(wp.total_applied or 0) if wp else int(ws.automation_total_sent or 0)
    cap = int(ws.automation_per_run_cap or 100)
    interval = max(1, int(ws.automation_interval_days or 1))
    plan_sa = compute_smart_automation_plan(
        days_elapsed, applied, max_per_run=cap, interval_days=interval, candidate_id=cand.id
    )
    batch = int(plan_sa["suggested_batch_per_run"])
    ws.automation_batch_size = batch
    ws.automation_max_emails = max(0, int(plan_sa["remaining_to_cap"]))
    if applied >= 1200 or batch <= 0:
        if applied >= 1200:
            ws.automation_enabled = False
        ws.automation_next_run = now + timedelta(days=interval)
        db.session.commit()
        return {
            "ok": True,
            "advanced": True,
            "batch": 0,
            "message": "No catch-up batch right now — schedule advanced (on track or cap reached).",
        }
    auto_start_single_run(
        ws.candidate_id,
        limit=batch,
        offset=ws.automation_target_index,
        industry=ws.industry,
        country=ws.country,
    )
    ws.automation_next_run = now + timedelta(days=interval)
    db.session.commit()
    return {"ok": True, "advanced": False, "batch": batch, "message": f"Queued batch of {batch}."}


def run_workflow_plan_batch_manual(plan: WorkflowPlan, now: Optional[datetime] = None, skip_sunday_check: bool = True):
    """Run one 6-month workflow batch (Run now / scheduler). Optionally allow Sundays."""
    now = now or datetime.utcnow()
    if (plan.status or "").lower() != "active":
        return {"ok": False, "error": "Plan is not active"}
    if plan.total_applied >= plan.total_target:
        plan.status = "completed"
        db.session.commit()
        return {"ok": False, "error": "Plan target already reached"}
    if not skip_sunday_check and now.weekday() == 6:
        new_run = (plan.next_run_date or now) + timedelta(days=1)
        while new_run <= now:
            new_run += timedelta(days=1)
        plan.next_run_date = new_run
        db.session.commit()
        return {"ok": False, "error": "Sunday — use Run now to override or wait until Monday"}
    elapsed = workflow_plan_elapsed_service_days(plan, now)
    phase_num = plan._get_phase_for_day(elapsed)
    phase_info = WORKFLOW_PHASES.get(phase_num, WORKFLOW_PHASES[7])
    batch_size = workflow_plan_effective_send_batch(plan, now, scheduled=False)
    if batch_size <= 0:
        if remaining_daily_send_quota(plan.candidate_id, now) <= 0:
            return {
                "ok": False,
                "error": "Daily automation limit (100) already reached for this candidate today. Remainder runs on the next day(s).",
            }
        return {"ok": False, "error": "Nothing to send for this plan"}
    auto_start_single_run(plan.candidate_id, limit=batch_size, offset=0, industry=plan.industry, country=plan.country)
    plan.current_phase = phase_num
    plan.last_run_date = now
    new_run = (plan.next_run_date or now) + timedelta(days=1)
    while new_run <= now:
        new_run += timedelta(days=1)
    plan.next_run_date = new_run
    if plan.total_applied >= plan.total_target:
        plan.status = "completed"
    db.session.commit()
    return {"ok": True, "batch_size": batch_size, "message": f"Queued phase batch ({batch_size})."}


@app.post("/api/runs/start")
def api_start_run():
    data = request.json or {}
    mode = data.get("mode") or "all"
    delay = int(data.get("delay", AUTOMATION_DEFAULT_DELAY_SECONDS))
    enable_bounce = bool(data.get("enableBounceCheck", ENABLE_BOUNCE_CHECK_DEFAULT))
    candidate_ids = data.get("candidateIds") or []
    industry = data.get("industry") or "Default"
    country = data.get("country") or "Global"

    if mode not in ("all", "single"):
        return jsonify({"error": "mode must be 'all' or 'single'"}), 400

    if mode == "all":
        cands = Candidate.query.order_by(Candidate.name.asc()).all()
        candidate_ids = [c.id for c in cands]
    else:
        if not candidate_ids:
            return jsonify({"error": "candidateIds required for single mode"}), 400

    if not candidate_ids:
        return jsonify({"error": "No candidates found"}), 400

    cand_objs = Candidate.query.filter(Candidate.id.in_(candidate_ids)).order_by(Candidate.name.asc()).all()
    names = ", ".join([c.name for c in cand_objs])

    pa_members = sorted({(c.pa_member or c.app_code or "").strip() for c in cand_objs if (c.pa_member or c.app_code or "").strip()})
    placement_officers = sorted({(c.placement_officer_member or "").strip() for c in cand_objs if (c.placement_officer_member or "").strip()})

    run = SendRun(
        mode=mode,
        status="queued",
        candidate_names=names,
        pa_members=", ".join(pa_members),
        placement_officer_members=", ".join(placement_officers),
        industry=industry,
        country=country,
        user_id=session.get("user_id"),
    )
    db.session.add(run)
    db.session.commit()

    run_states[str(run.id)] = {"stop": False, "pause": False}
    push_progress(str(run.id), "CONNECTED")

    t = Thread(target=run_sender_thread, args=(run.id, candidate_ids, delay, enable_bounce, None, 0, industry, country), daemon=True)
    t.start()

    return jsonify({"message": "Run started", "runId": run.id}), 200

@app.get("/api/runs/<int:run_id>/candidates/<int:candidate_id>/report")
def api_download_candidate_report(run_id, candidate_id):
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 401
    _backfill_run_reports_if_missing(run_id)
    rep = RunCandidateReport.query.filter(
        RunCandidateReport.run_id == run_id,
        RunCandidateReport.candidate_id == candidate_id
    ).first()
    if not rep or not os.path.exists(rep.report_csv_path):
        return "Report not found", 404
    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", rep.candidate_name)[:60]
    dl_name = f"run_{run_id}_{safe_name}.csv"
    return send_file(rep.report_csv_path, as_attachment=True, download_name=dl_name)

# =====================================================================
# EMAIL INTELLIGENCE SCANNER
# =====================================================================

SCANNER_ALLOWED_EXTENSIONS = {"xls", "xlsx"}

SCANNER_FREE_EMAIL_DOMAINS = {
    "gmail.com", "yahoo.com", "hotmail.com", "outlook.com", "live.com",
    "aol.com", "icloud.com", "proton.me", "protonmail.com",
}

SCANNER_HR_KEYWORDS = [
    "hr", "career", "careers", "jobs", "job", "recruit",
    "recruitment", "talent", "hiring",
]

SCANNER_EMAIL_REGEX = re.compile(
    r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}"
)

SCANNER_IMAGE_LIKE_TLDS = {
    "png", "jpg", "jpeg", "gif", "svg", "webp",
    "ico", "bmp", "tif", "tiff"
}
SCANNER_ASSET_OR_CODE_TLDS = {
    "css", "js", "map", "mp4", "webm", "woff", "woff2"
}


def _scanner_allowed_file(filename):
    name = os.path.basename(str(filename))
    return "." in name and name.rsplit(".", 1)[1].lower() in SCANNER_ALLOWED_EXTENSIONS


def _scanner_normalize_company_from_domain(domain):
    domain = domain.lower().strip()
    for prefix in ["www.", "mail.", "app.", "portal.", "webmail."]:
        if domain.startswith(prefix):
            domain = domain[len(prefix):]
    parts = domain.split(".")
    core = parts[-2] if len(parts) >= 2 else parts[0]
    words = re.split(r"[-_]", core)
    words = [w.capitalize() for w in words if w]
    return " ".join(words)


def _scanner_normalize_company_from_local_part(local_part):
    cleaned = re.sub(r"\d+", "", local_part)
    words = re.split(r"[._-]+", cleaned)
    words = [w for w in words if w]
    if not words:
        return local_part.capitalize()
    return " ".join(w.capitalize() for w in words)


def _scanner_is_hr_email(local_part):
    lp = local_part.lower()
    return any(kw in lp for kw in SCANNER_HR_KEYWORDS)


def _scanner_looks_like_tracking_or_hash(local):
    local = local.strip().lower()
    if len(local) >= 24 and all(c in "0123456789abcdef" for c in local):
        return True
    return False


def _scanner_validate_email_address(email_addr):
    email_addr = email_addr.strip()
    if not SCANNER_EMAIL_REGEX.fullmatch(email_addr):
        return False, "Invalid format"
    if "@" not in email_addr:
        return False, "Missing @"
    local, domain = email_addr.split("@", 1)
    if "." not in domain:
        return False, "Domain has no dot"
    if domain.startswith(".") or domain.endswith("."):
        return False, "Domain starts/ends with dot"
    if not local:
        return False, "Missing local part"
    tld = domain.rsplit(".", 1)[-1].lower()
    if tld in SCANNER_IMAGE_LIKE_TLDS or tld in SCANNER_ASSET_OR_CODE_TLDS:
        return False, "Looks like file/asset, not mailbox"
    if _scanner_looks_like_tracking_or_hash(local):
        return False, "Looks like tracking/hash id"
    return True, "Valid (format only)"


def _scanner_classify_email(email_addr, page_company_name=None):
    email_addr = email_addr.strip()
    if "@" not in email_addr:
        return {}
    local, domain = email_addr.split("@", 1)
    domain = domain.lower()
    is_free = domain in SCANNER_FREE_EMAIL_DOMAINS
    if is_free:
        company_guess = _scanner_normalize_company_from_local_part(local)
        domain_type = "free"
    else:
        company_guess = page_company_name if page_company_name else _scanner_normalize_company_from_domain(domain)
        domain_type = "corporate"
    hr_flag = _scanner_is_hr_email(local)
    is_valid, reason = _scanner_validate_email_address(email_addr)
    return {
        "email": email_addr,
        "company_name_guess": company_guess,
        "domain": domain,
        "domain_type": domain_type,
        "is_hr_email": hr_flag,
        "is_valid": is_valid,
        "validation_reason": reason,
    }


def _scanner_fetch_page(url):
    if not isinstance(url, str) or not url.strip():
        return None
    url = url.strip()
    if "@" in url:
        return None
    if not url.startswith(("http://", "https://")):
        url = "https://" + url
    try:
        headers = {"User-Agent": "Mozilla/5.0 (compatible; EmailScraper/1.0)"}
        resp = http_requests.get(url, headers=headers, timeout=12)
        resp.raise_for_status()
        return resp.text
    except Exception:
        return None


def _scanner_get_page_company_name(html, url):
    try:
        soup = BeautifulSoup(html, "html.parser")
        og_site = soup.find("meta", property="og:site_name")
        if og_site and og_site.get("content"):
            return og_site["content"].strip()
        if soup.title and soup.title.string:
            title_text = soup.title.string.strip()
            if " - " in title_text:
                title_text = title_text.split(" - ")[0].strip()
            elif "|" in title_text:
                title_text = title_text.split("|")[0].strip()
            return title_text
    except Exception:
        pass
    try:
        parsed = urlparse(url)
        domain = parsed.netloc or parsed.path
        return _scanner_normalize_company_from_domain(domain)
    except Exception:
        return None


def _scanner_extract_emails_from_html(html):
    emails = set()
    if not html:
        return emails
    soup = BeautifulSoup(html, "html.parser")
    for a in soup.find_all("a", href=True):
        href = a["href"].strip()
        if href.lower().startswith("mailto:"):
            addr = href.split(":", 1)[1]
            addr = addr.split("?", 1)[0].strip()
            if SCANNER_EMAIL_REGEX.fullmatch(addr):
                local, domain = addr.split("@", 1)
                tld = domain.rsplit(".", 1)[-1].lower()
                if tld in SCANNER_IMAGE_LIKE_TLDS or tld in SCANNER_ASSET_OR_CODE_TLDS:
                    continue
                if _scanner_looks_like_tracking_or_hash(local):
                    continue
                emails.add(addr)
    text = soup.get_text(" ", strip=True)
    for match in SCANNER_EMAIL_REGEX.findall(text):
        local, domain = match.split("@", 1)
        tld = domain.rsplit(".", 1)[-1].lower()
        if tld in SCANNER_IMAGE_LIKE_TLDS or tld in SCANNER_ASSET_OR_CODE_TLDS:
            continue
        if _scanner_looks_like_tracking_or_hash(local):
            continue
        emails.add(match)
    return emails


def _scanner_detect_columns(df):
    company_col = website_col = country_col = category_col = email_col = None
    for col in df.columns:
        lc = str(col).lower()
        if company_col is None and ("company" in lc or lc in ("name", "business name")):
            company_col = col
        if website_col is None and ("website" in lc or lc in ("url", "link", "website url", "site")):
            website_col = col
        if country_col is None and ("country" in lc or "nation" in lc):
            country_col = col
        if category_col is None and ("category" in lc or "industry" in lc or "segment" in lc):
            category_col = col
        if email_col is None and ("email" in lc or "e-mail" in lc):
            email_col = col
    if website_col is None and len(df.columns) >= 1:
        website_col = df.columns[0]
    if company_col is None and len(df.columns) > 1:
        company_col = df.columns[1]
    return company_col, website_col, country_col, category_col, email_col


def _scanner_process_excel(input_path, base_output_name):
    try:
        df = pd.read_excel(input_path)
    except Exception:
        return None, None, None, None, None, None

    output_columns = [
        "Excel_Company_Name", "Website", "Country", "Category",
        "Scraped_Email", "Company_Name_Guess", "Domain",
        "Domain_Type", "Is_HR_Email", "Is_Valid", "Validation_Reason",
    ]

    if df.empty:
        df_all = pd.DataFrame(columns=output_columns)
        df_valid = df_all.copy()
        df_invalid = df_all.copy()
    else:
        company_col, website_col, country_col, category_col, email_col = _scanner_detect_columns(df)
        output_rows = []
        seen_pairs = set()

        for _, row in df.iterrows():
            excel_company = str(row[company_col]) if company_col is not None and pd.notna(row[company_col]) else ""
            website_val = row[website_col] if website_col in df.columns else ""
            country = str(row[country_col]) if country_col is not None and pd.notna(row[country_col]) else ""
            category = str(row[category_col]) if category_col is not None and pd.notna(row[category_col]) else ""

            existing_email = None
            if email_col is not None and pd.notna(row[email_col]):
                existing_email = str(row[email_col]).strip()

            email_candidates = set()
            if existing_email:
                email_candidates.add(existing_email)

            page_company_name = None
            website_str = ""
            if isinstance(website_val, str) and website_val.strip():
                website_str = website_val.strip()
                if "@" not in website_str:
                    html = _scanner_fetch_page(website_str)
                    if html:
                        page_company_name = _scanner_get_page_company_name(html, website_str)
                        scraped_emails = _scanner_extract_emails_from_html(html)
                        email_candidates.update(scraped_emails)
                else:
                    website_str = ""

            if not email_candidates:
                continue

            for em in email_candidates:
                info = _scanner_classify_email(em, page_company_name)
                if not info:
                    continue
                key = (excel_company, em)
                if key in seen_pairs:
                    continue
                seen_pairs.add(key)
                output_rows.append({
                    "Excel_Company_Name": excel_company,
                    "Website": website_str,
                    "Country": country,
                    "Category": category,
                    "Scraped_Email": info["email"],
                    "Company_Name_Guess": info["company_name_guess"],
                    "Domain": info["domain"],
                    "Domain_Type": info["domain_type"],
                    "Is_HR_Email": "YES" if info["is_hr_email"] else "NO",
                    "Is_Valid": "YES" if info["is_valid"] else "NO",
                    "Validation_Reason": info["validation_reason"],
                })

        df_all = pd.DataFrame(output_rows, columns=output_columns) if output_rows else pd.DataFrame(columns=output_columns)
        df_valid = df_all[df_all["Is_Valid"] == "YES"].copy()
        df_invalid = df_all[df_all["Is_Valid"] == "NO"].copy()

    safe_base = secure_filename(os.path.splitext(base_output_name)[0]) or "result"
    file_all = f"{safe_base}_ALL.xlsx"
    file_valid = f"{safe_base}_VALID.xlsx"
    file_invalid = f"{safe_base}_INVALID.xlsx"

    try:
        df_all.to_excel(os.path.join(SCANNER_OUTPUT_FOLDER, file_all), index=False)
        df_valid.to_excel(os.path.join(SCANNER_OUTPUT_FOLDER, file_valid), index=False)
        df_invalid.to_excel(os.path.join(SCANNER_OUTPUT_FOLDER, file_invalid), index=False)
    except Exception:
        return df_all, df_valid, df_invalid, None, None, None

    return df_all, df_valid, df_invalid, file_all, file_valid, file_invalid


# --- Scanner Routes ---

@app.get("/scanner")
@app.post("/scanner")
def scanner_page():
    if "user_id" not in session:
        return redirect(url_for("login_page"))
    if request.method == "POST":
        if "file" not in request.files:
            flash("No file part in request.")
            return redirect("/scanner")
        file = request.files["file"]
        if file.filename == "":
            flash("No file selected.")
            return redirect("/scanner")
        if not _scanner_allowed_file(file.filename):
            flash("Please upload a valid Excel file (.xls or .xlsx).")
            return redirect("/scanner")

        original_name = file.filename
        filename = secure_filename(os.path.basename(original_name)) or "upload.xlsx"
        upload_path = os.path.join(SCANNER_UPLOAD_FOLDER, filename)
        try:
            file.save(upload_path)
        except Exception as e:
            flash(f"Error saving uploaded file: {e}")
            return redirect("/scanner")

        df_all, df_valid, df_invalid, file_all, file_valid, file_invalid = _scanner_process_excel(
            upload_path, filename
        )
        if df_all is None:
            flash("Failed to read or process the Excel file. Please check the file and try again.")
            return redirect("/scanner")

        return render_template(
            "terra.html",
            records_all=df_all.to_dict(orient="records"),
            records_valid=df_valid.to_dict(orient="records"),
            records_invalid=df_invalid.to_dict(orient="records"),
            file_all=file_all,
            file_valid=file_valid,
            file_invalid=file_invalid,
            total_all=len(df_all),
            total_valid=len(df_valid),
            total_invalid=len(df_invalid),
        )

    # GET
    return render_template(
        "terra.html",
        records_all=None,
        records_valid=None,
        records_invalid=None,
        file_all=None,
        file_valid=None,
        file_invalid=None,
        total_all=0,
        total_valid=0,
        total_invalid=0,
    )


@app.get("/scanner/download/<path:filename>")
def scanner_download_file(filename):
    safe_name = secure_filename(os.path.basename(filename))
    if not safe_name:
        flash("Invalid file requested.")
        return redirect("/scanner")
    file_path = os.path.join(SCANNER_OUTPUT_FOLDER, safe_name)
    if not os.path.exists(file_path):
        flash("Requested file does not exist.")
        return redirect("/scanner")
    try:
        return send_from_directory(SCANNER_OUTPUT_FOLDER, safe_name, as_attachment=True)
    except Exception as e:
        flash(f"Error while sending file: {e}")
        return redirect("/scanner")


# ─── Role-Based Candidate Analysis ────────────────────────────────────

def _split_name(full_name):
    """Split a single name field into (first_name, last_name)."""
    parts = (full_name or "").strip().split()
    if not parts:
        return ("", "")
    return (parts[0], " ".join(parts[1:]))


def _compute_role_groups(unique_roles, custom_groups=None):
    """
    Computes a grouping of roles.
    1. Mandatory 'custom_groups' provided by the user (as folders).
    2. Smart grouping of the remaining roles.
    """
    import re
    custom_groups = custom_groups or []

    # Dominant markers mapping
    # Note: These are checked in order.
    # We use a list of tuples to maintain priority.
    DOMINANT_MARKERS = [
        ('devsecops', 'DevOps Engineer'),
        ('devops', 'DevOps Engineer'),
        ('account', 'Account'),
        ('accounting', 'Account'),
        ('finance', 'Account'),
        ('audit', 'Account'),
        ('tax', 'Account'),
        ('frontend', 'Frontend Engineer'),
        ('backend', 'Backend Engineer'),
        ('full stack', 'Full Stack Developer'),
        ('fullstack', 'Full Stack Developer'),
        ('software', 'Software Engineer'),
        ('python', 'Python Developer'),
        ('java', 'Java Developer'),
        ('node', 'Node.js Developer'),
        ('react', 'React Developer'),
        ('cloud', 'Cloud Engineer'),
        ('data', 'Data Professional'),
        ('security', 'Security Specialist'),
        ('mobile', 'Mobile Developer'),
        ('android', 'Mobile Developer'),
        ('ios', 'Mobile Developer'),
        ('hr', 'Human Resources'),
        ('talent', 'Human Resources'),
        ('recruitment', 'Human Resources'),
        ('sales', 'Sales Professional'),
        ('marketing', 'Marketing'),
        ('product', 'Product Management'),
        ('project', 'Project Management'),
        ('ux', 'UX Designer'),
        ('user experience', 'UX Designer'),
        ('ui', 'UI Designer')
    ]

    NOISE_WORDS = {
        'senior', 'lead', 'staff', 'principal', 'junior', 'associate', 
        'trainee', 'graduate', 'level', 'iii', 'ii', 'i', 'specialist',
        'manager', 'analyst', 'consultant', 'coordinator', 'officer',
        'executive', 'professional', 'team', 'head', 'vp', 'director'
    }

    def get_bucket(role):
        # Normalize: replace separators and parentheses with space
        s = re.sub(r"[\(\)\-\–\—\/\|]", " ", role).lower().strip()
        
        # 1. Check for Dominant Markers
        # For 'devops' and 'account', we use substring match to be inclusive
        for key, label in DOMINANT_MARKERS:
            if key in ['devops', 'devsecops', 'account']:
                if key in s:
                    return label
            else:
                # For others, use word boundaries for safety
                if re.search(r'\b' + re.escape(key) + r'\b', s):
                    return label
        
        # 2. Fallback: Cleaned identity (strip noise words)
        words = s.split()
        if not words: return "Other"
        filtered = [w for w in words if w not in NOISE_WORDS]
        if not filtered: filtered = words
        return " ".join(filtered).title()

    # 1. Map roles to buckets
    bucket_to_roles = {}
    for role in unique_roles:
        # If the role name matches a custom group name, it belongs there
        if role in custom_groups:
            bucket_to_roles.setdefault(role, []).append(role)
            continue
        
        bucket = get_bucket(role)
        bucket_to_roles.setdefault(bucket, []).append(role)

    # 2. Final assignment
    groups = {}
    # Ensure all custom groups exist even if empty
    for cg in custom_groups:
        groups[cg] = bucket_to_roles.get(cg, [])

    for bucket, members in bucket_to_roles.items():
        if bucket not in groups:
            groups[bucket] = members

    # Sort each group's members
    for g in groups:
        groups[g].sort(key=str.lower)

    # DEBUG
    print(f"[_compute_role_groups] groups: {list(groups.keys())[:10]}... (Total: {len(groups)})")
    return groups


ROLE_GROUP_OVERRIDES_FILE = os.path.join(BASE_DIR, "config", "role_group_overrides.json")

def _load_role_group_overrides():
    try:
        if os.path.exists(ROLE_GROUP_OVERRIDES_FILE):
            with open(ROLE_GROUP_OVERRIDES_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
    except Exception:
        pass
    return {}

def _save_role_group_overrides(overrides):
    with open(ROLE_GROUP_OVERRIDES_FILE, "w", encoding="utf-8") as f:
        json.dump(overrides, f, indent=2, ensure_ascii=False)


@app.post("/api/role-groups/override")
def api_role_groups_override():
    data = request.json or {}
    overrides = data.get("overrides", {})
    if not isinstance(overrides, dict):
        return jsonify({"error": "overrides must be a dict"}), 400
    existing = _load_role_group_overrides()
    existing.update(overrides)
    _save_role_group_overrides(existing)
    return jsonify({"status": "ok", "count": len(existing)})


@app.get("/api/role-groups/override")
def api_role_groups_override_get():
    return jsonify(_load_role_group_overrides())


@app.delete("/api/role-groups/override")
def api_role_groups_override_delete():
    _save_role_group_overrides({})
    return jsonify({"status": "cleared"})


@app.post("/api/role-groups/custom")
def api_role_groups_custom_add():
    data = request.json or {}
    name = data.get("name", "").strip()
    if not name:
        return jsonify({"error": "name required"}), 400
    overrides = _load_role_group_overrides()
    custom_groups = overrides.setdefault("custom_groups", [])
    if name not in custom_groups:
        custom_groups.append(name)
    _save_role_group_overrides(overrides)
    return jsonify({"status": "ok", "custom_groups": custom_groups})


@app.delete("/api/role-groups/custom")
def api_role_groups_custom_delete():
    data = request.get_json(force=True, silent=True) or request.form or request.args or {}
    if isinstance(data, dict):
        name = (data.get("name") or "").strip()
    else:
        name = ""
    if not name:
        return jsonify({"error": "name required"}), 400
    overrides = _load_role_group_overrides()
    custom_groups = overrides.setdefault("custom_groups", [])
    if name in custom_groups:
        custom_groups.remove(name)
    c_overrides = overrides.setdefault("candidate_overrides", {})
    if not isinstance(c_overrides, dict):
        c_overrides = {}
        overrides["candidate_overrides"] = c_overrides
    to_delete_c = [cid for cid, g in list(c_overrides.items()) if g == name]
    for cid in to_delete_c:
        del c_overrides[cid]
    # Remove role -> folder mappings that pointed at this folder (same keys as drag-move)
    _ROLE_OVERRIDE_RESERVED = frozenset({"custom_groups", "candidate_overrides"})
    removed_roles = []
    for key in list(overrides.keys()):
        if key in _ROLE_OVERRIDE_RESERVED:
            continue
        val = overrides[key]
        if isinstance(val, str) and val.strip() == name:
            del overrides[key]
            removed_roles.append(key)
    _save_role_group_overrides(overrides)
    return jsonify({"status": "ok", "cleared_candidates": len(to_delete_c), "cleared_role_mappings": len(removed_roles)})


@app.post("/api/role-groups/move")
def api_role_groups_move():
    data = request.json or {}
    candidate_id = data.get("candidate_id")
    target_group = data.get("target_group")
    if not candidate_id or not target_group:
        return jsonify({"error": "candidate_id and target_group required"}), 400
    
    overrides = _load_role_group_overrides()
    c_overrides = overrides.setdefault("candidate_overrides", {})
    if target_group == "auto":
        if str(candidate_id) in c_overrides:
            del c_overrides[str(candidate_id)]
    else:
        c_overrides[str(candidate_id)] = target_group
    
    _save_role_group_overrides(overrides)
    return jsonify({"status": "ok"})


@app.get("/api/role-analysis")
def api_role_analysis():
    """Return all candidates grouped by role.
    Each candidate appears under every role from their roles_text.
    Includes smart grouping of similar roles by common prefix.
    """
    cands = Candidate.query.order_by(Candidate.id.asc()).all()
    roles_map = {}  # role_name -> [candidate_info, ...]
    flat_rows = []  # for the flat table view

    overrides = _load_role_group_overrides()
    candidate_overrides = overrides.get("candidate_overrides", {})
    custom_groups = overrides.get("custom_groups", [])

    for c in cands:
        first_name, last_name = _split_name(c.name)
        
        # Check for manual candidate override
        cid_str = str(c.id)
        if cid_str in candidate_overrides:
            roles = [candidate_overrides[cid_str]]
        else:
            roles = parse_roles(c.roles_text or "")
            if not roles:
                roles = ["Unassigned"]
        
        for role in roles:
            entry = {
                "candidate_id": c.id,
                "first_name": first_name,
                "last_name": last_name,
                "email": c.email or "",
                "resume_path": c.resume_path or "",
                "enrollment_id": c.enrollment_id or "",
                "industry_types": (c.industry_types or "").strip(),
                "role": role,
            }
            roles_map.setdefault(role, []).append(entry)
            flat_rows.append(entry)

    # Collect unique role names sorted
    unique_roles = sorted(roles_map.keys(), key=str.lower)

    # Smart grouping
    groups = _compute_role_groups(unique_roles, custom_groups=custom_groups)
    group_names = sorted(groups.keys(), key=str.lower)

    return jsonify({
        "unique_roles": unique_roles,
        "total_roles": len(unique_roles),
        "total_mappings": len(flat_rows),
        "total_candidates": len(cands),
        "roles": roles_map,
        "flat": flat_rows,
        "groups": groups,
        "group_names": group_names,
        "total_groups": len(group_names),
        "custom_groups": custom_groups,
        "candidate_overrides": candidate_overrides,
    })


@app.get("/api/role-analysis/export")
def api_role_analysis_export():
    """Export role-based candidate analysis as an Excel file.
    Includes both Individual Roles and Smart Groups.
    """
    cands = Candidate.query.order_by(Candidate.id.asc()).all()
    roles_map = {}  # role_name -> [row data]

    overrides = _load_role_group_overrides()
    candidate_overrides = overrides.get("candidate_overrides", {})
    custom_groups = overrides.get("custom_groups", [])

    for c in cands:
        first_name, last_name = _split_name(c.name)
        
        # Check for manual candidate override
        cid_str = str(c.id)
        if cid_str in candidate_overrides:
            roles = [candidate_overrides[cid_str]]
        else:
            roles = parse_roles(c.roles_text or "")
            if not roles:
                roles = ["Unassigned"]

        for role in roles:
            roles_map.setdefault(role, []).append({
                "first_name": first_name,
                "last_name": last_name,
                "email": c.email or "",
                "resume_path": c.resume_path or "",
                "enrollment_id": c.enrollment_id or "",
                "industry_types": (c.industry_types or "").strip(),
            })

    # Smart Grouping for export
    unique_roles = sorted(roles_map.keys(), key=str.lower)
    groups = _compute_role_groups(unique_roles, custom_groups=custom_groups)
    group_names = sorted(groups.keys(), key=str.lower)

    wb = openpyxl.Workbook()
    
    # 1. Summary Sheet (Individual Roles)
    ws_summary = wb.active
    ws_summary.title = "Roles Summary"
    ws_summary.append(["Role", "Total Candidates"])
    for role_name in unique_roles:
        ws_summary.append([role_name, len(roles_map[role_name])])

    # 2. Groups Summary Sheet
    ws_grp_summary = wb.create_sheet(title="Groups Summary")
    ws_grp_summary.append(["Smart Group", "Member Roles", "Unique Candidates"])
    for g_name in group_names:
        sub_roles = groups[g_name]
        # Deduplicate candidates across sub-roles in this group
        seen_emails = set()
        count = 0
        for sr in sub_roles:
            for entry in roles_map[sr]:
                if entry["email"].lower() not in seen_emails:
                    seen_emails.add(entry["email"].lower())
                    count += 1
        ws_grp_summary.append([g_name, ", ".join(sub_roles), count])

    # 3. Individual Role Sheets (limit to 30 most populated to avoid too many sheets)
    sorted_roles_by_pop = sorted(unique_roles, key=lambda x: len(roles_map[x]), reverse=True)
    headers = ["First Name", "Last Name", "Email", "Industry Types", "Resume Path", "Enrollment ID"]
    
    for role_name in sorted_roles_by_pop[:30]:
        safe_name = re.sub(r"[\\/*?\[\]:]", "", role_name)[:31] or "Role"
        ws = wb.create_sheet(title=safe_name)
        ws.append(headers)
        for entry in roles_map[role_name]:
            ws.append([entry["first_name"], entry["last_name"], entry["email"], entry.get("industry_types") or "", entry["resume_path"], entry["enrollment_id"]])

    # 4. Smart Group Sheets (one sheet per group if group has >1 role)
    for g_name in group_names:
        sub_roles = groups[g_name]
        if len(sub_roles) <= 1 and g_name in unique_roles:
            continue # Already covered in role sheets or redundant
        
        # Group sheet deduplicates candidates
        safe_gname = re.sub(r"[\\/*?\[\]:]", "", "Group " + g_name)[:31]
        # ensure unique sheet name
        if safe_gname in wb.sheetnames:
            safe_gname = safe_gname[:25] + "_" + str(hash(g_name))[:4]

        ws = wb.create_sheet(title=safe_gname)
        ws.append(["Original Role"] + headers)
        
        seen_emails = set()
        group_entries = []
        for sr in sub_roles:
            for entry in roles_map[sr]:
                if entry["email"].lower() not in seen_emails:
                    seen_emails.add(entry["email"].lower())
                    ws.append([sr, entry["first_name"], entry["last_name"], entry["email"], entry.get("industry_types") or "", entry["resume_path"], entry["enrollment_id"]])

    # 5. All Roles (Flat)
    ws_flat = wb.create_sheet(title="All Roles (Flat)")
    ws_flat.append(["Role"] + headers)
    for role_name in unique_roles:
        for entry in roles_map[role_name]:
            ws_flat.append([role_name, entry["first_name"], entry["last_name"], entry["email"], entry.get("industry_types") or "", entry["resume_path"], entry["enrollment_id"]])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return send_file(
        buf,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name=f"role_analysis_{ts}.xlsx"
    )


@app.route("/api/automation/download-template")
def download_batch_template():
    template_path = os.path.join(BASE_DIR, "static", "templates", "batch_application_template.xlsx")
    if not os.path.exists(template_path):
        return jsonify({"error": "Template file not found. Ensure static/templates/batch_application_template.xlsx exists."}), 404
    return send_file(template_path, as_attachment=True, download_name="batch_job_applications.xlsx")

# --- Selenium Script Support Routes ---

@app.get("/jsa-enrollment/index")
def jsa_enrollment_index():
    if "user_id" not in session:
        return redirect(url_for("login_page"))
    return "JSA Enrollment Index (Authenticated)"

@app.get("/jsa-enrollment/update")
def jsa_enrollment_update():
    if "user_id" not in session:
        return redirect(url_for("login_page"))
    # Return a minimal HTML that has the elements the Selenium script expects
    return """
    <!DOCTYPE html>
    <html>
    <head>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.3/dist/css/bootstrap.min.css" rel="stylesheet">
    </head>
    <body class="p-4 bg-light">
        <h3>JSA Enrollment Update Mock</h3>
        <p>This is a helper page for the Batch Job Application tool.</p>
        <button id="create-job" class="btn btn-primary">Add Job</button>
        
        <!-- Mock Modal Container -->
        <div id="modal-backdrop" class="modal-backdrop fade" style="display:none"></div>
        <div id="job-modal" class="modal fade" style="display:none">
          <div class="modal-dialog">
            <div class="modal-content p-4">
              <h5>Add Job Application</h5>
              <div class="mb-3">
                <label class="form-label">Country</label>
                <input type="text" name="country" class="form-control" placeholder="e.g. Germany">
              </div>
              <div class="mb-3">
                <label class="form-label">Company Name</label>
                <input type="text" name="company" class="form-control">
              </div>
              <div class="mb-3">
                <label class="form-label">Job Role</label>
                <input type="text" name="role" class="form-control">
              </div>
              <div class="mb-3">
                <label class="form-label">Applied Date</label>
                <input type="date" name="date" class="form-control">
              </div>
              <div class="mb-3">
                <label class="form-label">Screenshot</label>
                <input type="file" name="screenshot" class="form-control">
              </div>
              <button id="save" class="btn btn-success btn-primary">Save</button>
            </div>
          </div>
        </div>

        <script>
            const modal = document.getElementById('job-modal');
            const backdrop = document.getElementById('modal-backdrop');
            
            document.getElementById('create-job').onclick = () => {
                modal.classList.add('show');
                modal.style.display = 'block';
                backdrop.classList.add('show');
                backdrop.style.display = 'block';
                document.body.classList.add('modal-open');
            };

            document.getElementById('save').onclick = async () => {
                const data = {
                    candidate_id: new URLSearchParams(window.location.search).get('id'),
                    country: document.querySelector('[name="country"]').value,
                    company_name: document.querySelector('[name="company"]').value,
                    job_role: document.querySelector('[name="role"]').value,
                    applied_date: document.querySelector('[name="date"]').value,
                    screenshot_path: document.querySelector('[name="screenshot"]').value
                };
                
                const response = await fetch('/api/add-job', {
                    method: 'POST',
                    headers: {'Content-Type': 'application/json'},
                    body: JSON.stringify(data)
                });
                
                if (response.ok) {
                    modal.classList.remove('show');
                    modal.style.display = 'none';
                    backdrop.classList.remove('show');
                    backdrop.style.display = 'none';
                    document.body.classList.remove('modal-open');
                    alert('Job application saved successfully!');
                } else {
                    alert('Error saving application');
                }
            };
        </script>
    </body>
    </html>
    """

@app.post("/api/add-job")
def api_add_job():
    data = request.json or {}
    cand_id = data.get("candidate_id")
    if not cand_id:
        return jsonify({"error": "candidate_id required"}), 400
    
    app_record = JobApplication(
        candidate_id=cand_id,
        company_name=data.get("company_name", ""),
        job_role=data.get("job_role", ""),
        country=data.get("country", ""),
        applied_date=data.get("applied_date", ""),
        screenshot_path=data.get("screenshot_path", "")
    )
    db.session.add(app_record)
    db.session.commit()
    return jsonify({"status": "ok", "id": app_record.id})

@app.get("/api/candidates/<int:candidate_id>/job-applications")
def get_job_applications(candidate_id):
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 401
    
    apps = JobApplication.query.filter_by(candidate_id=candidate_id).order_by(JobApplication.created_at.desc()).all()
    return jsonify([a.to_dict() for a in apps])

@app.get("/api/candidates/<int:candidate_id>/email-events")
def get_candidate_email_events(candidate_id):
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 401
    
    events = EmailEvent.query.filter_by(candidate_id=candidate_id).order_by(EmailEvent.received_at.desc(), EmailEvent.created_at.desc()).all()
    return jsonify([e.to_dict() for e in events])

@app.get("/api/results/stats")
def api_get_results_stats():
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 401

    assessments = EmailEvent.query.filter_by(event_type="Assessment").count()
    interviews = EmailEvent.query.filter_by(event_type="Interview").count()
    offers = EmailEvent.query.filter_by(event_type="Offer").count()

    return jsonify({
        "assessments": assessments,
        "interviews": interviews,
        "offers": offers
    })

@app.delete("/api/email-events/<int:event_id>")
def api_delete_email_event(event_id):
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 401
    ev = db.session.get(EmailEvent, event_id)
    if not ev:
        return jsonify({"error": "Event not found"}), 404
    db.session.delete(ev)
    db.session.commit()
    return jsonify({"message": "Event deleted."})

@app.get("/api/results/milestones")
def api_get_results_milestones():
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 401
    
    print("[API] Fetching all milestone email events...")
    # 1. Pre-calculate name-to-staff map for all candidates as a smart fallback
    # This handles cases where milestones may be linked to shell candidate records 
    # but the main record with PO info shares the same name.
    all_cands = Candidate.query.order_by(Candidate.id.desc()).all()
    name_to_staff = {}
    for c in all_cands:
        name_key = (c.name or "").strip().lower()
        if not name_key:
            continue
        staff = (c.placement_officer_member or c.rm_member or c.pa_member or "").strip()
        if staff and name_key not in name_to_staff:
            name_to_staff[name_key] = staff

    # 2. Join EmailEvent with Candidate to get candidate name and direct staff info
    results = db.session.query(
        EmailEvent, 
        Candidate.name, 
        Candidate.placement_officer_member,
        Candidate.rm_member,
        Candidate.pa_member
    ).outerjoin(
        Candidate, EmailEvent.candidate_id == Candidate.id
    ).order_by(EmailEvent.received_at.desc(), EmailEvent.created_at.desc()).all()
    
    data = []
    for event_obj, cand_name, po_member, rm_member, pa_member in results:
        d = event_obj.to_dict()
        display_name = cand_name or "Unknown Candidate"
        d["candidate_name"] = display_name
        
        # Primary: Use direct join PO, fallback to RM, then PA
        po = (po_member or "").strip()
        rm = (rm_member or "").strip()
        pa = (pa_member or "").strip()
        staff_final = po or rm or pa
        
        # Smart Fallback: If still missing, look up by name
        if not staff_final:
            name_key = display_name.strip().lower()
            staff_final = name_to_staff.get(name_key, "")
        
        d["placement_officer"] = staff_final or "-"
        data.append(d)
        
    return jsonify(data)

@app.get("/api/candidates/<int:candidate_id>/app-count")
def api_get_candidate_app_count(candidate_id):
    count = JobApplication.query.filter_by(candidate_id=candidate_id).count()
    return jsonify({"count": count})

@app.post("/api/candidates/<int:candidate_id>/sync-apps")
def api_sync_candidate_apps(candidate_id):
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 401
    
    # Run sync in background or wait? Wait for now to show immediate result.
    res = sync_candidate_applications(candidate_id)
    if "failed" in res.lower() or "missing" in res.lower():
        return jsonify({"error": res}), 400
    return jsonify({"message": res})

def sync_candidate_applications(candidate_id):
    """
    Manual Email Sync: Scans inbox for Assessment, Interview, and Job Offer emails only.
    Only runs when explicitly triggered by the user pressing Sync.
    """
    imap = None
    try:
        candidate = db.session.get(Candidate, candidate_id)
        if not candidate or not candidate.email or not candidate.app_password:
            return "Missing candidate email or app password"

        print(f"[Email Sync] Scanning inbox for {candidate.name} ({candidate.email})...")

        # Clear existing events so each sync gives a fresh accurate count
        EmailEvent.query.filter_by(candidate_id=candidate_id).delete()
        db.session.commit()

        imap = imaplib.IMAP4_SSL(IMAP_HOST)
        imap.login(candidate.email, candidate.app_password)
        imap.select("INBOX")

# ── Subject-level keywords (high confidence) ─────────────────────────
        ASSESSMENT_SUBJ = [
            "assessment invitation", "assessment invite", "assessment scheduled",
            "assessment assigned", "assessment link", "assessment details shared",
            "online assessment", "technical assessment", "coding assessment",
            "skill assessment", "pre-employment assessment", "pre interview assessment",
            "pre screening assessment", "screening assessment", "evaluation test",
            "online test", "coding test", "programming test", "aptitude test",
            "logical reasoning test", "technical test", "domain assessment",
            "skill evaluation", "assessment round", "assessment stage",
            "assessment phase", "test invitation", "test invite", "test scheduled",
            "test assigned", "test link", "test details shared",
            "coding challenge", "coding challenge invitation", "hackathon invitation",
            "challenge invitation", "challenge link", "assessment deadline",
            "submission deadline", "time bound assessment", "assessment window",
            "assessment validity", "assessment expires", "take home assignment",
            "take home test", "take home assessment", "home assignment",
            "case study assignment", "case study assessment", "business case assignment",
            "practical assignment", "technical assignment",
            "hackerrank", "codility", "hackerearth", "hirevue", "testgorilla",
            "technical challenge",
        ]
        INTERVIEW_SUBJ = [
            "interview scheduled", "interview invitation", "interview invite",
            "invited for interview", "shortlisted for interview",
            "profile shortlisted", "resume shortlisted", "application shortlisted",
            "shortlisting confirmation", "screening shortlisted",
            "selected for interview", "selected for interview round",
            "interview confirmed", "interview confirmation",
            "interview slot confirmed", "interview time confirmed",
            "interview date confirmed", "interview arranged", "interview setup",
            "interview setup completed", "interview planned", "interview booked",
            "interview booking confirmation", "next round interview",
            "next interview round", "next stage interview", "next step interview",
            "technical interview scheduled", "technical round scheduled",
            "hr interview scheduled", "hr round scheduled",
            "managerial interview scheduled", "panel interview scheduled",
            "final round interview scheduled", "screening call scheduled",
            "initial screening scheduled", "phone interview scheduled",
            "telephonic interview scheduled", "video interview scheduled",
            "virtual interview scheduled", "online interview scheduled",
            "in-person interview scheduled", "onsite interview scheduled",
            "face to face interview scheduled", "walk in interview",
            "walk-in interview invitation",
            "interview link", "meeting link for interview",
            "google meet interview", "zoom interview", "microsoft teams interview",
            "interview details shared", "interview instructions shared",
            "interview agenda shared", "interview rescheduled",
            "interview postponed", "interview delayed",
            "interview reminder", "reminder for interview",
            "follow-up interview schedule",
            "phone interview scheduled", "video interview scheduled",
            "virtual interview scheduled", "technical interview",
            "panel interview scheduled", "final round interview", "phone screen scheduled",
        ]
        OFFER_SUBJ = [
            "offer letter", "job offer", "offer of employment", "employment offer",
            "formal offer", "official offer", "conditional offer", "final offer",
            "offer release", "offer has been released", "offer issued",
            "offer extended", "offer made", "offer approved", "offer confirmed",
            "offer finalized", "offer letter attached", "offer acceptance",
            "selected for the role", "you have been selected", "you are selected",
            "selection confirmed", "successfully selected", "successful candidate",
            "joining confirmation", "confirmation of employment",
            "confirmation of role", "employment confirmation", "joining letter",
            "joining date confirmed", "date of joining", "expected date of joining",
            "start date confirmed", "onboarding details", "onboarding initiation",
            "pre-onboarding", "background verification initiated",
            "background check initiated", "document verification initiated",
            "probation period", "terms and conditions", "employment terms",
            "compensation details", "salary details", "package details",
            "ctc details", "total compensation", "benefits package",
            "remuneration", "employment agreement", "employment contract",
            "contract attached", "contingent offer",
            "congratulations", "welcome to the team", "joining date",
        ]

# ── Body-level keywords (lower confidence, only if subject missed) ───
        ASSESSMENT_BODY = [
            "we would like you to complete an assessment",
            "please complete the assessment", "kindly complete the assessment",
            "request to complete assessment", "assessment required",
            "assessment mandatory", "assessment as part of selection process",
            "next step assessment", "next stage assessment",
            "shortlisting assessment", "qualification assessment",
            "eligibility test", "screening test", "initial assessment",
            "first round assessment", "round one assessment", "online evaluation",
            "complete within deadline", "assessment completion confirmation",
            "assessment submission received", "assessment under review",
            "assessment in progress", "assessment evaluation in progress",
            "assessment results pending",
            "complete the following assessment", "take our online assessment",
            "complete this online assessment", "assessment link",
            "complete the assessment",
            "hackerrank.com", "codility.com", "hackerearth.com", "hirevue.com",
            "coding challenge link", "technical test link", "online assessment",
        ]
        INTERVIEW_BODY = [
            "we would like to schedule an interview",
            "we would like to invite you for an interview",
            "we would like to invite you to interview",
            "we would like to proceed with your interview",
            "you have been shortlisted for an interview",
            "you have been selected for an interview",
            "please share your availability for an interview",
            "kindly confirm your availability for the interview",
            "preferred interview slot", "available time slots for interview",
            "proposed interview time", "interview slot",
            "invite you to interview", "schedule an interview with you",
            "interested in scheduling an interview",
            "phone screen with our", "video interview with our",
            "conduct an interview with you",
            "please select an interview time", "book your interview slot",
            "confirm your interview time", "confirm your interview date",
        ]
        OFFER_BODY = [
            "we are pleased to inform you", "we are delighted to inform you",
            "we are happy to inform you", "we are excited to inform you",
            "please find your offer letter attached",
            "attached is your offer letter",
            "we would like to offer you", "we are pleased to offer you",
            "we would like to extend an offer", "we are extending an offer",
            "we are happy to extend an offer", "we are excited to extend an offer",
            "successful completion of interview",
            "cleared all interview rounds", "cleared the interview process",
            "passed the interview", "passed the assessment",
            "met all selection criteria", "met the hiring bar",
            "recommended for hire", "approved for hire",
            "hiring approval granted", "hiring committee approval",
            "offer approval", "pre-offer approval",
            "compensation approved", "salary approved",
            "role confirmed", "position confirmed",
            "offer subject to background verification",
            "offer subject to verification", "offer subject to references",
            "offer subject to approval", "conditional upon verification",
            "contingent upon checks",
            "please review and accept", "kindly acknowledge acceptance",
            "acceptance required", "please confirm acceptance",
            "sign and return", "e-sign required", "digital signature required",
            "offer of employment", "pleased to extend an offer",
            "we are pleased to offer you", "we are happy to offer you",
            "employment offer", "formal offer letter", "offer letter is attached",
            "congratulations on your offer",
        ]

# ── Generic sender domains to skip for company name ──────────────────
        GENERIC_SENDER_DOMAINS = {
            "gmail.com", "outlook.com", "yahoo.com", "hotmail.com", "icloud.com",
            "me.com", "googlemail.com", "live.com", "msn.com", "aol.com",
            "mail.com", "protonmail.com", "zoho.com"
        }
        COMPANY_NOISE_WORDS = {
            "hr", "recruiting", "recruitment", "careers", "career", "talent",
            "team", "jobs", "hiring", "notifications", "no-reply", "noreply",
            "donotreply", "do-not-reply", "support", "info", "contact", "people",
            "resourcing", "operations", "staffing", "acquisition"
        }

# ── Role titles to detect in subject ─────────────────────────────────
        ROLE_TITLES = [
            "software engineer", "data engineer", "data scientist",
            "machine learning engineer", "ml engineer", "devops engineer",
            "cloud engineer", "backend engineer", "frontend engineer",
            "full stack engineer", "fullstack engineer", "mobile engineer",
            "ios engineer", "android engineer", "product manager",
            "project manager", "program manager", "data analyst",
            "business analyst", "systems analyst", "ux designer",
            "ui designer", "product designer", "java developer",
            "python developer", "react developer", "node developer",
            "software developer", "web developer", "full stack developer",
            "frontend developer", "backend developer", "qa engineer",
            "test engineer", "site reliability engineer", "security engineer",
            "network engineer", "database administrator", "solution architect",
            "cloud architect", "technical lead", "tech lead", "scrum master"
        ]

        since = (datetime.now() - timedelta(days=90)).strftime("%d-%b-%Y")
        updated_count = 0

        typ, data = imap.search(None, f'(SINCE "{since}")')
        if typ != "OK" or not data[0]:
            imap.logout()
            return "No emails found in the last 90 days"

        msg_ids = data[0].split()
        print(f"[Email Sync] Checking {len(msg_ids)} emails from last 90 days...")

        for num in reversed(msg_ids):
            typ, msgdata = imap.fetch(num, "(RFC822)")
            if typ != "OK":
                continue
            raw_msg = email.message_from_bytes(msgdata[0][1])

            raw_subject = raw_msg.get("Subject") or ""
            raw_from = raw_msg.get("From") or ""
            subject_clean = re.sub(r'^(re|fwd|fw):\s*', '', raw_subject, flags=re.IGNORECASE).strip().lower()

# ── Classify: subject first (priority: Offer > Interview > Assessment)
            found_status = None
            if any(k in subject_clean for k in OFFER_SUBJ):
                found_status = "Offer"
            elif any(k in subject_clean for k in INTERVIEW_SUBJ):
                found_status = "Interview"
            elif any(k in subject_clean for k in ASSESSMENT_SUBJ):
                found_status = "Assessment"

# ── Fallback: scan full body ───────────────────────────────────────
            if not found_status:
                body_text = ""
                if raw_msg.is_multipart():
                    for part in raw_msg.walk():
                        if part.get_content_type() in ("text/plain", "text/html"):
                            try:
                                body_text += part.get_payload(decode=True).decode("utf-8", "ignore").lower()
                            except:
                                pass
                else:
                    try:
                        body_text = raw_msg.get_payload(decode=True).decode("utf-8", "ignore").lower()
                    except:
                        pass

                if any(k in body_text for k in OFFER_BODY):
                    found_status = "Offer"
                elif any(k in body_text for k in INTERVIEW_BODY):
                    found_status = "Interview"
                elif any(k in body_text for k in ASSESSMENT_BODY):
                    found_status = "Assessment"

            if not found_status:
                continue

# ── Extract company name ──────────────────────────────────────────
            company_name = "Unknown Company"
            from_lc = raw_from.lower()

            # Try From display name: "Google Careers <hr@google.com>"
            if "<" in raw_from:
                display = raw_from.split("<")[0].strip().strip('"').strip("'")
                clean_words = [w for w in display.lower().split() if w not in COMPANY_NOISE_WORDS and len(w) > 1]
                if clean_words:
                    company_name = " ".join(clean_words).title()

            # Fall back to sender domain
            if company_name == "Unknown Company" and "@" in from_lc:
                domain = from_lc.split("@")[-1].split(">")[0].strip()
                if domain not in GENERIC_SENDER_DOMAINS:
                    company_name = domain.split(".")[0].capitalize()

            # Refine from subject patterns: "Interview at Google", "from Siemens"
            for pat in [
                r'(?:at|with|from)\s+([A-Z][A-Za-z0-9\s&.,-]{2,35}?)(?:\s+(?:for|regarding|–|-|\||$))',
                r'([A-Z][A-Za-z0-9\s&]{2,30}?)\s+(?:Interview|Assessment|Offer)\b',
            ]:
                m = re.search(pat, raw_subject)
                if m:
                    candidate_co = m.group(1).strip().rstrip(".,")
                    if len(candidate_co) > 2 and candidate_co.lower() not in {"the", "our", "your", "re", "this", "a"}:
                        company_name = candidate_co
                        break

# ── Extract job role from subject ─────────────────────────────────
            job_role = "Job Application"
            for title in ROLE_TITLES:
                if title in subject_clean:
                    idx = subject_clean.find(title)
                    prefix_slice = subject_clean[max(0, idx - 15):idx]
                    seniority = ""
                    for pref in ["senior ", "sr. ", "lead ", "staff ", "principal ", "junior ", "jr. ", "associate "]:
                        if prefix_slice.strip().endswith(pref.strip()):
                            seniority = pref.strip().title() + " "
                            break
                    job_role = seniority + title.title()
                    break

# ── Parse email received date ─────────────────────────────────────
            msg_date = None
            try:
                date_header = raw_msg.get("Date")
                if date_header:
                    msg_date = email.utils.parsedate_to_datetime(date_header)
            except:
                pass

# ── Deduplicate: same candidate + same subject = skip ─────────────
            existing = EmailEvent.query.filter_by(
                candidate_id=candidate_id,
                subject=raw_subject[:500]
            ).first()
            if existing:
                continue

            new_event = EmailEvent(
                candidate_id=candidate_id,
                event_type=found_status,
                company_name=company_name[:250],
                job_role=job_role[:250],
                subject=raw_subject[:500],
                received_at=msg_date
            )
            db.session.add(new_event)
            updated_count += 1
            print(f"[Email Sync] [{found_status}] {company_name} | {raw_subject[:60]}")

        db.session.commit()
        imap.logout()
        return f"Sync complete. Found {updated_count} new milestone(s)."

    except Exception as e:
        db.session.rollback()
        print(f"[Email Sync] Error for candidate {candidate_id}: {e}")
        if imap:
            try:
                imap.logout()
            except:
                pass
        return f"Sync failed: {str(e)}"

@app.get("/api/candidates/all-job-applications")
def api_get_all_job_applications():
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 401
    
    print("[API] Fetching all job applications (Outer Join)...")
    # Use outerjoin so orphaned apps (e.g. from batch tool with wrong IDs) still show up
    results = db.session.query(JobApplication, Candidate.name).outerjoin(Candidate, JobApplication.candidate_id == Candidate.id).order_by(JobApplication.created_at.desc()).all()
    
    print(f"[API] Found {len(results)} records.")
    data = []
    for app_obj, cand_name in results:
        d = app_obj.to_dict()
        d["candidate_name"] = cand_name or f"Unknown (ID:{app_obj.candidate_id})"
        data.append(d)
        
    return jsonify(data)

@app.delete("/api/job-applications/<int:app_id>")
def api_delete_job_application(app_id):
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 401
    app_obj = db.session.get(JobApplication, app_id)
    if not app_obj:
        return jsonify({"error": "Application not found"}), 404
    db.session.delete(app_obj)
    db.session.commit()
    return jsonify({"message": "Application deleted successfully."})

@app.post("/api/job-applications/delete-all")
def api_delete_all_job_applications():
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 401
    # Allow targetting orphans only or everything
    mode = request.json.get("mode", "all")
    if mode == "orphans":
        # Delete apps where candidate_id does not exist in Candidate table
        # Subquery for existing candidate IDs
        valid_ids = db.session.query(Candidate.id)
        db.session.query(JobApplication).filter(JobApplication.candidate_id.not_in(valid_ids)).delete(synchronize_session=False)
    else:
        db.session.query(JobApplication).delete()
    
    db.session.commit()
    return jsonify({"message": f"Successfully deleted {mode} applications."})


def bulk_sync_worker(app_context, cands):
    global is_bulk_sync_running, last_bulk_sync_status
    with app_context:
        is_bulk_sync_running = True
        total_updated = 0
        total_candidates = len(cands)
        print(f"[Bulk Sync] Starting background sync for {total_candidates} candidates.")
        
        for i, c in enumerate(cands, 1):
            try:
                print(f"[Bulk Sync] [{i}/{total_candidates}] Processing {c.name}...")
                res = sync_candidate_applications(c.id)
                print(f"[Bulk Sync]   -> {res}")
                if "Updated" in res:
                    match = re.search(r"Updated (\d+)", res)
                    if match:
                        total_updated += int(match.group(1))
            except Exception as e:
                db.session.rollback()
                print(f"[Bulk Sync] Error for {c.name}: {e}")
                
        is_bulk_sync_running = False
        last_bulk_sync_status = f"Completed at {datetime.now().strftime('%H:%M:%S')}. Updated {total_updated} total."
        print(f"[Bulk Sync] {last_bulk_sync_status}")

@app.get("/api/candidates/sync-status")
def api_get_sync_status():
    global is_bulk_sync_running, last_bulk_sync_status
    return jsonify({
        "running": is_bulk_sync_running,
        "message": last_bulk_sync_status
    })

@app.post("/api/candidates/sync-all")
def api_sync_all_candidates_apps():
    if "user_id" not in session:
        return jsonify({"error": "Unauthorized"}), 401
    
    cands = Candidate.query.filter(Candidate.app_password != None, Candidate.app_password != "").all()
    if not cands:
        return jsonify({"message": "No candidates with app passwords found."}), 200

    # Start bulk sync in a background thread to prevent timeout
    thread = Thread(target=bulk_sync_worker, args=(app.app_context(), cands))
    thread.daemon = True
    thread.start()
            
    return jsonify({
        "message": f"Bulk sync started for {len(cands)} candidates. This will run in the background. Check server logs for details."
    })


# ---- Workflow Plan API Routes ----

@app.get("/api/workflow-plans")
def api_list_workflow_plans():
    plans = WorkflowPlan.query.order_by(WorkflowPlan.id.desc()).all()
    return jsonify([p.to_dict() for p in plans])


@app.get("/api/dashboard/workflow-phase-tracker")
def api_dashboard_workflow_phase_tracker():
    """Dashboard: every candidate; latest workflow plan per candidate when present, else 'no plan' stub row."""
    cands = Candidate.query.order_by(Candidate.name.asc(), Candidate.id.asc()).all()
    plans_all = WorkflowPlan.query.order_by(WorkflowPlan.id.desc()).all()
    latest_by_cand = {}
    for p in plans_all:
        cid = int(p.candidate_id)
        if cid not in latest_by_cand:
            latest_by_cand[cid] = p
    rows = []
    for c in cands:
        p = latest_by_cand.get(int(c.id))
        if p:
            row = workflow_plan_phase_tracker_rows(p)
            row["candidate_name"] = (c.name or "").strip() or row.get("candidate_name") or ""
            rows.append(row)
        else:
            rows.append(workflow_tracker_row_no_plan(c))
    return jsonify({"plans": rows})

@app.get("/api/workflow-plans/<int:plan_id>")
def api_get_workflow_plan(plan_id):
    plan = WorkflowPlan.query.get_or_404(plan_id)
    data = plan.to_dict()
    # Add per-phase breakdown
    phases = []
    for pnum, pinfo in WORKFLOW_PHASES.items():
        phases.append({
            "phase": pnum,
            "label": pinfo["label"],
            "days_range": f"{pinfo['days'][0]}-{pinfo['days'][1]}",
            "target": pinfo["cumulative"],
            "per_batch": pinfo["per_batch"],
            "alternate": pinfo["alternate"],
            "is_current": pnum == data["current_phase"],
            "is_completed": data["total_applied"] >= pinfo["cumulative"],
        })
    data["phases"] = phases
    return jsonify(data)

@app.post("/api/workflow-plans")
def api_create_workflow_plan():
    data = request.json or {}
    candidate_id = data.get("candidate_id")
    if not candidate_id:
        return jsonify({"error": "candidate_id is required"}), 400

    cand = db.session.get(Candidate, int(candidate_id))
    if not cand:
        return jsonify({"error": "Candidate not found"}), 404

    # Check for existing active plan
    existing = WorkflowPlan.query.filter_by(candidate_id=int(candidate_id), status="active").first()
    if existing:
        return jsonify({"error": f"Candidate already has an active workflow plan (Plan #{existing.id})"}), 400

    # Parse manual inputs for transition
    initial_days = int(data.get("initial_days", 0))
    initial_applied = int(data.get("initial_applied", 0))
    total_target = int(data.get("total_target", 1200))
    country = data.get("country", "")
    industry = data.get("industry", "")
    start_datetime_str = data.get("scheduled_start_time", "")

    now = datetime.utcnow()
    
    if start_datetime_str:
        try:
            # Frontend passes ISO string e.g. "2026-10-15T09:30:00.000Z"
            next_run_date = datetime.fromisoformat(start_datetime_str.replace("Z", "+00:00")).astimezone(timezone.utc).replace(tzinfo=None)
        except Exception as e:
            print(f"Error parsing start_datetime: {e}")
            next_run_date = now
    else:
        next_run_date = now

    # Calculate backdated start date if initial_days provided
    now = datetime.utcnow()
    service_start_date = now - timedelta(days=initial_days)
    service_end_date = service_start_date + timedelta(days=180)

    # Initial phase calculation
    temp_plan = WorkflowPlan(service_start_date=service_start_date)
    initial_phase = temp_plan._get_phase_for_day(initial_days)

    plan = WorkflowPlan(
        candidate_id=cand.id,
        candidate_name=cand.name,
        enrollment_id=cand.enrollment_id or "",
        service_start_date=service_start_date,
        service_end_date=service_end_date,
        status="active",
        current_phase=initial_phase,
        total_target=total_target,
        total_applied=initial_applied,
        country=country,
        industry=industry,
        scheduled_start_time=next_run_date,
        is_alternate_day=True,
        next_run_date=next_run_date, # Start automation precisely at scheduled time
        reported_service_day=initial_days,
        reported_service_anchor_at=now,
    )
    db.session.add(plan)
    db.session.commit()
    return jsonify({"message": "Workflow plan initialized", "plan": plan.to_dict()}), 201

@app.put("/api/workflow-plans/<int:plan_id>")
def api_update_workflow_plan(plan_id):
    plan = WorkflowPlan.query.get_or_404(plan_id)
    data = request.json or {}

    action = data.get("action", "").lower()
    if action == "pause":
        plan.status = "paused"
    elif action == "resume":
        plan.status = "active"
        if not plan.next_run_date or plan.next_run_date < datetime.utcnow():
            plan.next_run_date = datetime.utcnow()
    elif action == "reset":
        plan.total_applied = 0
        plan.current_phase = 1
        plan.status = "active"
        plan.last_run_date = None
        plan.next_run_date = plan.scheduled_start_time or plan.service_start_date
    elif action == "send_now":
        batch_size = workflow_plan_effective_send_batch(plan, datetime.utcnow(), scheduled=False)
        if batch_size > 0:
            auto_start_single_run(plan.candidate_id, limit=batch_size, offset=0, industry=plan.industry, country=plan.country)
            plan.last_run_date = datetime.utcnow()
            db.session.commit()
            return jsonify({"message": f"Sent {batch_size} applications immediately."})
        else:
            appd = int(plan.total_applied or 0)
            tgt = int(plan.total_target or 1200)
            if appd >= tgt:
                msg = "Total target already reached."
            elif remaining_daily_send_quota(plan.candidate_id) <= 0:
                msg = "Daily automation limit already reached for this candidate today."
            else:
                msg = "Nothing to send in this slot."
            return jsonify({"message": msg}), 400
    elif action == "reschedule":
        # Band setup / dashboard: set next run time and optionally sync service curve fields (existing plan).
        start_datetime_str = data.get("scheduled_start_time", "") or ""
        now = datetime.utcnow()
        if start_datetime_str:
            try:
                next_run_date = datetime.fromisoformat(start_datetime_str.replace("Z", "+00:00")).astimezone(
                    timezone.utc
                ).replace(tzinfo=None)
            except Exception:
                next_run_date = now
        else:
            next_run_date = now
        plan.next_run_date = next_run_date
        plan.scheduled_start_time = next_run_date
        plan.status = "active"
        if "initial_days" in data:
            days = max(0, int(data["initial_days"]))
            plan.service_start_date = now - timedelta(days=days)
            plan.service_end_date = plan.service_start_date + timedelta(days=180)
            plan.reported_service_day = days
            plan.reported_service_anchor_at = now
            plan.current_phase = plan._get_phase_for_day(days)
        if "initial_applied" in data:
            plan.total_applied = max(0, int(data["initial_applied"]))
        if "total_target" in data:
            plan.total_target = max(1, int(data["total_target"]))
        if "country" in data:
            plan.country = (data.get("country") or "").strip()
        if "industry" in data:
            plan.industry = (data.get("industry") or "").strip()
    elif action == "adjust":
        # Recalculate service_start_date based on current days
        if "initial_days" in data:
            days = int(data["initial_days"])
            now_adj = datetime.utcnow()
            plan.service_start_date = now_adj - timedelta(days=days)
            plan.reported_service_day = days
            plan.reported_service_anchor_at = now_adj
        if "total_applied" in data:
            plan.total_applied = int(data["total_applied"])
        if "total_target" in data:
            plan.total_target = int(data["total_target"])
        # Re-evaluate phase after adjustment
        plan.current_phase = plan._get_phase_for_day((datetime.utcnow() - plan.service_start_date).days)
    else:
        # Field updates
        if "total_target" in data:
            plan.total_target = int(data["total_target"])
        if "total_applied" in data:
            plan.total_applied = int(data["total_applied"])
        if "status" in data:
            plan.status = data["status"]

    db.session.commit()
    return jsonify({"message": "Updated", "plan": plan.to_dict()})

@app.delete("/api/workflow-plans/<int:plan_id>")
def api_delete_workflow_plan(plan_id):
    plan = WorkflowPlan.query.get_or_404(plan_id)
    db.session.delete(plan)
    db.session.commit()
    return jsonify({"message": "Workflow plan deleted"})

@app.get("/api/workflow-plans/<int:plan_id>/status")
def api_workflow_plan_status(plan_id):
    plan = WorkflowPlan.query.get_or_404(plan_id)
    status_info = plan.compute_status()
    now = datetime.utcnow()
    nr_raw = plan.next_run_date
    nr_show = workflow_next_run_effective_display(nr_raw, now, plan.status or "active")
    status_info["id"] = plan.id
    status_info["candidate_name"] = plan.candidate_name or ""
    status_info["total_applied"] = plan.total_applied or 0
    status_info["total_target"] = plan.total_target or 1200
    status_info["status"] = plan.status or "active"
    status_info["next_run_date"] = nr_show.isoformat() if nr_show else ""
    status_info["next_run_date_stored"] = nr_raw.isoformat() if nr_raw else ""
    status_info["last_run_date"] = plan.last_run_date.isoformat() if plan.last_run_date else ""
    return jsonify(status_info)

@app.get("/api/workflow-phases")
def api_workflow_phases():
    """Return the phase definitions for the frontend."""
    phases = []
    for pnum, pinfo in WORKFLOW_PHASES.items():
        phases.append({
            "phase": pnum,
            "label": pinfo["label"],
            "days_range": f"{pinfo['days'][0]}-{pinfo['days'][1]}",
            "target": pinfo["target"],
            "cumulative": pinfo["cumulative"],
            "per_batch": pinfo["per_batch"],
            "alternate": pinfo["alternate"],
        })
    return jsonify(phases)


@app.post("/api/smart-automation/preview", strict_slashes=False)
@app.post("/api/sa/preview", strict_slashes=False)
def api_smart_automation_preview():
    """Preview 6-month–aligned batch size and phase for Smart Automation (no DB changes)."""
    data = request.json or {}
    days = int(data.get("days_in_system") or data.get("days", 0) or 0)
    applied = int(data.get("applied", 0) or 0)
    max_per_run = int(data.get("max_per_run", 100) or 100)
    interval_days = int(data.get("interval_days", 1) or 1)
    cand_id = None
    cid_raw = data.get("candidate_id")
    if cid_raw not in (None, ""):
        try:
            cand_id = int(cid_raw)
        except (TypeError, ValueError):
            cand_id = None
    plan = compute_smart_automation_plan(
        days, applied, max_per_run=max_per_run, interval_days=interval_days, candidate_id=cand_id
    )
    return jsonify(plan)


@app.get("/api/smart-automation/dashboard", strict_slashes=False)
@app.get("/api/sa/dashboard", strict_slashes=False)
def api_smart_automation_dashboard():
    """Active Smart workspaces with live days, bucket, PO, industry, country, pending."""
    now = datetime.utcnow()
    rows = []
    try:
        for ws in Workspace.query.filter_by(automation_enabled=True).order_by(Workspace.id.desc()).all():
            cand = db.session.get(Candidate, ws.candidate_id)
            if not cand:
                continue
            svc = getattr(cand, "smart_service_start_date", None) or ws.service_start_date
            days_elapsed = max(0, (now - svc).days) if svc else 0
            wp = (
                WorkflowPlan.query.filter_by(candidate_id=cand.id)
                .filter(WorkflowPlan.status.in_(["paused", "active"]))
                .order_by(WorkflowPlan.id.desc())
                .first()
            )
            applied = int(wp.total_applied or 0) if wp else int(ws.automation_total_sent or 0)
            cap = int(ws.automation_per_run_cap or 100)
            interval = max(1, int(ws.automation_interval_days or 1))
            plan = compute_smart_automation_plan(
                days_elapsed, applied, max_per_run=cap, interval_days=interval, candidate_id=cand.id
            )
            tgt = int(wp.total_target or 1200) if wp else 1200
            wf_pred = workflow_phases_prediction_payload(
                days_elapsed, applied, tgt, service_start=svc
            )
            rows.append({
                "workspace_id": ws.id,
                "candidate_id": cand.id,
                "candidate_name": cand.name,
                "po_name": (cand.placement_officer_member or "").strip(),
                "enrollment_id": (cand.enrollment_id or "").strip(),
                "total_applied": applied,
                "days_in_system": days_elapsed,
                "days_bucket": plan.get("days_bucket") or days_bucket_short(days_elapsed),
                "phase_label": plan.get("phase_label") or "",
                "industry": ws.industry or "Default",
                "country": ws.country or "Global",
                "pending_on_track": plan.get("pending_on_track", 0),
                "next_batch": plan.get("suggested_batch_per_run", 0),
                "backlog_mode": bool(plan.get("backlog_mode")),
                "backlog_safe_cap": int(plan.get("backlog_safe_cap") or SMART_AUTOMATION_BACKLOG_SAFE_CAP),
                "remaining_to_cap": plan.get("remaining_to_cap", 0),
                "expected_applications_by_now": int(plan.get("expected_applications_by_now") or 0),
                "phase_batch_cap": int(plan.get("phase_batch_cap") or 0),
                "cadence_days": interval,
                "next_run_utc": ws.automation_next_run.isoformat() if ws.automation_next_run else "",
                "service_start_utc": svc.isoformat() if svc else "",
                "workflow_plan_status": (wp.status or "") if wp else "",
                "total_target": tgt,
                "prediction_pct": wf_pred["overall_progress_pct"],
                "workflow_phases": wf_pred["phases"],
                "workflow_current_phase": wf_pred["current_phase"],
                "sent_today_for_daily_cap": int(plan.get("sent_today_for_daily_cap") or 0),
                "remaining_daily_quota": int(plan.get("remaining_daily_quota") or 0),
            })
    except Exception as ex:
        print(f"[Smart Automation] dashboard error: {ex}")
        return jsonify({"error": "Smart dashboard failed", "detail": str(ex)}), 500
    return jsonify({"rows": rows})


@app.post("/api/candidates/<int:candidate_id>/smart-service")
def api_save_candidate_smart_service(candidate_id):
    """One-time service profile: start date, total applications till date (incl. manual), HR country/industry."""
    data = request.json or {}
    cand = Candidate.query.get_or_404(candidate_id)
    raw_start = (data.get("smart_service_start_date") or data.get("service_start_date") or "").strip()
    if raw_start:
        try:
            if "T" in raw_start:
                cand.smart_service_start_date = datetime.fromisoformat(raw_start.replace("Z", "+00:00"))
                if cand.smart_service_start_date.tzinfo is not None:
                    cand.smart_service_start_date = cand.smart_service_start_date.astimezone(timezone.utc).replace(tzinfo=None)
            else:
                cand.smart_service_start_date = datetime.strptime(raw_start[:10], "%Y-%m-%d")
        except Exception:
            return jsonify({"error": "Invalid smart_service_start_date"}), 400
    if "smart_baseline_applied" in data:
        cand.smart_baseline_applied = max(0, int(data.get("smart_baseline_applied") or 0))
    if "smart_country" in data:
        cand.smart_country = (data.get("smart_country") or "").strip() or None
    if "smart_industry" in data:
        cand.smart_industry = (data.get("smart_industry") or "").strip() or None
    if "days_in_system" in data:
        d = max(0, int(data.get("days_in_system") or 0))
        cand.smart_service_start_date = datetime.utcnow() - timedelta(days=d)
    if "smart_country" in data or "smart_industry" in data:
        sync_workflow_plan_hr_from_candidate(cand)
    db.session.commit()
    wf_s, ws_s = candidate_aux_service_starts(cand.id)
    return jsonify(
        {
            "message": "Saved",
            "candidate": cand.to_dict_summary(
                workflow_service_start=wf_s,
                workspace_service_start=ws_s,
            ),
        }
    ), 200


@app.post("/api/smart-automation/apply", strict_slashes=False)
@app.post("/api/sa/apply", strict_slashes=False)
def api_smart_automation_apply():
    """Enable Campaign workspace automation: recomputed batch each run from service dates + applied count."""
    data = request.json or {}
    candidate_ids = data.get("candidate_ids") or []
    if not candidate_ids:
        return jsonify({"error": "candidate_ids required"}), 400

    # Prefer days_in_system when present so 0 is not treated as falsy and skipped.
    if "days_in_system" in data:
        try:
            days_input = max(0, int(data.get("days_in_system") or 0))
        except (TypeError, ValueError):
            days_input = 0
    else:
        try:
            days_input = max(0, int(data.get("days") or 0))
        except (TypeError, ValueError):
            days_input = 0
    applied = int(data.get("applied", 0) or 0)
    country = (data.get("country") or "Global").strip()
    industry = (data.get("industry") or "Default").strip()
    max_per_run = int(data.get("max_per_run", 100) or 100)
    interval_days = max(1, min(7, int(data.get("interval_days", 1) or 1)))
    next_run_iso = (data.get("next_run_iso") or "").strip()
    service_start_iso = (data.get("service_start_iso") or "").strip()

    if next_run_iso:
        try:
            dt = datetime.fromisoformat(next_run_iso.replace("Z", "+00:00"))
            if dt.tzinfo is not None:
                dt = dt.astimezone(timezone.utc).replace(tzinfo=None)
            else:
                dt = dt.replace(tzinfo=None)
        except Exception:
            dt = datetime.utcnow() + timedelta(minutes=2)
    else:
        dt = datetime.utcnow() + timedelta(minutes=2)

    results = []
    last_plan = None
    now = datetime.utcnow()

    for cid_raw in candidate_ids:
        try:
            cid = int(cid_raw)
        except Exception:
            results.append({"candidate_id": cid_raw, "error": "invalid id"})
            continue
        cand = db.session.get(Candidate, cid)
        if not cand:
            results.append({"candidate_id": cid, "error": "not found"})
            continue

        svc_start = None
        if service_start_iso:
            try:
                svc_start = datetime.fromisoformat(service_start_iso.replace("Z", "+00:00"))
                if svc_start.tzinfo is not None:
                    svc_start = svc_start.astimezone(timezone.utc).replace(tzinfo=None)
                else:
                    svc_start = svc_start.replace(tzinfo=None)
            except Exception:
                svc_start = None
        if svc_start is None:
            # days_in_system: service "day 1" was N days ago → elapsed days increase automatically each day
            svc_start = now - timedelta(days=max(0, days_input))
        cand.smart_service_start_date = svc_start
        # Do not overwrite candidate profile Country Type / industry — those are edited on the candidate form.
        # Campaign country & industry live on Workspace + WorkflowPlan (ws.country, wp.country).
        days_i = max(0, (now - svc_start).days)
        # `applied` from client = total job applications completed till now (manual + CRM); CRM adds more after this via sends.
        applied_i = max(0, int(applied))
        cand.smart_baseline_applied = applied_i
        plan_i = compute_smart_automation_plan(
            days_i, applied_i, max_per_run=max_per_run, interval_days=interval_days, candidate_id=cid
        )
        batch_i = int(plan_i["suggested_batch_per_run"])
        rem_i = int(plan_i["remaining_to_cap"])
        if batch_i <= 0 or rem_i <= 0:
            results.append({"candidate_id": cid, "error": "Nothing to schedule for this candidate at current numbers."})
            continue

        ws = Workspace.query.filter_by(candidate_id=cid, industry=industry, country=country).first()
        if not ws:
            ws = Workspace(
                name=f"Smart Auto — {cand.name}",
                candidate_id=cid,
                industry=industry,
                country=country,
            )
            db.session.add(ws)
            db.session.flush()

        ws.automation_enabled = True
        ws.automation_batch_size = batch_i
        ws.automation_interval_days = interval_days
        ws.automation_per_run_cap = max_per_run
        ws.automation_max_emails = rem_i
        ws.automation_next_run = dt
        ws.automation_type = "interval"
        ws.industry = industry
        ws.country = country
        ws.service_start_date = svc_start
        ws.service_end_date = svc_start + timedelta(days=180)

        wp = WorkflowPlan.query.filter_by(candidate_id=cid, status="active").first()
        if wp:
            wp.status = "paused"
        else:
            wp = WorkflowPlan.query.filter_by(candidate_id=cid, status="paused").first()

        if not wp:
            wp = WorkflowPlan(
                candidate_id=cid,
                candidate_name=cand.name,
                enrollment_id=cand.enrollment_id,
                service_start_date=svc_start,
                service_end_date=svc_start + timedelta(days=180),
                status="paused",
                total_target=1200,
                total_applied=applied_i,
                country=country,
                industry=industry,
            )
            db.session.add(wp)
        else:
            wp.total_applied = applied_i
            wp.total_target = 1200
            wp.service_start_date = svc_start
            wp.service_end_date = svc_start + timedelta(days=180)
            wp.industry = industry
            wp.country = country
            wp.candidate_name = cand.name

        last_plan = plan_i
        results.append({"candidate_id": cid, "workspace_id": ws.id, "ok": True, "plan": plan_i})

    db.session.commit()
    if not any(r.get("ok") for r in results):
        return jsonify({
            "error": "No candidate could be enabled (check service dates, applied counts, and on-track target).",
            "results": results,
        }), 400
    return jsonify({
        "message": "Smart automation enabled for campaign workspaces.",
        "plan": last_plan,
        "results": results,
    }), 200


if __name__ == "__main__":
    with app.app_context():
        db.create_all()
        # Ensure at least one admin exists
        admin = User.query.filter_by(role='admin').first()
        if not admin:
            from werkzeug.security import generate_password_hash
            print("[Startup] No admin found. Creating default admin...")
            admin = User(
                username='admin',
                email='admin@terratern.com',
                password_hash=generate_password_hash('admin123'),
                is_verified=True,
                is_approved=True,
                role='admin'
            )
            db.session.add(admin)
            db.session.commit()
            print("[Startup] Default admin 'admin' created with password 'admin123'")
        
        # Enable WAL mode for SQLite to handle multi-threaded access
        try:
            db.session.execute(text("PRAGMA journal_mode=WAL;"))
            db.session.commit()
        except Exception as e:
            print(f"[Startup] Error enabling WAL: {e}")

    
    debug = False
    
    sched_t = Thread(target=run_scheduler_loop, daemon=True)
    sched_t.start()
    
    import socket
    hostname = socket.gethostname()
    local_ip = socket.gethostbyname(hostname)
    print(f"\n[Server] Access CRM locally: http://127.0.0.1:8080")
    print(f"[Server] Access CRM on WiFi: http://{local_ip}:8080")
    print("[Server] SQLite DB backup: GET /api/candidates/backup_sqlite (also sqlite_backup, /api/backup/database)")
    print("[Server] CRM manifest CSV: POST/GET /api/reports/crm-manifest | /api/runs?download=crm_manifest_csv | /api/crm-manifest-export\n")

    port = int(os.environ.get("PORT", 8080))
    app.run(host="0.0.0.0", port=port, debug=debug, threaded=True)
